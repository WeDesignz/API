from rest_framework import status, filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.db.models import Q, Count, F, Sum
from django.utils import timezone
from django.conf import settings
from datetime import timedelta
from django.core.paginator import Paginator
from django.core.files.storage import default_storage
import random
import os
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from .models import Product, Category, CollectionBundle, Tags, ProductCounter, PDFDownload
from .serializers import (
    ProductSerializer, CategorySerializer, CategoryListSerializer, CollectionBundleSerializer,
    TagsSerializer, ProductCounterSerializer, PDFDownloadSerializer,
    PDFDownloadListSerializer, PDFDownloadCreateSerializer, PDFDownloadRequestSerializer,
    PDFDownloadStatusSerializer, PDFDownloadPaymentSerializer
)
from common.business_config import BusinessConfig
from .tasks import generate_pdf_task


class CustomPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


@api_view(['GET'])
@permission_classes([AllowAny])
def landing_page(request):
    """
    Landing page for unauthenticated users - shows overview and plans.
    """
    try:
        # Get active plans
        from Plans.models import Plan
        plans = Plan.objects.filter(status='active').order_by('plan_name', 'plan_duration')
        
        # Get featured categories
        featured_categories = Category.objects.annotate(
            product_count=Count('products', filter=Q(products__status='active'))
        ).filter(product_count__gt=0).order_by('-product_count')[:6]
        
        # Get some sample products for preview
        sample_products = Product.objects.filter(
            status='active',
            visibility_status='show'
        ).order_by('-created_at')[:8]
        
        try:
            categories_data = CategorySerializer(featured_categories, many=True).data
        except Exception:
            categories_data = []
        
        try:
            products_data = ProductSerializer(sample_products, many=True, context={'request': request}).data
        except Exception:
            products_data = []
        
        return Response({
            'message': 'Welcome to WeDesignz',
            'featured_categories': categories_data,
            'sample_products': products_data,
            'plans': [{
                'id': plan.id,
                'plan_name': plan.get_plan_name_display(),
                'duration': plan.get_plan_duration_display(),
                'price': float(plan.price),
                'description': plan.description
            } for plan in plans]
        })
    except Exception as e:
        return Response({
            'message': 'Welcome to WeDesignz',
            'featured_categories': [],
            'sample_products': [],
            'plans': [],
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])
def home_feed(request):
    """
    Dynamic home page feed with designs and bundles.
    Mix of individual designs and bundles (max 20% bundles).
    """
    try:
        page = int(request.GET.get('page', 1))
        page_size = 20
        
        # Calculate offset for pagination
        offset = (page - 1) * page_size
        
        # Get all active products
        products = Product.objects.filter(
            status='active',
            visibility_status='show'
        ).order_by('-created_at')
        
        # Get all active bundles
        bundles = CollectionBundle.objects.filter(status='available')
        bundles_count = bundles.count()
        
        # Calculate how many bundles to include (max 20% of page_size)
        max_bundles = min(bundles_count, int(page_size * 0.2))
        
        # Randomly select bundles (handle case where max_bundles > len(bundles))
        if bundles.exists() and max_bundles > 0:
            bundles_list = list(bundles)
            if len(bundles_list) >= max_bundles:
                selected_bundles = random.sample(bundles_list, max_bundles)
            else:
                selected_bundles = bundles_list
        else:
            selected_bundles = []
        
        # Get remaining slots for products
        remaining_slots = page_size - len(selected_bundles)
        
        # Get products for remaining slots with pagination
        total_products = products.count()
        products_for_page = list(products[offset:offset + remaining_slots])
        
        # Mix bundles and products randomly
        all_items = products_for_page + selected_bundles
        random.shuffle(all_items)
        
        # Separate back into products and bundles for serialization
        feed_products = [item for item in all_items if isinstance(item, Product)]
        feed_bundles = [item for item in all_items if isinstance(item, CollectionBundle)]
        
        # Check if there are more items
        total_items_so_far = offset + len(all_items)
        has_next = total_items_so_far < (total_products + bundles_count)
        
        try:
            products_data = ProductSerializer(feed_products, many=True, context={'request': request}).data
        except Exception as e:
            # If serialization fails, return empty list
            products_data = []
        
        try:
            bundles_data = CollectionBundleSerializer(feed_bundles, many=True).data
        except Exception as e:
            # If serialization fails, return empty list
            bundles_data = []
        
        return Response({
            'page': page,
            'products': products_data,
            'bundles': bundles_data,
            'has_next': has_next
        })
    except Exception as e:
        # Return error response instead of 500
        return Response({
            'page': int(request.GET.get('page', 1)),
            'products': [],
            'bundles': [],
            'has_next': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Trending Designs',
    operation_description='Trending Designs endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([AllowAny])
def trending_designs(request):
    """
    Top trending designs based on downloads/purchases in last 7 days.
    """
    seven_days_ago = timezone.now() - timedelta(days=7)
    
    trending_products = Product.objects.filter(
        status='active',
        visibility_status='show',
        productcounters__product_counter_type__in=['downloaded', 'purchased'],
        productcounters__created_at__gte=seven_days_ago
    ).annotate(
        trend_score=Count('productcounters', filter=Q(
            productcounters__product_counter_type__in=['downloaded', 'purchased'],
            productcounters__created_at__gte=seven_days_ago
        ))
    ).order_by('-trend_score')[:20]
    
    return Response({
        'trending_designs': ProductSerializer(trending_products, many=True, context={'request': request}).data
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Recently Added',
    operation_description='Recently Added endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([AllowAny])
def recently_added(request):
    """
    Recently added designs.
    """
    recent_products = Product.objects.filter(
        status='active',
        visibility_status='show'
    ).order_by('-created_at')[:20]
    
    return Response({
        'recently_added': ProductSerializer(recent_products, many=True, context={'request': request}).data
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def hero_section_designs(request):
    """
    Get designs for hero section CardSwap.
    Returns full product data with image, title, creator, price.
    """
    from CoreAdmin.models import SystemConfig
    from django.core.cache import cache
    
    try:
        config = SystemConfig.get_config()
        product_ids_raw = config.hero_section_designs or []
        
        if not product_ids_raw:
            return Response({
                'designs': [],
                'message': 'No designs configured for hero section'
            })
        
        # Convert string IDs to integers (handle both string and int formats)
        product_ids = []
        for pid in product_ids_raw:
            try:
                product_ids.append(int(pid))
            except (ValueError, TypeError):
                continue
        
        if not product_ids:
            return Response({
                'designs': [],
                'message': 'No valid product IDs found in hero section configuration'
            })
        
        # Cache key
        cache_key = f'hero_section_designs_{config.updated_at.timestamp()}'
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)
        
        # Fetch products with optimized query
        products = Product.objects.filter(
            id__in=product_ids,
            status='active',
            visibility_status='show'
        ).select_related('category', 'created_by').only(
            'id', 'title', 'price', 'created_by', 'category', 'product_number'
        )
        
        # Maintain order from config
        product_dict = {p.id: p for p in products}
        ordered_products = [product_dict[pid] for pid in product_ids if pid in product_dict]
        
        # Log if some products weren't found
        import logging
        import os
        from MediaFiles.models import Relation
        
        logger = logging.getLogger(__name__)
        
        if len(ordered_products) < len(product_ids):
            missing_ids = set(product_ids) - {p.id for p in ordered_products}
            logger.warning(f'Hero section: {len(missing_ids)} products not found or not active: {missing_ids}')
        
        logger.info(f'Hero section: Found {len(ordered_products)} products out of {len(product_ids)} requested')
        
        designs = []
        
        for product in ordered_products:
            # Get media files - same logic as dome gallery (prefer mockup, fallback to any image)
            image_url = None
            mockup_media = None
            fallback_media = None
            
            try:
                media = product.get_media()
            except Exception as e:
                logger.error(f'Error getting media for product {product.id}: {e}')
                media = None
            
            if media:
                try:
                    media_list = list(media)
                except Exception as e:
                    logger.error(f'Error converting media to list for product {product.id}: {e}')
                    media_list = []
                
                logger.info(f'Product {product.id} has {len(media_list)} media files')
                
                # Prefer design images (JPG/PNG, non-mockup), then fallback to mockup
                for m in media_list:
                    if not m.file:
                        continue
                    
                    # Only process image media files (skip vectors, documents, etc.)
                    media_type = getattr(m, 'media_type', 'image')
                    if media_type and media_type.lower() != 'image':
                        continue
                    
                    # Check if file extension is an image format
                    file_name = m.file.name if hasattr(m.file, 'name') else ''
                    is_image_file = False
                    if file_name:
                        file_name_lower = file_name.lower()
                        is_image_file = any(ext in file_name_lower for ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp'])
                    
                    # Skip if not an image file
                    if not is_image_file:
                        continue
                    
                    # Check filename for mockup
                    is_mockup_by_name = False
                    if file_name:
                        file_name_lower = file_name.lower()
                        base_name = os.path.splitext(os.path.basename(file_name_lower))[0]
                        is_mockup_by_name = base_name == 'mockup'
                    
                    # Check Relation metadata for mockup
                    is_mockup_by_meta = False
                    try:
                        from MediaFiles.models import Relation
                        relation = Relation.objects.filter(
                            relation_type='Product:Media',
                            id_1=product.pk,
                            id_2=m.pk
                        ).first()
                        
                        if relation and relation.meta:
                            meta = relation.meta
                            if isinstance(meta, dict):
                                is_mockup_by_meta = meta.get('is_mockup', False) or meta.get('type') == 'mockup'
                            elif isinstance(meta, str):
                                meta_lower = str(meta).lower()
                                is_mockup_by_meta = 'mockup' in meta_lower or '"is_mockup":true' in meta_lower
                    except Exception:
                        pass
                    
                    is_mockup = is_mockup_by_name or is_mockup_by_meta
                    
                    # Check if it's JPG or PNG (design image)
                    is_jpg_png = False
                    if file_name:
                        file_name_lower = file_name.lower()
                        is_jpg_png = file_name_lower.endswith(('.jpg', '.jpeg', '.png'))
                    
                    # Prefer design images (JPG/PNG, non-mockup)
                    if is_jpg_png and not is_mockup:
                        try:
                            test_url = m.file.url
                            selected_media = m
                            logger.info(f'Product {product.id}: Found design image {m.pk}')
                            break
                        except:
                            pass
                    # Store mockup as fallback
                    elif is_mockup and not mockup_media:
                        try:
                            test_url = m.file.url
                            mockup_media = m
                            logger.debug(f'Product {product.id}: Storing mockup media {m.pk} as fallback')
                        except:
                            pass
                    # Store first valid image as last fallback
                    elif not fallback_media:
                        try:
                            test_url = m.file.url
                            fallback_media = m
                            logger.debug(f'Product {product.id}: Storing image media {m.pk} as last fallback')
                        except:
                            pass
                
                # Use design image if found, otherwise mockup, otherwise fallback
                if 'selected_media' not in locals():
                    selected_media = mockup_media or fallback_media
                
                if selected_media and selected_media.file:
                    try:
                        url = selected_media.file.url
                        if request and url.startswith('/'):
                            url = request.build_absolute_uri(url)
                        image_url = url
                        logger.info(f'Product {product.id}: Selected image URL: {url}')
                    except Exception as e:
                        logger.error(f'Error getting image URL for product {product.id}: {e}')
                        pass
                else:
                    logger.warning(f'Product {product.id}: No valid image media found')
            else:
                logger.warning(f'Product {product.id}: No media files found')
            
            # Get creator info
            creator = product.created_by
            creator_name = f"@{creator.username}" if creator and creator.username else "@wedesignz"
            
            designs.append({
                'id': product.id,
                'title': product.title,
                'creator': creator_name,
                'product_number': product.product_number or None,
                'price': f"₹{float(product.price)}" if product.price else "Free",
                'image': image_url or None  # Return None instead of non-existent placeholder
            })
        
        response_data = {
            'designs': designs,
            'count': len(designs)
        }
        
        # Log the response for debugging
        logger.info(f'Hero section response: {len(designs)} designs')
        for design in designs:
            logger.info(f'  - Design {design["id"]}: title="{design["title"]}", image={"SET" if design["image"] else "NULL"}')
            if design["image"]:
                logger.info(f'    Image URL: {design["image"]}')
        
        # Cache for 1 hour
        cache.set(cache_key, response_data, 3600)
        return Response(response_data)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f'Error in hero_section_designs: {str(e)}')
        logger.error(f'Traceback: {traceback.format_exc()}')
        return Response({
            'designs': [],
            'error': str(e),
            'message': 'An error occurred while fetching hero section designs'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])
def dome_gallery_images(request):
    """
    Get optimized image URLs for dome gallery.
    Returns only image URLs for fast loading.
    """
    from CoreAdmin.models import SystemConfig
    from django.core.cache import cache
    
    try:
        config = SystemConfig.get_config()
        product_ids_raw = config.dome_gallery_designs or []
        
        if not product_ids_raw:
            return Response({'images': [], 'count': 0})
        
        # Convert string IDs to integers (handle both string and int formats)
        product_ids = []
        for pid in product_ids_raw:
            try:
                product_ids.append(int(pid))
            except (ValueError, TypeError):
                continue
        
        if not product_ids:
            return Response({'images': [], 'count': 0, 'message': 'No valid product IDs found'})
        
        cache_key = f'dome_gallery_images_{config.updated_at.timestamp()}'
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)
        
        # Optimized query - only fetch IDs and titles
        products = Product.objects.filter(
            id__in=product_ids,
            status='active',
            visibility_status='show'
        ).only('id', 'title')
        
        images = []
        import os
        from MediaFiles.models import Relation
        
        for product in products:
            media = product.get_media()
            image_url = None
            mockup_media = None
            fallback_media = None
            
            if media:
                media_list = list(media)
                
                # Prefer design images (JPG/PNG, non-mockup), then fallback to mockup
                for m in media_list:
                    if not m.file:
                        continue
                    
                    # Only process image media files (skip vectors, documents, etc.)
                    media_type = getattr(m, 'media_type', 'image')
                    if media_type and media_type.lower() != 'image':
                        continue
                    
                    # Check if file extension is an image format
                    file_name = m.file.name if hasattr(m.file, 'name') else ''
                    is_image_file = False
                    if file_name:
                        file_name_lower = file_name.lower()
                        is_image_file = any(ext in file_name_lower for ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp'])
                    
                    # Skip if not an image file
                    if not is_image_file:
                        continue
                    
                    # Check filename for mockup
                    is_mockup_by_name = False
                    if file_name:
                        file_name_lower = file_name.lower()
                        base_name = os.path.splitext(os.path.basename(file_name_lower))[0]
                        is_mockup_by_name = base_name == 'mockup'
                    
                    # Check Relation metadata for mockup
                    is_mockup_by_meta = False
                    try:
                        relation = Relation.objects.filter(
                            relation_type='Product:Media',
                            id_1=product.pk,
                            id_2=m.pk
                        ).first()
                        
                        if relation and relation.meta:
                            meta = relation.meta
                            if isinstance(meta, dict):
                                is_mockup_by_meta = meta.get('is_mockup', False) or meta.get('type') == 'mockup'
                            elif isinstance(meta, str):
                                meta_lower = str(meta).lower()
                                is_mockup_by_meta = 'mockup' in meta_lower or '"is_mockup":true' in meta_lower
                    except Exception:
                        pass
                    
                    is_mockup = is_mockup_by_name or is_mockup_by_meta
                    
                    # Check if it's JPG or PNG (design image)
                    is_jpg_png = False
                    if file_name:
                        file_name_lower = file_name.lower()
                        is_jpg_png = file_name_lower.endswith(('.jpg', '.jpeg', '.png'))
                    
                    # Prefer design images (JPG/PNG, non-mockup)
                    if is_jpg_png and not is_mockup:
                        try:
                            test_url = m.file.url
                            selected_media = m
                            break
                        except:
                            pass
                    # Store mockup as fallback
                    elif is_mockup and not mockup_media:
                        try:
                            test_url = m.file.url
                            mockup_media = m
                        except:
                            pass
                    # Store first valid image as last fallback
                    elif not fallback_media:
                        try:
                            test_url = m.file.url
                            fallback_media = m
                        except:
                            pass
                
                # Use design image if found, otherwise mockup, otherwise fallback
                if 'selected_media' not in locals():
                    selected_media = mockup_media or fallback_media
                
                if selected_media and selected_media.file:
                    try:
                        url = selected_media.file.url
                        if request and url.startswith('/'):
                            url = request.build_absolute_uri(url)
                        image_url = url
                    except Exception:
                        continue
            
            if image_url:
                images.append({
                    'src': image_url,
                    'alt': product.title or f'Design {product.id}'
                })
        
        response_data = {
            'images': images,
            'count': len(images)
        }
        
        # Cache for 1 hour
        cache.set(cache_key, response_data, 3600)
        return Response(response_data)
    except Exception as e:
        return Response({
            'images': [],
            'count': 0,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])
def featured_designs(request):
    """
    Get featured designs for the landing page slider.
    Returns product data with image, title, creator, price.
    """
    from CoreAdmin.models import SystemConfig
    from django.core.cache import cache
    import logging
    import os
    from MediaFiles.models import Relation
    
    logger = logging.getLogger(__name__)
    
    try:
        config = SystemConfig.get_config()
        product_ids_raw = config.featured_designs or []
        
        if not product_ids_raw:
            return Response({
                'designs': [],
                'message': 'No designs configured for featured section'
            })
        
        # Convert string IDs to integers (handle both string and int formats)
        product_ids = []
        for pid in product_ids_raw:
            try:
                product_ids.append(int(pid))
            except (ValueError, TypeError):
                continue
        
        if not product_ids:
            return Response({
                'designs': [],
                'message': 'No valid product IDs found in featured designs configuration'
            })
        
        # Cache key
        cache_key = f'featured_designs_{config.updated_at.timestamp()}'
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)
        
        # Fetch products with optimized query
        products = Product.objects.filter(
            id__in=product_ids,
            status='active',
            visibility_status='show'
        ).select_related('category', 'created_by').only(
            'id', 'title', 'price', 'created_by', 'category', 'product_number'
        )
        
        # Maintain order from config
        product_dict = {p.id: p for p in products}
        ordered_products = [product_dict[pid] for pid in product_ids if pid in product_dict]
        
        if len(ordered_products) < len(product_ids):
            missing_ids = set(product_ids) - {p.id for p in ordered_products}
            logger.warning(f'Featured designs: {len(missing_ids)} products not found or not active: {missing_ids}')
        
        designs = []
        
        for product in ordered_products:
            # Get media files - same logic as hero section and dome gallery (prefer mockup, fallback to any image)
            image_url = None
            mockup_media = None
            fallback_media = None
            
            try:
                media = product.get_media()
            except Exception as e:
                logger.error(f'Error getting media for product {product.id}: {e}')
                media = None
            
            if media:
                try:
                    media_list = list(media)
                except Exception as e:
                    logger.error(f'Error converting media to list for product {product.id}: {e}')
                    media_list = []
                
                # Prefer design images (JPG/PNG, non-mockup), then fallback to mockup
                for m in media_list:
                    if not m.file:
                        continue
                    
                    # Only process image media files (skip vectors, documents, etc.)
                    media_type = getattr(m, 'media_type', 'image')
                    if media_type and media_type.lower() != 'image':
                        continue
                    
                    # Check if file extension is an image format
                    file_name = m.file.name if hasattr(m.file, 'name') else ''
                    is_image_file = False
                    if file_name:
                        file_name_lower = file_name.lower()
                        is_image_file = any(ext in file_name_lower for ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp'])
                    
                    # Skip if not an image file
                    if not is_image_file:
                        continue
                    
                    # Check filename for mockup
                    is_mockup_by_name = False
                    if file_name:
                        file_name_lower = file_name.lower()
                        base_name = os.path.splitext(os.path.basename(file_name_lower))[0]
                        is_mockup_by_name = base_name == 'mockup'
                    
                    # Check Relation metadata for mockup
                    is_mockup_by_meta = False
                    try:
                        relation = Relation.objects.filter(
                            relation_type='Product:Media',
                            id_1=product.pk,
                            id_2=m.pk
                        ).first()
                        
                        if relation and relation.meta:
                            meta = relation.meta
                            if isinstance(meta, dict):
                                is_mockup_by_meta = meta.get('is_mockup', False) or meta.get('type') == 'mockup'
                            elif isinstance(meta, str):
                                meta_lower = str(meta).lower()
                                is_mockup_by_meta = 'mockup' in meta_lower or '"is_mockup":true' in meta_lower
                    except Exception:
                        pass
                    
                    is_mockup = is_mockup_by_name or is_mockup_by_meta
                    
                    # Check if it's JPG or PNG (design image)
                    is_jpg_png = False
                    if file_name:
                        file_name_lower = file_name.lower()
                        is_jpg_png = file_name_lower.endswith(('.jpg', '.jpeg', '.png'))
                    
                    # Prefer design images (JPG/PNG, non-mockup)
                    if is_jpg_png and not is_mockup:
                        try:
                            test_url = m.file.url
                            selected_media = m
                            break
                        except:
                            pass
                    # Store mockup as fallback
                    elif is_mockup and not mockup_media:
                        try:
                            test_url = m.file.url
                            mockup_media = m
                        except:
                            pass
                    # Store first valid image as last fallback
                    elif not fallback_media:
                        try:
                            test_url = m.file.url
                            fallback_media = m
                        except:
                            pass
                
                # Use design image if found, otherwise mockup, otherwise fallback
                if 'selected_media' not in locals():
                    selected_media = mockup_media or fallback_media
                
                if selected_media and selected_media.file:
                    try:
                        url = selected_media.file.url
                        if request and url.startswith('/'):
                            url = request.build_absolute_uri(url)
                        image_url = url
                    except Exception:
                        continue
            
            # Get creator info
            creator = product.created_by
            creator_name = f"@{creator.username}" if creator and creator.username else "@wedesignz"
            
            designs.append({
                'id': product.id,
                'title': product.title,
                'creator': creator_name,
                'product_number': product.product_number or None,
                'price': f"₹{float(product.price)}" if product.price else "Free",
                'image': image_url or None,
                'category': product.category.name if product.category else 'Uncategorized'
            })
        
        response_data = {
            'designs': designs,
            'count': len(designs)
        }
        
        # Cache for 1 hour
        cache.set(cache_key, response_data, 3600)
        return Response(response_data)
    except Exception as e:
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f'Error in featured_designs: {str(e)}')
        logger.error(f'Traceback: {traceback.format_exc()}')
        return Response({
            'designs': [],
            'error': str(e),
            'message': 'An error occurred while fetching featured designs'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Popular Categories',
    operation_description='Popular Categories endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([AllowAny])
def popular_categories(request):
    """
    Popular categories ranked by number of active designs.
    """
    popular_categories = Category.objects.annotate(
        active_products=Count('products', filter=Q(products__status='active'))
    ).filter(active_products__gt=0).order_by('-active_products')[:10]
    
    return Response({
        'popular_categories': CategorySerializer(popular_categories, many=True).data
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Search and Filter Products',
    operation_description='Search designs with filters by categories, tags, pricing, ratings, and date.',
    manual_parameters=[
        openapi.Parameter('q', openapi.IN_QUERY, description='Search query', type=openapi.TYPE_STRING),
        openapi.Parameter('category', openapi.IN_QUERY, description='Category ID', type=openapi.TYPE_INTEGER),
        openapi.Parameter('tag', openapi.IN_QUERY, description='Tag ID', type=openapi.TYPE_INTEGER),
        openapi.Parameter('pricing', openapi.IN_QUERY, description='Pricing filter', type=openapi.TYPE_STRING, enum=['free', 'paid']),
        openapi.Parameter('min_rating', openapi.IN_QUERY, description='Minimum rating', type=openapi.TYPE_NUMBER),
        openapi.Parameter('date_from', openapi.IN_QUERY, description='Start date (YYYY-MM-DD)', type=openapi.TYPE_STRING),
        openapi.Parameter('date_to', openapi.IN_QUERY, description='End date (YYYY-MM-DD)', type=openapi.TYPE_STRING),
        openapi.Parameter('page', openapi.IN_QUERY, description='Page number', type=openapi.TYPE_INTEGER)
    ],
    responses={
        200: openapi.Response(
            description='Search results retrieved successfully',
            examples={
                'application/json': {
                    'results': [
                        {
                            'id': 1,
                            'title': 'Sample Design',
                            'description': 'Beautiful design',
                            'price': 29.99,
                            'category': 'Graphics',
                            'tags': ['modern', 'minimal'],
                            'rating': 4.5,
                            'created_at': '2024-01-01T00:00:00Z'
                        }
                    ],
                    'total_pages': 5,
                    'current_page': 1,
                    'total_count': 100
                }
            }
        ),
        400: openapi.Response(description='Bad request - invalid parameters')
    },
    tags=['Catalog']
)

@api_view(['GET'])
@permission_classes([AllowAny])
def search_and_filter(request):
    """
    Search designs with filters by categories, tags, pricing, ratings, date.
    """
    query = request.GET.get('q', '')
    category_id = request.GET.get('category')
    tag_id = request.GET.get('tag')
    pricing = request.GET.get('pricing')  # 'free' or 'paid'
    min_rating = request.GET.get('min_rating')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    page = int(request.GET.get('page', 1))
    
    # Base queryset
    products = Product.objects.filter(
        status='active',
        visibility_status='show'
    )
    
    # Search by title, description, product_number, and studio_design_number
    if query:
        products = products.filter(
            Q(title__icontains=query) | 
            Q(description__icontains=query) |
            Q(product_number__icontains=query) |
            Q(studio_design_number__icontains=query)
        )
    
    # Filter by category (include subcategories)
    if category_id:
        try:
            from .models import Category
            # Convert category_id to integer
            category_id_int = int(category_id)
            category = Category.objects.get(id=category_id_int)
            # Get all subcategory IDs (including nested subcategories)
            def get_all_subcategory_ids(cat):
                subcategory_ids = [cat.id]
                for subcat in cat.subcategories.all():
                    subcategory_ids.extend(get_all_subcategory_ids(subcat))
                return subcategory_ids
            
            all_category_ids = get_all_subcategory_ids(category)
            products = products.filter(category_id__in=all_category_ids)
        except (Category.DoesNotExist, ValueError, TypeError):
            # If category doesn't exist or invalid ID, return empty results
            products = products.none()
    
    # Filter by tag
    if tag_id:
        products = products.filter(tags__id=tag_id)
    
    # Filter by pricing
    if pricing == 'free':
        products = products.filter(price__isnull=True)
    elif pricing == 'paid':
        products = products.filter(price__isnull=False)
    
    # Filter by date range
    if date_from:
        products = products.filter(created_at__gte=date_from)
    if date_to:
        products = products.filter(created_at__lte=date_to)
    
    # Pagination
    paginator = Paginator(products, 20)
    page_obj = paginator.get_page(page)
    
    return Response({
        'results': ProductSerializer(page_obj.object_list, many=True, context={'request': request}).data,
        'total_pages': paginator.num_pages,
        'current_page': page,
        'total_count': paginator.count
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Get Product Details',
    operation_description='Get detailed information about a specific product.',
    responses={
        200: openapi.Response(
            description='Product details retrieved successfully',
            examples={
                'application/json': {
                    'product': {
                        'id': 1,
                        'title': 'Sample Design',
                        'description': 'Beautiful design with modern aesthetics',
                        'price': 29.99,
                        'category': 'Graphics',
                        'tags': ['modern', 'minimal'],
                        'rating': 4.5,
                        'downloads': 150,
                        'views': 500,
                        'created_at': '2024-01-01T00:00:00Z',
                        'designer': {
                            'id': 1,
                            'username': 'designer1',
                            'studio_name': 'Creative Studio'
                        }
                    }
                }
            }
        ),
        404: openapi.Response(description='Product not found')
    },
    tags=['Catalog']
)

@api_view(['GET'])
@permission_classes([AllowAny])
def product_detail(request, product_id):
    """
    Get detailed information about a specific product.
    """
    try:
        product = Product.objects.get(id=product_id, status='active')
        
        # Track product view
        if request.user.is_authenticated:
            ProductCounter.objects.create(
                product_counter_type='opened',
                created_by=request.user
            )
            product.attach_counter(
                ProductCounter.objects.filter(
                    product_counter_type='opened',
                    created_by=request.user
                ).last()
            )
        
        return Response({
            'product': ProductSerializer(product, context={'request': request}).data
        })
    except Product.DoesNotExist:
        return Response({
            'error': 'Product not found'
        }, status=status.HTTP_404_NOT_FOUND)


@swagger_auto_schema(
    method='get',
    operation_summary='Categories List',
    operation_description='Categories List endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def categories_list(request):
    """
    Get all categories with subcategories (GET) or create a new category (POST).
    """
    if request.method == 'POST':
        # Handle category creation
        return create_category(request)
    
    # Handle GET request
    categories = Category.objects.filter(parent__isnull=True).prefetch_related('subcategories')
    return Response({
        'categories': CategorySerializer(categories, many=True).data
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Get Subcategories by Category ID',
    operation_description='Get all subcategories for a specific parent category.',
    responses={
        200: openapi.Response(
            description='Subcategories retrieved successfully',
            examples={
                'application/json': {
                    'subcategories': [
                        {
                            'id': 1,
                            'name': 'Subcategory Name',
                            'parent': 'Parent Category Name',
                            'parent_id': 1
                        }
                    ]
                }
            }
        ),
        404: openapi.Response(description='Category not found')
    },
    tags=['Catalog']
)
@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def category_subcategories(request, category_id):
    """
    Get all subcategories for a specific parent category (GET) or create a new subcategory (POST).
    """
    if request.method == 'POST':
        # Handle subcategory creation
        return create_subcategory(request, category_id)
    
    # Handle GET request
    try:
        parent_category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        return Response({
            'error': 'Category not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    # Get all subcategories where parent_id matches the category_id
    subcategories = Category.objects.filter(parent_id=category_id)
    
    return Response({
        'subcategories': CategoryListSerializer(subcategories, many=True).data
    })


def create_category(request):
    """
    Create a new parent category (for designers).
    """
    # Check authentication for POST
    if not request.user.is_authenticated:
        return Response({
            'error': 'Authentication required'
        }, status=status.HTTP_401_UNAUTHORIZED)
    
    # Only allow creating parent categories (no parent_id)
    data = request.data.copy()
    data['parent_id'] = None  # Ensure it's a parent category
    
    serializer = CategorySerializer(data=data)
    if serializer.is_valid():
        category = serializer.save(created_by=request.user)
        
        return Response({
            'category': CategorySerializer(category).data
        }, status=status.HTTP_201_CREATED)
    else:
        return Response({
            'error': 'Invalid data',
            'details': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


def create_subcategory(request, category_id):
    """
    Create a new subcategory under a parent category (for designers).
    """
    # Check authentication for POST
    if not request.user.is_authenticated:
        return Response({
            'error': 'Authentication required'
        }, status=status.HTTP_401_UNAUTHORIZED)
    
    # Verify parent category exists
    try:
        parent_category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        return Response({
            'error': 'Parent category not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    # Set parent_id to the category_id
    data = request.data.copy()
    data['parent_id'] = category_id
    # Remove any created_by fields (read-only) - we'll set it via context
    data.pop('created_by', None)
    data.pop('created_by_id', None)
    
    # Pass created_by and request via context so serializer can use it
    serializer = CategorySerializer(
        data=data,
        context={
            'created_by': request.user,
            'request': request
        }
    )
    if serializer.is_valid():
        subcategory = serializer.save()
        
        return Response({
            'subcategory': CategoryListSerializer(subcategory).data
        }, status=status.HTTP_201_CREATED)
    else:
        return Response({
            'error': 'Invalid data',
            'details': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='get',
    operation_summary='Tags List',
    operation_description='Tags List endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def tags_list(request):
    """
    Get all available tags (GET) or create a new tag (POST).
    """
    if request.method == 'POST':
        # Handle tag creation
        return create_tag(request)
    
    # Handle GET request
    tags = Tags.objects.all().order_by('name')
    return Response({
        'tags': TagsSerializer(tags, many=True).data
    })


def create_tag(request):
    """
    Create a new tag (for designers).
    """
    # Check authentication for POST
    if not request.user.is_authenticated:
        return Response({
            'error': 'Authentication required'
        }, status=status.HTTP_401_UNAUTHORIZED)
    
    # Get tag name from request
    tag_name = request.data.get('name', '').strip()
    if not tag_name:
        return Response({
            'error': 'Tag name is required'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Check if tag already exists (case-insensitive)
    existing_tag = Tags.objects.filter(name__iexact=tag_name).first()
    if existing_tag:
        # Return existing tag instead of creating a duplicate
        return Response({
            'tag': TagsSerializer(existing_tag).data
        }, status=status.HTTP_200_OK)
    
    # Create new tag
    serializer = TagsSerializer(data={
        'name': tag_name,
        'tags_type': 'manually_added'
    })
    
    if serializer.is_valid():
        tag = serializer.save(created_by=request.user)
        
        return Response({
            'tag': TagsSerializer(tag).data
        }, status=status.HTTP_201_CREATED)
    else:
        return Response({
            'error': 'Invalid data',
            'details': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='get',
    operation_summary='Bundles List',
    operation_description='Bundles List endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([AllowAny])
def bundles_list(request):
    """
    Get all available collection bundles.
    """
    bundles = CollectionBundle.objects.filter(status='available').order_by('-created_at')
    return Response({
        'bundles': CollectionBundleSerializer(bundles, many=True).data
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Bundle Detail',
    operation_description='Bundle Detail endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([AllowAny])
def bundle_detail(request, bundle_id):
    """
    Get detailed information about a specific bundle.
    """
    try:
        bundle = CollectionBundle.objects.get(id=bundle_id, status='available')
        return Response({
            'bundle': CollectionBundleSerializer(bundle).data
        })
    except CollectionBundle.DoesNotExist:
        return Response({
            'error': 'Bundle not found'
        }, status=status.HTTP_404_NOT_FOUND)


# ==================== DESIGNER CONSOLE - DESIGN MANAGEMENT ====================

@swagger_auto_schema(
    method='post',
    operation_summary='Upload Design',
    operation_description='Upload Design endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_design(request):
    """
    Upload a new design with metadata and files.
    Files are processed asynchronously for instant response.
    """
    from django.db import transaction
    from common.relations import attach_relation
    from Catalog.tasks import process_single_design_upload
    import uuid
    import logging
    import os
    import time
    
    logger = logging.getLogger(__name__)
    
    try:
        # Validate required fields
        required_fields = ['title', 'description', 'category_id', 'product_plan_type']
        missing_fields = [f for f in required_fields if f not in request.data]
        if missing_fields:
            return Response({
                'error': f'Missing required fields: {", ".join(missing_fields)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate files
        if 'design_files' not in request.FILES:
            return Response({
                'error': 'Design files are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        design_files = request.FILES.getlist('design_files')
        if not design_files:
            return Response({
                'error': 'At least one design file is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get category (this will be subcategory if selected, parent category otherwise)
        try:
            category = Category.objects.get(id=request.data['category_id'])
        except Category.DoesNotExist:
            return Response({
                'error': f'Category with id {request.data["category_id"]} does not exist'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Generate platform ID
        platform_id = f"WD{uuid.uuid4().hex[:9].upper()}"
        
        # Determine product_plan_type: 'free' or 'basic' (paid)
        # NOTE: This field is for categorization only (free vs paid designs).
        # It does NOT restrict design access based on subscription plans.
        # All paid designs are available to all subscription plans.
        product_plan_type = request.data.get('product_plan_type', 'free')
        
        # Handle price - use global price from config if paid (basic)
        # Designers can no longer set individual prices; all paid designs use the global price
        price_value = None
        if product_plan_type == 'basic':
            from common.business_config import BusinessConfig
            price_value = BusinessConfig.get_design_price()
        
        # Determine studio for design association
        # For studio members, get studio from membership
        # For studio owners, get owned studio
        studio = None
        uploaded_by_member_id = None  # Track if uploaded by studio member
        from Profiles.models import Studio, StudioMember
        studio_membership = StudioMember.objects.filter(
            member=request.user, 
            status='active'
        ).select_related('studio').first()
        
        if studio_membership:
            # User is a studio member - use their studio
            studio = studio_membership.studio
            uploaded_by_member_id = request.user.id  # Track the member who uploaded
            # Set created_by to studio owner (not the member) - ownership belongs to studio owner
            product_owner = studio.created_by
            logger.info(f'User is studio member, associating design with studio: {studio.name} (ID: {studio.id}), owner: {product_owner.id}, uploaded by member: {uploaded_by_member_id}')
        else:
            # Check if user owns a studio
            studio = Studio.objects.filter(created_by=request.user).first()
            if studio:
                logger.info(f'User owns studio, associating design with studio: {studio.name} (ID: {studio.id})')
            # For studio owners or individual designers, created_by is themselves
            product_owner = request.user
        
        # Generate studio design number if studio exists
        studio_design_number = None
        if studio and studio.wedesignz_auto_name:
            from common.studio_name_generator import generate_design_numbers
            design_numbers = generate_design_numbers(studio.wedesignz_auto_name)
            studio_design_number = design_numbers['studio_number']
            logger.info(f'Generated studio design number: {studio_design_number}')
        
        # Build product_metadata to track uploader if member uploaded
        product_metadata = {}
        if uploaded_by_member_id:
            product_metadata['uploaded_by_member_id'] = uploaded_by_member_id
        
        # Create product FIRST (instant operation)
        # Temporarily disable the pre_save signal to avoid Studio query overhead
        # Since we're already setting product_number, the signal would skip anyway,
        # but disabling it entirely is more efficient
        logger.info('Creating product in database...')
        from django.db.models.signals import pre_save
        from Catalog.models import generate_design_numbers_signal
        
        with transaction.atomic():
            # Temporarily disconnect the signal to avoid any overhead
            pre_save.disconnect(generate_design_numbers_signal, sender=Product)
            try:
                # Create product without signal overhead
                product = Product.objects.create(
                    title=request.data['title'],
                    description=request.data['description'],
                    category_id=request.data['category_id'],  # This is subcategory if selected
                    product_plan_type=product_plan_type,  # 'free' or 'basic'
                    status='draft',  # Default status
                    product_number=platform_id,  # Already generated
                    studio_design_number=studio_design_number,  # Set studio design number if available
                    price=price_value,  # Only set if paid
                    color=request.data.get('color') or None,
                    created_by=product_owner,  # Studio owner if member uploaded, else request.user
                    product_metadata=product_metadata if product_metadata else {}  # Store member ID if uploaded by member
                )
                logger.info(f'Product created with ID: {product.id}, owner: {product.created_by.id}, uploaded_by_member: {uploaded_by_member_id}')
                # Note: Studio association is handled via studio_design_number
                # Designs can be queried by studio using the design number prefix
            finally:
                # Reconnect the signal
                pre_save.connect(generate_design_numbers_signal, sender=Product)
            
            # Create Category:Product relation
            attach_relation('Category:Product', category, product, created_by=request.user)
            logger.info('Category:Product relation created')
            
            # Attach tags - establish relation between tags and product
            tag_ids = request.data.getlist('tags') if hasattr(request.data, 'getlist') else (
                request.data.get('tags', []) if isinstance(request.data.get('tags'), list) else 
                [request.data.get('tags')] if request.data.get('tags') else []
            )
            
            logger.info(f'Attaching {len(tag_ids)} tags...')
            for tag_id in tag_ids:
                try:
                    tag = Tags.objects.get(id=tag_id)
                    # attach_tag() already creates the relation, no need to call attach_relation again
                    product.attach_tag(tag, created_by=request.user)
                except Tags.DoesNotExist:
                    logger.warning(f'Tag with ID {tag_id} not found')
        
        # Save files synchronously (must be done before response to avoid file handle issues)
        # Then queue Celery task asynchronously
        from django.conf import settings
        
        timestamp = int(time.time())
        saved_files_data = []
        
        logger.info(f'Saving {len(design_files)} files to temp location...')
        for idx, file in enumerate(design_files):
            try:
                file_name = f"{timestamp}_{file.name}"
                relative_path = f'{request.user.id}/temp/{file_name}'
                
                logger.info(f'Saving file {idx+1}/{len(design_files)}: {file.name} ({file.size} bytes)')
                # Use default_storage.save() which handles file writing efficiently
                # This is faster and more reliable than manual file writing
                saved_path = default_storage.save(relative_path, file)
                
                saved_files_data.append({
                    'path': saved_path,
                    'name': file.name,
                    'size': file.size
                })
                logger.info(f'File {idx+1} saved successfully to {saved_path}')
            except Exception as e:
                logger.error(f'Error saving file {file.name}: {str(e)}', exc_info=True)
                # Continue with other files even if one fails
        
        # Process files synchronously to ensure they're attached immediately
        # This ensures media files are available right after upload
        logger.info('Processing files synchronously to ensure immediate availability...')
        try:
            from Catalog.tasks import process_single_design_upload
            # Call the task function using apply() for synchronous execution
            # This ensures files are processed and attached before response is sent
            task_result = process_single_design_upload.apply(
                args=(product.id, saved_files_data, tag_ids, platform_id)
            )
            if task_result.successful():
                result_data = task_result.result
                processed_count = result_data.get('processed_files', 0) if isinstance(result_data, dict) else 0
                logger.info(f'Files processed successfully: {processed_count} files attached to product {product.id}')
            else:
                logger.warning(f'Task completed but may have had issues: {task_result.result}')
        except Exception as sync_error:
            logger.error(f'Failed to process files synchronously: {str(sync_error)}', exc_info=True)
            # Try async as fallback
            try:
                logger.warning('Falling back to async processing...')
                result = process_single_design_upload.delay(
                product_id=product.id,
                design_files_data=saved_files_data,
                tag_ids=tag_ids,
                platform_id=platform_id
            )
                logger.info(f'Celery task queued as fallback with task_id: {result.id}')
            except Exception as async_error:
                logger.error(f'Failed to queue async task: {str(async_error)}', exc_info=True)
                # Don't fail the request - files are saved, can be processed later manually
        
        # Return success response
        logger.info(f'Returning success response for product {product.id}')
        return Response({
            'message': 'Design uploaded successfully',
            'product_id': product.id,
            'platform_id': platform_id,
            'status': 'draft'
        }, status=status.HTTP_201_CREATED)
            
    except Exception as e:
        logger.error(f'Upload design error: {str(e)}', exc_info=True)
        return Response({
            'error': f'Failed to upload design: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='post',
    operation_summary='Upload Designs Bulk',
    operation_description='Bulk upload multiple designs using a zip file. All design information is read from metadata.xlsx.',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'zip_file': openapi.Schema(type=openapi.TYPE_STRING, format=openapi.FORMAT_BINARY, description='Zip file containing designs')
        },
        required=['zip_file']
    ),
    responses={
        200: openapi.Response(description='Success'),
        400: openapi.Response(description='Bad request - validation failed')
    },
    tags=['API']
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_designs_bulk(request):
    """
    Bulk upload designs using a zip file.
    Validates zip file structure, metadata.xlsx, and design folders.
    Does NOT enforce minimum design requirement (unlike onboarding).
    """
    import zipfile
    import io
    import os
    import logging
    from openpyxl import load_workbook
    from django.db import transaction
    from django.core.files.storage import default_storage
    from django.utils import timezone
    from Profiles.models import DesignProcessingTask
    from Profiles.tasks import process_design_upload_task
    
    logger = logging.getLogger(__name__)
    
    # Check if zip file is provided
    if 'zip_file' not in request.FILES:
        return Response({
            'error': '❌ No file uploaded: Please select a zip file to upload.',
            'validation_errors': ['❌ No file uploaded: Please select a zip file to upload.']
        }, status=status.HTTP_400_BAD_REQUEST)
    
    zip_file = request.FILES['zip_file']
    
    # Validate file size (1GB = 1073741824 bytes)
    MAX_FILE_SIZE = 1073741824  # 1GB
    if zip_file.size > MAX_FILE_SIZE:
        file_size_mb = zip_file.size / (1024*1024)
        return Response({
            'error': f'❌ File too large: Your file is {file_size_mb:.2f} MB, but the maximum allowed size is 1GB (1024 MB). Please compress your files or split them into smaller zip files.',
            'validation_errors': [f'❌ File too large: Your file is {file_size_mb:.2f} MB, but the maximum allowed size is 1GB (1024 MB). Please compress your files or split them into smaller zip files.']
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Validate file extension
    if not zip_file.name.lower().endswith('.zip'):
        return Response({
            'error': '❌ Invalid file type: Please upload a .zip file. The file you selected is not a zip archive.',
            'validation_errors': ['❌ Invalid file type: Please upload a .zip file. The file you selected is not a zip archive.']
        }, status=status.HTTP_400_BAD_REQUEST)
    
    validation_errors = []
    zip_folders = {}
    valid_design_folders = {}
    all_files = []
    root_folder = ''
    metadata_file = None
    zip_content = None
    
    try:
        # Read zip file into memory (only once)
        zip_content = zip_file.read()
        zip_buffer = io.BytesIO(zip_content)
        
        # Open zip file
        with zipfile.ZipFile(zip_buffer, 'r') as zip_ref:
            # Get all file names in zip
            all_files = zip_ref.namelist()
            
            # Check for metadata.xlsx (can be at root or in a subfolder)
            metadata_file = None
            for file_name in all_files:
                lower_name = file_name.lower()
                if lower_name == 'metadata.xlsx' or lower_name.endswith('/metadata.xlsx'):
                    metadata_file = file_name
                    break
            
            if not metadata_file:
                validation_errors.append('❌ metadata.xlsx file not found: Your zip file must contain a file named "metadata.xlsx" at the root level (inside the main folder).')
            else:
                # Read and parse metadata.xlsx
                try:
                    metadata_data = zip_ref.read(metadata_file)
                    metadata_buffer = io.BytesIO(metadata_data)
                    workbook = load_workbook(metadata_buffer, data_only=True)
                    sheet = workbook.active
                    
                    # Find folder_name column
                    header_row = None
                    folder_name_col = None
                    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=False), 1):
                        values = [str(cell.value).lower() if cell.value else '' for cell in row]
                        if 'folder_name' in values:
                            header_row = idx
                            folder_name_col = values.index('folder_name') + 1
                            break
                    
                    if not folder_name_col:
                        validation_errors.append('❌ Missing required column: The metadata.xlsx file must have a column named "folder_name" (case-insensitive). This column should list all your design folder names.')
                    else:
                        # Extract folder names from Excel
                        excel_folders = set()
                        for row in sheet.iter_rows(min_row=header_row + 1, values_only=False):
                            folder_name_cell = row[folder_name_col - 1]
                            if folder_name_cell.value:
                                folder_name = str(folder_name_cell.value).strip()
                                if folder_name:
                                    excel_folders.add(folder_name)
                        
                        # System folders to ignore
                        SYSTEM_FOLDERS = ['__macosx', '.ds_store', 'rar', '.rar', 'thumbs.db']
                        
                        # Find root folder
                        root_folder = ''
                        for file_name in all_files:
                            if '/' in file_name and not file_name.endswith('/'):
                                parts = file_name.split('/')
                                if len(parts) >= 2:
                                    root_folder = parts[0]
                                    break
                        
                        # Extract actual design folders from zip - handle both 2-part and 3-part paths
                        zip_folders = {}
                        required_files = {'.eps', '.cdr', '.jpg', '.png'}
                        folders_with_wrong_structure = []  # Track folders with wrong path structure
                        
                        for file_name in all_files:
                            if file_name == metadata_file or file_name.endswith('/'):
                                continue
                            
                            # Skip Mac resource fork files (files starting with ._)
                            file_name_only_check = file_name.split('/')[-1]
                            if file_name_only_check.startswith('._'):
                                continue
                            
                            if '/' in file_name:
                                parts = file_name.split('/')
                                
                                # Handle both cases:
                                # 1. root_folder/design_folder/file.ext (3 parts) - e.g., "123/WD01/WD01.eps"
                                # 2. design_folder/file.ext (2 parts) - e.g., "WD01/WD01.eps" (if zip created from inside root folder)
                                
                                if len(parts) < 2:
                                    # Skip files at root level (not in any folder)
                                    continue
                                
                                # Determine folder name and file info based on path length
                                if len(parts) == 2:
                                    # Case 2: design_folder/file.ext (zip created from inside root folder)
                                    folder_name = parts[0]
                                    file_name_only = parts[1]
                                elif len(parts) == 3:
                                    # Case 1: root_folder/design_folder/file.ext (normal case)
                                    root_folder_name = parts[0].lower()
                                    folder_name = parts[1]
                                    file_name_only = parts[2]
                                    
                                    # Skip system folders
                                    if root_folder_name in SYSTEM_FOLDERS or folder_name.lower() in SYSTEM_FOLDERS:
                                        continue
                                else:
                                    # More than 3 parts - wrong structure
                                    # First check if root folder is a system folder (e.g., __MACOSX)
                                    root_folder_name = parts[0].lower()
                                    if root_folder_name in SYSTEM_FOLDERS:
                                        continue  # Skip __MACOSX and other system folders
                                    
                                    if len(parts) >= 2:
                                        folder_name = parts[1]  # Still use second part as folder name for error reporting
                                    else:
                                        continue
                                    if folder_name not in folders_with_wrong_structure:
                                        folders_with_wrong_structure.append(folder_name)
                                    continue
                                
                                file_ext = os.path.splitext(file_name_only)[1].lower()
                                file_name_lower = file_name_only.lower()
                                
                                # Check if it's an optional mockup file (case insensitive - any case of "mockup")
                                is_mockup_file = file_name_lower == 'mockup.jpg' or file_name_lower == 'mockup.png'
                                
                                # Skip system folders (check folder name)
                                if folder_name.lower() in SYSTEM_FOLDERS:
                                    continue
                                
                                # Process files in design folders
                                # Accept required files or optional mockup files
                                if len(parts) == 2 or len(parts) == 3:
                                    if file_ext in required_files or is_mockup_file:
                                        if folder_name not in zip_folders:
                                            zip_folders[folder_name] = set()
                                        # Only add required files to the set (mockup is optional)
                                        if file_ext in required_files:
                                            zip_folders[folder_name].add(file_ext)
                        
                        # Check for folder structure issues first
                        if folders_with_wrong_structure:
                            structure_examples = list(folders_with_wrong_structure)[:5]
                            more_count = len(folders_with_wrong_structure) - 5 if len(folders_with_wrong_structure) > 5 else 0
                            validation_errors.append(f'❌ Incorrect folder structure detected: Some folders have the wrong path structure. Expected format: root_folder/design_folder/file.ext (3 levels) or design_folder/file.ext (2 levels). Found issues in: {", ".join(structure_examples)}{f" (and {more_count} more)" if more_count > 0 else ""}. Please ensure your zip has the structure: root_folder/design_folder/file.ext or design_folder/file.ext')
                        
                        # Check if any folders were detected at all
                        if not zip_folders:
                            # Provide more diagnostic information
                            design_file_count = 0
                            sample_paths = []
                            for file_name in all_files:
                                if file_name == metadata_file or file_name.endswith('/'):
                                    continue
                                if '/' in file_name:
                                    parts = file_name.split('/')
                                    if len(parts) >= 2:
                                        file_ext = os.path.splitext(parts[-1])[1].lower()
                                        if file_ext in required_files:
                                            design_file_count += 1
                                            if len(sample_paths) < 3:
                                                sample_paths.append(file_name)
                            
                            error_msg = '❌ No design folders detected: The system could not find any design folders in your zip file. '
                            if design_file_count > 0:
                                error_msg += f'Found {design_file_count} design files, but they may have incorrect folder structure. '
                                if sample_paths:
                                    error_msg += f'Sample file paths: {", ".join(sample_paths[:3])}. '
                            error_msg += 'Please ensure: (1) Your zip has the structure: root_folder/design_folder/file.ext (3 levels) or design_folder/file.ext (2 levels), (2) Each folder contains files with extensions: .eps, .cdr, .jpg, .png'
                            validation_errors.append(error_msg)
                        else:
                            # Analyze which folders are missing which files BEFORE filtering
                            invalid_folders_detailed = []
                            for folder_name, files in zip_folders.items():
                                missing_files = required_files - files
                                if missing_files:
                                    invalid_folders_detailed.append({
                                        'folder': folder_name,
                                        'missing': list(missing_files),
                                        'has': list(files)
                                    })
                            
                            # Filter folders that have all required files
                            valid_design_folders = {}
                            for folder_name, files in zip_folders.items():
                                if all(ext in files for ext in required_files):
                                    valid_design_folders[folder_name] = files
                            
                            # Provide detailed feedback about missing files
                            if invalid_folders_detailed:
                                total_invalid = len(invalid_folders_detailed)
                                total_detected = len(zip_folders)
                                
                                # Show first few examples with details
                                examples = invalid_folders_detailed[:5]
                                example_texts = []
                                for item in examples:
                                    has_files = ', '.join(item['has']) if item['has'] else 'none'
                                    example_texts.append(f"{item['folder']} (missing: {', '.join(item['missing'])}, has: {has_files})")
                                
                                more_count = total_invalid - 5 if total_invalid > 5 else 0
                                
                                if len(valid_design_folders) == 0:
                                    # All folders are invalid - provide comprehensive error
                                    validation_errors.append(f'❌ All {total_detected} design folders are missing required files: Each folder must contain all 4 file types (.eps, .cdr, .jpg, .png). Examples: {"; ".join(example_texts)}{f" (and {more_count} more folders with similar issues)" if more_count > 0 else ""}. Please add the missing file types to each folder.')
                                else:
                                    # Some folders are valid, some are not
                                    validation_errors.append(f'❌ {total_invalid} out of {total_detected} design folders are missing required files. Each folder must contain all 4 file types (.eps, .cdr, .jpg, .png). Examples: {"; ".join(example_texts)}{f" (and {more_count} more)" if more_count > 0 else ""}')
                        
                        # NOTE: No minimum design requirement for regular bulk uploads
                        # (unlike onboarding which requires minimum designs)
                        
                        # Validate folder_name mapping (only if we have valid folders)
                        if valid_design_folders:
                            zip_folder_names = set(valid_design_folders.keys())
                            missing_in_zip = excel_folders - zip_folder_names
                            missing_in_excel = zip_folder_names - excel_folders
                            
                            if missing_in_zip:
                                missing_list = list(missing_in_zip)[:10]
                                more_count = len(missing_in_zip) - 10 if len(missing_in_zip) > 10 else 0
                                validation_errors.append(f'❌ Folders listed in metadata.xlsx but not found in zip file: {", ".join(missing_list)}{f" (and {more_count} more)" if more_count > 0 else ""}. Please ensure these folders exist in your zip file and contain all required files (.eps, .cdr, .jpg, .png).')
                            
                            if missing_in_excel:
                                missing_list = list(missing_in_excel)[:10]
                                more_count = len(missing_in_excel) - 10 if len(missing_in_excel) > 10 else 0
                                validation_errors.append(f'❌ Folders found in zip file but not listed in metadata.xlsx: {", ".join(missing_list)}{f" (and {more_count} more)" if more_count > 0 else ""}. Please add these folders to the "folder_name" column in your metadata.xlsx file.')
                        
                        # Validate each design folder has required files
                        invalid_folders = []
                        for folder_name, files in valid_design_folders.items():
                            missing_files = required_files - files
                            if missing_files:
                                invalid_folders.append(f'{folder_name} (missing: {", ".join(missing_files)})')
                        
                        if invalid_folders:
                            more_count = len(invalid_folders) - 10 if len(invalid_folders) >= 10 else 0
                            validation_errors.append(f'❌ Some design folders are missing required files. Each folder must contain all 4 file types (.eps, .cdr, .jpg, .png). Affected folders: {"; ".join(invalid_folders[:10])}{f" (and {more_count} more)" if more_count > 0 else ""}')
                
                except Exception as e:
                    validation_errors.append(f'❌ Error reading metadata.xlsx: {str(e)}. The Excel file appears to be corrupted or in an invalid format. Please ensure it\'s a valid .xlsx file.')
        
        # If validation errors exist, return them
        if validation_errors:
            return Response({
                'error': 'Zip file validation failed',
                'validation_errors': validation_errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get total designs count
        total_designs = len(valid_design_folders)
        
        # Save zip file to storage
        try:
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            user_upload_dir = os.path.join(settings.MEDIA_ROOT, str(request.user.id), 'uploads')
            os.makedirs(user_upload_dir, exist_ok=True)
            
            zip_file_path = f'{request.user.id}/uploads/{timestamp}_{zip_file.name}'
            
            # Reset file pointer before saving
            zip_file.seek(0)
            saved_path = default_storage.save(zip_file_path, zip_file)
            
            # Verify the file was saved correctly
            logger.info(f"Saved zip file - requested path: {zip_file_path}, saved path: {saved_path}")
            if not default_storage.exists(saved_path):
                logger.error(f"WARNING: File was saved but cannot be verified at path: {saved_path}")
            else:
                logger.info(f"Verified: File exists at saved path: {saved_path}")
            
            # Create DesignProcessingTask record
            with transaction.atomic():
                processing_task = DesignProcessingTask.objects.create(
                    user=request.user,
                    zip_file_path=saved_path,
                    total_designs=total_designs,
                    status='pending'
                )
                
                # Queue Celery task to process designs asynchronously
                process_design_upload_task.delay(processing_task.id, saved_path)
            
            # Return success response immediately with task_id
            return Response({
                'message': 'Zip file uploaded successfully. Designs are being processed in the background.',
                'data': {
                    'task_id': processing_task.id,
                    'total_designs': total_designs,
                    'status': 'pending'
                }
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            import traceback
            error_traceback = traceback.format_exc()
            logger.error(f"Error saving zip file: {str(e)}")
            logger.error(error_traceback)
            return Response({
                'error': f'Failed to save zip file: {str(e)}',
                'traceback': error_traceback if settings.DEBUG else None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    except zipfile.BadZipFile:
        return Response({
            'error': 'Invalid zip file format'
        }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        logger.error(f"Error in upload_designs_bulk: {str(e)}")
        logger.error(error_traceback)
        return Response({
            'error': f'Failed to process zip file: {str(e)}',
            'traceback': error_traceback if settings.DEBUG else None
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='My Designs',
    operation_description='My Designs endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_designs(request):
    """
    Get designer's designs with filtering and pagination.
    For studio owners: shows all designs from their studio
    For studio members: shows only designs they uploaded (but owned by studio owner)
    """
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        from django.core.paginator import Paginator
        from Profiles.models import DesignerProfile, Studio, StudioMember
        
        # Get query parameters
        status_filter = request.GET.get('status')
        category_id = request.GET.get('category_id')
        search = request.GET.get('search')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        try:
            page = int(request.GET.get('page', 1))
        except (ValueError, TypeError):
            page = 1
        
        try:
            limit = int(request.GET.get('limit', 20))  # Default to 20 if not provided
        except (ValueError, TypeError):
            limit = 20
        
        # Validate limit (prevent abuse)
        if limit < 1:
            limit = 20
        if limit > 100:
            limit = 100  # Max 100 items per page
        
        # Check if user is a studio owner or member
        try:
            designer_profile = DesignerProfile.objects.get(created_by=request.user)
            is_studio_owner = designer_profile.is_studio_owner()
            is_studio_member = designer_profile.is_studio_member()
        except DesignerProfile.DoesNotExist:
            # For users without DesignerProfile, check StudioMember directly
            # This handles studio members who don't have a DesignerProfile
            is_studio_owner = Studio.objects.filter(created_by=request.user).exists()
            is_studio_member = StudioMember.objects.filter(member=request.user, status='active').exists()
        
        from django.db.models import Q
        
        # Base queryset logic:
        # - Studio owners: Show all designs from their studio (created_by = owner OR uploaded by members)
        # - Studio members: Show only designs they uploaded (query by product_metadata['uploaded_by_member_id'])
        # - Individual designers: Show only their own designs
        if is_studio_owner:
            # Get the studio owned by this user
            studio = Studio.objects.filter(created_by=request.user).first()
            if studio:
                # Get all studio members
                studio_members = StudioMember.objects.filter(studio=studio, status='active').values_list('member', flat=True)
                # Show designs:
                # 1. Created by studio owner directly (created_by = owner)
                # 2. Uploaded by any studio member (product_metadata['uploaded_by_member_id'] in studio_members)
                # Use fallback method for JSONB queries to avoid type casting issues
                owner_designs = Product.objects.filter(created_by=request.user).exclude(status='deleted')
                # For member designs, use Python filtering to avoid JSONB type casting issues
                all_designs = Product.objects.exclude(status='deleted').exclude(product_metadata__isnull=True)
                member_design_ids = []
                for design in all_designs.iterator(chunk_size=100):
                    if design.product_metadata and isinstance(design.product_metadata, dict):
                        member_id = design.product_metadata.get('uploaded_by_member_id')
                        if member_id and member_id in list(studio_members):
                            member_design_ids.append(design.id)
                member_designs = Product.objects.filter(id__in=member_design_ids) if member_design_ids else Product.objects.none()
                designs = (owner_designs | member_designs).distinct()
            else:
                # Fallback: just owner's designs
                designs = Product.objects.filter(created_by=request.user).exclude(status='deleted')
        elif is_studio_member:
            # Studio members: Show only designs they uploaded
            # Query by product_metadata['uploaded_by_member_id'] == request.user.id
            # Note: These designs have created_by = studio owner, but were uploaded by the member
            # Use Python filtering to avoid JSONB type casting issues with PostgreSQL
            all_designs = Product.objects.exclude(status='deleted').exclude(product_metadata__isnull=True)
            design_ids = []
            for design in all_designs.iterator(chunk_size=100):
                if design.product_metadata and isinstance(design.product_metadata, dict):
                    if design.product_metadata.get('uploaded_by_member_id') == request.user.id:
                        design_ids.append(design.id)
            designs = Product.objects.filter(id__in=design_ids) if design_ids else Product.objects.none()
            logger.info(f'Studio member {request.user.id}: Found {len(design_ids)} designs')
        else:
            # Individual designer: Show only their own designs
            designs = Product.objects.filter(created_by=request.user).exclude(status='deleted')
        
        # Apply filters
        if status_filter:
            designs = designs.filter(status=status_filter)
        
        if category_id:
            designs = designs.filter(category_id=category_id)
        
        if search:
            designs = designs.filter(
                Q(title__icontains=search) | Q(description__icontains=search)
            )
        
        if date_from:
            designs = designs.filter(created_at__gte=date_from)
        if date_to:
            designs = designs.filter(created_at__lte=date_to)
        
        # Order by creation date
        designs = designs.order_by('-created_at')
        
        # Pagination
        paginator = Paginator(designs, limit)
        page_obj = paginator.get_page(page)
        
        return Response({
            'designs': ProductSerializer(page_obj.object_list, many=True, context={'request': request}).data,
            'total_pages': paginator.num_pages,
            'current_page': page,
            'total_count': paginator.count,
            'filters_applied': {
                'status': status_filter,
                'category_id': category_id,
                'search': search,
                'date_from': date_from,
                'date_to': date_to
            }
        })
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error in my_designs endpoint for user {request.user.id}: {str(e)}', exc_info=True)
        return Response({
            'error': 'Failed to fetch designs',
            'detail': str(e) if settings.DEBUG else 'An error occurred while fetching your designs'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Design Detail',
    operation_description='Design Detail endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def design_detail(request, design_id):
    """
    Get, update, or delete a specific design.
    Allows access to:
    - Design owner (created_by)
    - Studio members who uploaded the design (product_metadata['uploaded_by_member_id'])
    """
    from django.db.models import Q
    from Profiles.models import StudioMember
    
    # Check if user is a studio member
    is_studio_member = StudioMember.objects.filter(member=request.user, status='active').exists()
    
    # Try to get design - allow access if user is owner OR uploaded it as studio member
    try:
        if is_studio_member:
            # For studio members: check if they uploaded it (via product_metadata)
            design = Product.objects.get(
                Q(id=design_id) & (
                    Q(created_by=request.user) | 
                    Q(product_metadata__uploaded_by_member_id=request.user.id)
                )
            )
        else:
            # For studio owners and individual designers: check ownership only
            design = Product.objects.get(id=design_id, created_by=request.user)
    except Product.DoesNotExist:
        return Response({
            'error': 'Design not found or you do not have permission'
        }, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        from .serializers import DesignDetailSerializer
        return Response({
            'design': DesignDetailSerializer(design, context={'request': request}).data
        })
    
    elif request.method == 'PUT':
        # Check if design can be edited
        if design.status == 'active':
            return Response({
                'error': 'Cannot edit approved designs. Contact admin for changes.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        from django.db import transaction
        from MediaFiles.models import Media
        from django.utils import timezone
        
        # Check if files are being updated
        has_file_updates = 'design_files' in request.FILES
        
        try:
            with transaction.atomic():
                # Prepare update data
                update_data = {}
                
                # Handle text fields
                if 'title' in request.data:
                    update_data['title'] = request.data['title']
                if 'description' in request.data:
                    update_data['description'] = request.data['description']
                if 'category_id' in request.data:
                    update_data['category_id'] = request.data['category_id']
                # Price is now managed globally via SystemConfig, so we don't allow designers to update it directly
                # If product_plan_type changes, update price accordingly
                if 'product_plan_type' in request.data:
                    update_data['product_plan_type'] = request.data['product_plan_type']
                    # Update price based on plan type using global config
                    from common.business_config import BusinessConfig
                    if request.data['product_plan_type'] == 'basic':
                        update_data['price'] = BusinessConfig.get_design_price()
                    elif request.data['product_plan_type'] == 'free':
                        update_data['price'] = None
                if 'color' in request.data:
                    update_data['color'] = request.data['color']
                
                # If files are being updated, set status to draft
                if has_file_updates:
                    update_data['status'] = 'draft'
                
                # Update product fields
                serializer = ProductSerializer(design, data=update_data, partial=True, context={'request': request})
                if serializer.is_valid():
                    serializer.save(updated_by=request.user)
                    
                    # Handle file uploads if present
                    if has_file_updates:
                        design_files = request.FILES.getlist('design_files')
                        
                        # Ensure product_number is available
                        if not design.product_number:
                            design.refresh_from_db()
                            if not design.product_number:
                                return Response({
                                    'error': 'Product has no product_number. Please contact support.'
                                }, status=status.HTTP_400_BAD_REQUEST)
                        
                        product_number = design.product_number
                        
                        # Set product context for file path generation
                        # #region agent log
                        import json
                        import os
                        log_path = os.getenv('DEBUG_LOG_PATH', os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs', 'debug.log'))
                        try:
                            with open(log_path, 'a') as f:
                                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"B","location":"Catalog/views.py:design_detail","message":"Setting product context before Media.create (design_detail view)","data":{"product_id":design.id,"user_id":request.user.id},"timestamp":int(__import__('time').time()*1000)})+'\n')
                        except: pass
                        # #endregion
                        Media.set_product_context(design.id)
                        try:
                            for file in design_files:
                                # Detect if this is a mockup file
                                file_name_lower = file.name.lower()
                                is_mockup = 'mockup' in file_name_lower
                                
                                # Generate new filename using product_number
                                file_ext = os.path.splitext(file.name)[1].lower()
                                if is_mockup:
                                    new_filename = f'{product_number}_MOCKUP{file_ext}'
                                else:
                                    new_filename = f'{product_number}{file_ext}'
                                
                                # Create a new file-like object with the renamed filename
                                from django.core.files.base import ContentFile
                                file_content = file.read()
                                renamed_file = ContentFile(file_content, name=new_filename)
                                
                                # Create Media instance and set temp product_id as additional fallback
                                media_obj = Media(
                                    file=renamed_file,
                                    media_type='image' if file.content_type and file.content_type.startswith('image/') else 'video',
                                    created_by=request.user
                                )
                                # Set instance-level product_id as fallback
                                media_obj.set_temp_product_id(design.id)
                                # Save the instance
                                media_obj.save()
                                # #region agent log
                                try:
                                    with open(log_path, 'a') as f:
                                        f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"B","location":"Catalog/views.py:design_detail","message":"Media object created (design_detail view)","data":{"media_id":media_obj.id,"saved_path":media_obj.file.name,"product_id":design.id,"filename":new_filename},"timestamp":int(__import__('time').time()*1000)})+'\n')
                                except: pass
                                # #endregion
                                
                                # Validate file location - ensure it's in the correct product design folder
                                expected_path_prefix = f'{request.user.id}/designs/{design.id}/'
                                if not media_obj.file.name.startswith(expected_path_prefix):
                                    error_msg = f'Media file saved to wrong location! Expected: {expected_path_prefix}*, Got: {media_obj.file.name}'
                                    logger.error(error_msg)
                                    # #region agent log
                                    try:
                                        with open(log_path, 'a') as f:
                                            f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"B","location":"Catalog/views.py:design_detail","message":"VALIDATION ERROR: File in wrong location","data":{"media_id":media_obj.id,"saved_path":media_obj.file.name,"expected_prefix":expected_path_prefix,"product_id":design.id},"timestamp":int(__import__('time').time()*1000)})+'\n')
                                    except: pass
                                    # #endregion
                                
                                # Attach upload metadata
                                upload_metadata = {
                                    'timestamp': timezone.now().isoformat(),
                                    'original_filename': file.name,
                                    'file_size': file.size,
                                    'content_type': file.content_type if file.content_type else 'application/octet-stream',
                                    'is_mockup': is_mockup
                                }
                                
                                if is_mockup:
                                    upload_metadata['type'] = 'mockup'
                                
                                design.attach_media(media_obj, meta=upload_metadata, created_by=request.user)
                        finally:
                            # Clear product context
                            Media.clear_product_context()
                    
                    # Handle tags if provided
                    if 'tags' in request.data:
                        # Clear existing tags first
                        design.detach_all_tags()
                        
                        tag_ids = request.data.getlist('tags') if hasattr(request.data, 'getlist') else (
                            request.data['tags'] if isinstance(request.data['tags'], list) else [request.data['tags']]
                        )
                        for tag_id in tag_ids:
                            try:
                                tag = Tags.objects.get(id=tag_id)
                                design.attach_tag(tag, created_by=request.user)
                            except Tags.DoesNotExist:
                                pass
                    
                    return Response({
                        'message': 'Design updated successfully',
                        'design': ProductSerializer(design, context={'request': request}).data
                    })
                else:
                    # Return errors in a format that matches DRF and frontend expectations
                    error_response = {
                        'error': 'Validation failed',
                        'detail': 'Please check the field errors below'
                    }
                    error_response.update(serializer.errors)
                    return Response(error_response, status=status.HTTP_400_BAD_REQUEST)
                    
        except Exception as e:
            return Response({
                'error': f'Failed to update design: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    elif request.method == 'DELETE':
        # Only the design owner (created_by) can delete designs
        # Studio members cannot delete designs, even if they uploaded them
        if design.created_by != request.user:
            return Response({
                'error': 'You do not have permission to delete this design. Only the design owner can delete it.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check if design can be deleted
        if design.status == 'active':
            # TODO: Implement transfer ownership logic for approved designs
            # designs.services.transfer_ownership_to_wedesignz(design, request.user)
            return Response({
                'error': 'Cannot delete approved designs. Contact admin for assistance.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        design.delete()
        return Response({
            'message': 'Design deleted successfully'
        })


@swagger_auto_schema(
    method='get',
    operation_summary='Design Status',
    operation_description='Design Status endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def design_status(request, design_id):
    """
    Get design status, rejection reasons, and review information.
    """
    try:
        design = Product.objects.get(id=design_id, created_by=request.user)
        
        status_info = {
            'design_id': design.id,
            'title': design.title,
            'status': design.status,
            'created_at': design.created_at,
            'updated_at': design.updated_at,
            'rejection_reason': None,
            'admin_notes': None,
            'review_status': 'pending'
        }
        
        # TODO: Get rejection reasons and admin notes from review system
        # if design.status == 'rejected':
        #     review = DesignReview.objects.get(design=design)
        #     status_info['rejection_reason'] = review.rejection_reason
        #     status_info['admin_notes'] = review.admin_notes
        
        return Response(status_info)
        
    except Product.DoesNotExist:
        return Response({
            'error': 'Design not found'
        }, status=status.HTTP_404_NOT_FOUND)


@swagger_auto_schema(
    method='get',
    operation_summary='Pending Reviews Count',
    operation_description='Pending Reviews Count endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pending_reviews_count(request):
    """
    Get count of designs pending review.
    """
    pending_count = Product.objects.filter(
        created_by=request.user,
        status='draft'
    ).count()
    
    return Response({
        'pending_reviews_count': pending_count
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Design Analytics',
    operation_description='Design Analytics endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def design_analytics(request, design_id):
    """
    Get analytics for a specific design.
    """
    from Catalog.models import ProductCounter
    from Orders.models import Order, Cart
    from django.db.models import Sum, Q
    
    try:
        design = Product.objects.get(id=design_id, created_by=request.user)
        
        # Get all counters for this product
        counters = design.get_counters()
        
        # Calculate views (opened)
        views = counters.filter(product_counter_type='opened').count()
        
        # Calculate downloads
        downloads = counters.filter(product_counter_type='downloaded').count()
        
        # Calculate purchases and revenue
        # Get all successful orders containing this product
        purchases = 0
        revenue = 0.0
        
        # Get all successful orders (cart and subscription orders)
        orders = Order.objects.filter(
            status='success',
            order_type__in=['cart', 'subscription']
        ).exclude(product_ids__isnull=True).exclude(product_ids='')
        
        design_id_str = str(design.id)
        for order in orders:
            if order.product_ids:
                try:
                    # Parse product IDs from comma-separated string
                    order_product_ids = [pid.strip() for pid in order.product_ids.split(',') if pid.strip()]
                    # Check if design's ID is in this order's product IDs
                    if design_id_str in order_product_ids:
                        purchases += 1
                        # Calculate revenue: get the portion of total_amount for this product
                        # For simplicity, divide total_amount by number of items in order
                        # In a more complex system, you'd track individual item prices
                        num_items = len(order_product_ids)
                        if num_items > 0:
                            if design.price:
                                # Use product price if available
                                revenue += float(design.price)
                            else:
                                # Divide total amount equally among items
                                revenue += float(order.total_amount) / num_items
                except (ValueError, TypeError) as e:
                    # Skip orders with invalid product_ids format
                    continue
        
        # Also count purchased counters
        purchased_counters = counters.filter(product_counter_type='purchased').count()
        purchases = max(purchases, purchased_counters)
        
        # Calculate performance score
        # Formula: 0.2*views + 0.6*purchases + 0.2*downloads
        # Normalize to 0-10 scale for better readability
        # Use logarithmic scaling to handle large numbers
        import math
        
        # Normalize each metric to 0-10 scale using log scaling
        # This prevents any single metric from dominating
        def normalize_metric(value, base=10):
            if value <= 0:
                return 0
            # Use log scale: log(value + 1) / log(base + 1) * 10
            # This gives us a 0-10 scale that handles large numbers gracefully
            normalized = (math.log10(value + 1) / math.log10(base + 1)) * 10
            return min(normalized, 10)  # Cap at 10
        
        normalized_views = normalize_metric(views, 1000)
        normalized_downloads = normalize_metric(downloads, 100)
        normalized_purchases = normalize_metric(purchases, 50)
        
        # Apply weights: 0.2*views + 0.6*purchases + 0.2*downloads
        performance_score = (
            0.2 * normalized_views +
            0.6 * normalized_purchases +
            0.2 * normalized_downloads
        )
        
        analytics_data = {
            'design_id': design.id,
            'title': design.title,
            'views': views,
            'downloads': downloads,
            'purchases': purchases,
            'revenue': float(revenue),
            'performance_score': round(performance_score, 2)
        }
        
        return Response(analytics_data)
        
    except Product.DoesNotExist:
        return Response({
            'error': 'Design not found'
        }, status=status.HTTP_404_NOT_FOUND)


# ==================== PDF DOWNLOAD FUNCTIONALITY ====================

@swagger_auto_schema(
    method='get',
    operation_summary='Check Free Download Eligibility',
    operation_description='Check Free Download Eligibility endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def check_free_download_eligibility(request):
    """
    Check if user is eligible for free PDF download.
    """
    user = request.user
    
    # Check if user has already used their free download
    # Get all PDF downloads and filter by user via relations
    all_pdf_downloads = PDFDownload.objects.filter(download_type='free', status='completed')
    user_free_downloads = [pdf for pdf in all_pdf_downloads if pdf.get_user() == user]
    free_downloads = len(user_free_downloads)
    
    is_eligible = free_downloads == 0
    
    return Response({
        'is_eligible': is_eligible,
        'free_downloads_used': free_downloads,
        'free_pdf_designs_count': settings.PAID_PDF_DESIGNS_OPTIONS[0] if settings.PAID_PDF_DESIGNS_OPTIONS else 50,
        'message': 'You are eligible for a free download' if is_eligible else 'You have already used your free download'
    })


@swagger_auto_schema(
    method='post',
    operation_summary='Create Pdf Download Request',
    operation_description='Create Pdf Download Request endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_pdf_download_request(request):
    """
    Create a PDF download request (free or paid).
    Handles both specific product selection and search results.
    Now supports using subscription's mock PDF downloads.
    """
    from Plans.models import Subscription
    import logging
    logger = logging.getLogger(__name__)
    
    serializer = PDFDownloadRequestSerializer(data=request.data, context={'request': request})
    
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    data = serializer.validated_data
    user = request.user
    
    # Check for active subscription
    active_subscription = Subscription.objects.filter(
        created_by=user,
        status='active'
    ).select_related('plan').first()
    
    # Check if user wants to use subscription mock PDF (if download_type is 'free')
    use_subscription_mock_pdf = data.get('use_subscription_mock_pdf', False)
    
    # Validate total_pages based on download type
    # Get free design count from first value of PAID_PDF_DESIGNS_OPTIONS
    free_designs_count = settings.PAID_PDF_DESIGNS_OPTIONS[0] if settings.PAID_PDF_DESIGNS_OPTIONS else 50
    
    if data['download_type'] == 'free':
        # Free downloads (both regular and subscription) must use first value of PAID_PDF_DESIGNS_OPTIONS
        if data['total_pages'] != free_designs_count:
            return Response({
                'error': f'Free PDF downloads must contain exactly {free_designs_count} designs.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # If user wants to use subscription mock PDF
        if use_subscription_mock_pdf:
            if not active_subscription:
                return Response({
                    'error': 'No active subscription found. Cannot use subscription mock PDF download.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if not active_subscription.can_use_mock_pdf_download():
                return Response({
                    'error': f'No mock PDF downloads remaining. You have used {active_subscription.mock_pdf_downloads_used} of {active_subscription.plan.mock_pdf_count} available.',
                    'remaining_mock_pdf_downloads': active_subscription.get_remaining_mock_pdf_downloads(),
                    'total_mock_pdf_downloads': active_subscription.plan.mock_pdf_count
                }, status=status.HTTP_400_BAD_REQUEST)
        else:
            # Regular free download - check if user has already used their one-time free download
            from common.relations import get_related_ids_for_right
            pdf_download_ids = get_related_ids_for_right(user, 'User:PDFDownload')
            free_downloads = PDFDownload.objects.filter(
                id__in=pdf_download_ids,
                download_type='free',
                status='completed'
            ).count()
            
            if free_downloads > 0:
                return Response({
                    'error': 'You have already used your free download. Please use paid download for additional PDFs or use your subscription mock PDF downloads if available.'
                }, status=status.HTTP_400_BAD_REQUEST)
    else:
        # Paid downloads must use one of the configured options
        if data['total_pages'] not in settings.PAID_PDF_DESIGNS_OPTIONS:
            return Response({
                'error': f'Paid PDF downloads must use one of the following design counts: {", ".join(map(str, settings.PAID_PDF_DESIGNS_OPTIONS))}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Create PDF download record
        pdf_download = PDFDownload.objects.create(
            download_type=data['download_type'],
            total_pages=data['total_pages'],
            selection_type=data.get('selection_type', 'search_results'),
            selected_products=data.get('selected_products', []),
            search_filters=data.get('search_filters', {}),
            status='pending',
            products_count=0,
            included_products=[],
            customer_name=data.get('customer_name', ''),
            customer_mobile=data.get('customer_mobile', '')
        )
        
        # Attach user via relation system
        pdf_download.attach_user(user)
        
        # If using subscription mock PDF, mark it and set amount to 0
        if use_subscription_mock_pdf and active_subscription:
            pdf_download.price_per_design = 0.00
            pdf_download.total_amount = 0.00
        else:
            # Calculate pricing based on selection type using configurable values
            if data['download_type'] == 'paid':
                if data.get('selection_type') == 'specific' and data.get('selected_products'):
                    # Specific products selected - use configured price per design
                    pdf_download.price_per_design = settings.PAID_PDF_PRICE_PER_DESIGN_SELECTED
                    pdf_download.total_amount = len(data['selected_products']) * settings.PAID_PDF_PRICE_PER_DESIGN_SELECTED
                else:
                    # First N products from search - use configured price per design
                    pdf_download.price_per_design = settings.PAID_PDF_PRICE_PER_DESIGN_FIRSTN
                    pdf_download.total_amount = data['total_pages'] * settings.PAID_PDF_PRICE_PER_DESIGN_FIRSTN
            else:
                # Regular free download
                pdf_download.price_per_design = 0.00
                pdf_download.total_amount = 0.00
        
        pdf_download.save()
        
        # For free downloads (both regular free and subscription mock PDF), create order and process immediately
        if data['download_type'] == 'free':
            # Create order for free PDF download
            from Orders.models import Order
            product_ids_str = ','.join([str(pid) for pid in pdf_download.selected_products]) if pdf_download.selected_products else ''
            
            order = Order.objects.create(
                order_type='mock_pdf',
                product_ids=product_ids_str,
                total_amount=0.00,  # Free download
                status='success',  # Free downloads are automatically successful
                pdf_download=pdf_download,  # Link order to PDF download
                subscription=active_subscription if use_subscription_mock_pdf else None,  # Link subscription if using subscription mock PDF
                created_by=user
            )
            
            pdf_download.status = 'processing'
            pdf_download.save()
            
            # If using subscription mock PDF, increment counter immediately
            if use_subscription_mock_pdf and active_subscription:
                try:
                    active_subscription.use_mock_pdf_download()
                    logger.info(f"Used mock PDF download for subscription {active_subscription.id}. Remaining: {active_subscription.get_remaining_mock_pdf_downloads()}")
                except ValueError as e:
                    # This shouldn't happen as we checked above, but handle gracefully
                    logger.error(f"Failed to use mock PDF download for subscription {active_subscription.id}: {str(e)}")
            
            # Trigger PDF generation task
            generate_pdf_task.delay(pdf_download.id)
            
            return Response({
                'message': 'Free PDF download request created successfully',
                'download_id': pdf_download.id,
                'order_id': order.id,
                'status': 'processing',
                'estimated_completion': '5-10 minutes',
                'total_pages': pdf_download.total_pages,
                'selection_type': pdf_download.selection_type,
                'used_subscription_mock_pdf': use_subscription_mock_pdf,
                'remaining_mock_pdf_downloads': active_subscription.get_remaining_mock_pdf_downloads() if (use_subscription_mock_pdf and active_subscription) else None
            }, status=status.HTTP_201_CREATED)
        
        # For paid downloads, return payment information
        return Response({
            'message': 'Paid PDF download request created successfully',
            'download_id': pdf_download.id,
            'total_amount': float(pdf_download.total_amount),
            'price_per_design': float(pdf_download.price_per_design),
            'total_pages': pdf_download.total_pages,
            'selection_type': pdf_download.selection_type,
            'payment_required': True,
            'next_step': 'Make payment to proceed with PDF generation'
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return Response({
            'error': f'Failed to create PDF download request: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Get Pdf Download Status',
    operation_description='Get Pdf Download Status endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_pdf_download_status(request, download_id):
    """
    Get status of a PDF download request.
    """
    try:
        pdf_download = PDFDownload.objects.get(id=download_id)
        
        # Verify user owns this PDF download via relations system
        pdf_user = pdf_download.get_user()
        if pdf_user != request.user:
            return Response({
                'error': 'You do not have permission to view this PDF download'
            }, status=status.HTTP_403_FORBIDDEN)
        
        return Response({
            'download_id': pdf_download.id,
            'status': pdf_download.status,
            'download_type': pdf_download.download_type,
            'total_pages': pdf_download.total_pages,
            'total_amount': float(pdf_download.total_amount),
            'payment_status': pdf_download.payment_status,
            'created_at': pdf_download.created_at,
            'completed_at': pdf_download.completed_at,
            'pdf_file_path': pdf_download.pdf_file_path,
            'file_size': pdf_download.file_size
        })
        
    except PDFDownload.DoesNotExist:
        return Response({
            'error': 'PDF download not found'
        }, status=status.HTTP_404_NOT_FOUND)


@swagger_auto_schema(
    method='post',
    operation_summary='Add Products To Pdf',
    operation_description='Add Products To Pdf endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_products_to_pdf(request, download_id):
    """
    Add products to a PDF download.
    """
    try:
        pdf_download = PDFDownload.objects.get(id=download_id)
        # Verify user owns this PDF download via relations
        if pdf_download.get_user() != request.user:
            return Response({
                'error': 'You do not have permission to modify this PDF download'
            }, status=status.HTTP_403_FORBIDDEN)
        
        if pdf_download.status not in ['pending', 'processing']:
            return Response({
                'error': 'Cannot add products to a completed or failed download'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        products_data = request.data.get('products', [])
        if not products_data:
            return Response({
                'error': 'Products data is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Add products to PDF
        for product_data in products_data:
            product_id = product_data.get('product_id')
            page_number = product_data.get('page_number', 1)
            
            if not product_id:
                continue
                
            # Validate product exists
            try:
                product = Product.objects.get(
                    id=product_id,
                    status='active',
                    visibility_status='show'
                )
                pdf_download.add_product_to_pdf(product_id, page_number)
            except Product.DoesNotExist:
                continue
        
        return Response({
            'message': 'Products added to PDF successfully',
            'download_id': pdf_download.id,
            'products_count': pdf_download.products_count,
            'included_products': pdf_download.get_included_products()
        })
        
    except PDFDownload.DoesNotExist:
        return Response({
            'error': 'PDF download not found'
        }, status=status.HTTP_404_NOT_FOUND)


@swagger_auto_schema(
    method='get',
    operation_summary='List User Pdf Downloads',
    operation_description='List User Pdf Downloads endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_user_pdf_downloads(request):
    """
    List all PDF downloads for the authenticated user with filtering.
    """
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    download_type = request.GET.get('download_type')  # 'free', 'paid', or None for all
    status_filter = request.GET.get('status')  # 'pending', 'processing', 'completed', 'failed', or None for all
    
    # Get all PDF downloads and filter by user via relations
    all_downloads = PDFDownload.objects.all().order_by('-created_at')
    downloads = [pdf for pdf in all_downloads if pdf.get_user() == request.user]
    
    # Apply filters
    if download_type in ['free', 'paid']:
        downloads = [pdf for pdf in downloads if pdf.download_type == download_type]
    
    if status_filter in ['pending', 'processing', 'completed', 'failed']:
        downloads = [pdf for pdf in downloads if pdf.status == status_filter]
    
    # Pagination
    paginator = Paginator(downloads, page_size)
    page_obj = paginator.get_page(page)
    
    return Response({
        'downloads': PDFDownloadListSerializer(page_obj.object_list, many=True).data,
        'total_pages': paginator.num_pages,
        'current_page': page,
        'total_count': paginator.count
    })


@swagger_auto_schema(
    method='post',
    operation_summary='Process Pdf Payment',
    operation_description='Process Pdf Payment endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def process_pdf_payment(request):
    """
    Process payment for a paid PDF download.
    """
    serializer = PDFDownloadPaymentSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    data = serializer.validated_data
    
    try:
        pdf_download = PDFDownload.objects.get(
            id=data['download_id'],
            download_type='paid',
            payment_status='pending'
        )
        
        # Verify user owns this PDF download via relations system
        pdf_user = pdf_download.get_user()
        if pdf_user != request.user:
            return Response({
                'error': 'You do not have permission to process payment for this PDF download'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # TODO: Verify Razorpay payment
        # razorpay_payment = verify_razorpay_payment(data['razorpay_payment_id'])
        
        # For now, simulate successful payment
        pdf_download.payment_status = 'paid'
        pdf_download.status = 'processing'
        pdf_download.save()
        
        # Trigger PDF generation task
        generate_pdf_task.delay(pdf_download.id)
        
        return Response({
            'message': 'Payment processed successfully',
            'download_id': pdf_download.id,
            'status': 'processing',
            'estimated_completion': '5-10 minutes'
        })
        
    except PDFDownload.DoesNotExist:
        return Response({
            'error': 'PDF download not found or not eligible for payment'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            'error': f'Payment processing failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Download Pdf File',
    operation_description='Download Pdf File endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_pdf_file(request, download_id):
    """
    Download the generated PDF file.
    """
    try:
        pdf_download = PDFDownload.objects.get(id=download_id)
        
        # Verify user owns this PDF download via relations system
        pdf_user = pdf_download.get_user()
        if pdf_user != request.user:
            return Response({
                'error': 'You do not have permission to download this PDF'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check if PDF is still processing
        if pdf_download.status == 'processing':
            return Response({
                'error': 'PDF is still being generated. Please try again in a few moments.',
                'status': 'processing'
            }, status=status.HTTP_202_ACCEPTED)
        
        if pdf_download.status != 'completed':
            return Response({
                'error': f'PDF is not ready for download. Current status: {pdf_download.status}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not pdf_download.pdf_file_path:
            # If file path is not set but status is completed, try to regenerate it
            # This handles cases where the task completed but file path wasn't saved
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f'PDF download {download_id} has no file path but status is completed. Attempting to locate file.')
            
            # Try to find the file - check both new and old locations
            from django.http import FileResponse
            import os
            expected_filename = f'pdf_download_{download_id}.pdf'
            
            # Try new location first (user-specific)
            user = pdf_download.get_user()
            if user:
                user_id = user.id
                pdf_dir = os.path.join(settings.MEDIA_ROOT, str(user_id), 'pdfs')
                expected_path = os.path.join(pdf_dir, expected_filename)
                if os.path.exists(expected_path):
                    pdf_download.pdf_file_path = f'{user_id}/pdfs/{expected_filename}'
                    pdf_download.file_size = os.path.getsize(expected_path)
                    pdf_download.save()
                else:
                    # Try old location as fallback
                    pdf_dir = os.path.join(settings.MEDIA_ROOT, 'pdfs')
                    expected_path = os.path.join(pdf_dir, expected_filename)
                    if os.path.exists(expected_path):
                        pdf_download.pdf_file_path = f'pdfs/{expected_filename}'
                        pdf_download.file_size = os.path.getsize(expected_path)
                        pdf_download.save()
                    else:
                        return Response({
                            'error': 'PDF file not available. The file may not have been generated yet.'
                        }, status=status.HTTP_404_NOT_FOUND)
            else:
                # No user found, try old location
                pdf_dir = os.path.join(settings.MEDIA_ROOT, 'pdfs')
                expected_path = os.path.join(pdf_dir, expected_filename)
                if os.path.exists(expected_path):
                    pdf_download.pdf_file_path = f'pdfs/{expected_filename}'
                    pdf_download.file_size = os.path.getsize(expected_path)
                    pdf_download.save()
                else:
                    return Response({
                        'error': 'PDF file not available. The file may not have been generated yet.'
                    }, status=status.HTTP_404_NOT_FOUND)
        
        # Implement file download logic
        from django.http import FileResponse
        import os
        import logging
        logger = logging.getLogger(__name__)
        
        file_path = os.path.join(settings.MEDIA_ROOT, pdf_download.pdf_file_path)
        logger.info(f'Attempting to download PDF from: {file_path}')
        logger.info(f'MEDIA_ROOT: {settings.MEDIA_ROOT}')
        logger.info(f'pdf_file_path from DB: {pdf_download.pdf_file_path}')
        logger.info(f'Full file path: {file_path}')
        logger.info(f'File exists check: {os.path.exists(file_path)}')
        
        if os.path.exists(file_path):
            try:
                # Generate filename using customer name if available
                import re
                customer_name = pdf_download.customer_name or ''
                logger.info(f'Download request for PDF {download_id}: customer_name="{customer_name}"')
                if customer_name:
                    # Sanitize customer name for filename (remove special characters, keep spaces)
                    sanitized_name = re.sub(r'[^a-zA-Z0-9\s-]', '', customer_name)
                    sanitized_name = re.sub(r'\s+', ' ', sanitized_name.strip())  # Keep spaces, just trim
                    sanitized_name = sanitized_name[:50]  # Limit length
                    filename = f'{sanitized_name}.pdf'  # Remove _mock_pdf suffix
                    logger.info(f'Generated filename from customer name: "{filename}"')
                else:
                    filename = f'designs_{download_id}.pdf'
                    logger.info(f'No customer name found, using default filename: "{filename}"')
                
                response = FileResponse(
                    open(file_path, 'rb'),
                    as_attachment=True,
                    filename=filename,
                    content_type='application/pdf'
                )
                # Explicitly set Content-Disposition header with proper encoding for filenames with spaces
                # Use both filename and filename* (RFC 5987) for better browser compatibility
                from urllib.parse import quote
                encoded_filename = quote(filename)
                response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
                # Also set a custom header that's easier to access (will be exposed via CORS)
                response['X-Filename'] = filename
                logger.info(f'Setting Content-Disposition header: {response["Content-Disposition"]}')
                logger.info(f'Setting X-Filename header: {filename}')
                return response
            except Exception as e:
                logger.error(f'Error opening PDF file: {str(e)}', exc_info=True)
                return Response({
                    'error': f'Error reading PDF file: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        else:
            logger.error(f'PDF file not found at path: {file_path}')
            # Try alternative paths (check new location first, then old location)
            user = pdf_download.get_user()
            alt_path = None
            if user:
                # Try new user-specific location
                alt_path = os.path.join(settings.MEDIA_ROOT, str(user.id), 'pdfs', f'pdf_download_{download_id}.pdf')
                if not os.path.exists(alt_path):
                    # Fallback to old location
                    alt_path = os.path.join(settings.MEDIA_ROOT, 'pdfs', f'pdf_download_{download_id}.pdf')
            else:
                # No user, try old location
                alt_path = os.path.join(settings.MEDIA_ROOT, 'pdfs', f'pdf_download_{download_id}.pdf')
            
            if alt_path and os.path.exists(alt_path):
                try:
                    # Generate filename using customer name if available
                    import re
                    customer_name = pdf_download.customer_name or ''
                    if customer_name:
                        # Sanitize customer name for filename (remove special characters, keep spaces)
                        sanitized_name = re.sub(r'[^a-zA-Z0-9\s-]', '', customer_name)
                        sanitized_name = re.sub(r'\s+', ' ', sanitized_name.strip())  # Keep spaces, just trim
                        sanitized_name = sanitized_name[:50]  # Limit length
                        filename = f'{sanitized_name}.pdf'  # Remove _mock_pdf suffix
                    else:
                        filename = f'designs_{download_id}.pdf'
                    
                    response = FileResponse(
                        open(alt_path, 'rb'),
                        as_attachment=True,
                        filename=filename,
                        content_type='application/pdf'
                    )
                    # Explicitly set Content-Disposition header with proper encoding for filenames with spaces
                    from urllib.parse import quote
                    encoded_filename = quote(filename)
                    response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
                    # Also set a custom header that's easier to access (will be exposed via CORS)
                    response['X-Filename'] = filename
                    logger.info(f'Setting Content-Disposition header (alt path): {response["Content-Disposition"]}')
                    logger.info(f'Setting X-Filename header (alt path): {filename}')
                    # Update database with correct path based on location found
                    user = pdf_download.get_user()
                    if user and alt_path.startswith(os.path.join(settings.MEDIA_ROOT, str(user.id))):
                        pdf_download.pdf_file_path = f'{user.id}/pdfs/pdf_download_{download_id}.pdf'
                    else:
                        pdf_download.pdf_file_path = f'pdfs/pdf_download_{download_id}.pdf'
                    pdf_download.save()
                    return response
                except Exception as e:
                    logger.error(f'Error opening PDF file from alt path: {str(e)}', exc_info=True)
            
            # If file doesn't exist but status is completed, try to regenerate it
            # This handles cases where the task completed but file wasn't created
            if pdf_download.status == 'completed':
                logger.warning(f'PDF file missing for completed download {download_id}. Attempting to regenerate.')
                try:
                    from .tasks import generate_pdf_task
                    # Trigger task asynchronously to regenerate the file
                    generate_pdf_task.delay(download_id)
                    return Response({
                        'error': 'PDF file is being regenerated. Please try again in a few moments.',
                        'status': 'regenerating',
                        'download_id': download_id
                    }, status=status.HTTP_202_ACCEPTED)
                except Exception as e:
                    logger.error(f'Error triggering PDF regeneration: {str(e)}', exc_info=True)
                    return Response({
                        'error': 'PDF file not found and regeneration failed. Please contact support.',
                        'details': str(e)
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            return Response({
                'error': 'PDF file not found on server. The file may not have been generated yet.',
                'file_path': file_path,
                'status': pdf_download.status,
                'suggestion': 'Please try again in a few moments or contact support if the issue persists.'
            }, status=status.HTTP_404_NOT_FOUND)
        
    except PDFDownload.DoesNotExist:
        return Response({
            'error': 'PDF download not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error downloading PDF {download_id}: {str(e)}', exc_info=True)
        return Response({
            'error': f'An error occurred while downloading the PDF: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Get Pdf Configuration',
    operation_description='Get PDF download configuration including design counts and pricing',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)
@api_view(['GET'])
@permission_classes([AllowAny])
def get_pdf_config(request):
    """
    Get PDF download configuration (design counts and pricing).
    This endpoint is public and can be accessed without authentication.
    """
    return Response({
        'free_pdf_designs_count': settings.PAID_PDF_DESIGNS_OPTIONS[0] if settings.PAID_PDF_DESIGNS_OPTIONS else 50,
        'paid_pdf_designs_options': settings.PAID_PDF_DESIGNS_OPTIONS,
        'pricing': {
            'first_n_per_design': float(settings.PAID_PDF_PRICE_PER_DESIGN_FIRSTN),
            'selected_per_design': float(settings.PAID_PDF_PRICE_PER_DESIGN_SELECTED)
        }
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Get Pdf Pricing Info',
    operation_description='Get Pdf Pricing Info endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_pdf_pricing_info(request):
    """
    Get pricing information for PDF downloads.
    """
    return Response({
        'pricing': {
            'free_download': {
                'available': True,
                'description': 'One free download per user (any page count)',
                'price': 0.00,
                'currency': 'INR'
            },
            'paid_downloads': {
                'specific_products': {
                    'price_per_design': 4.00,
                    'description': 'Select specific designs of your choice (Rs. 4 per design)',
                    'currency': 'INR',
                    'example': '100 specific designs = Rs. 400'
                },
                'search_results': {
                    'price_per_design': 2.00,
                    'description': 'First N designs from search results (Rs. 2 per design)',
                    'currency': 'INR',
                    'example': '100 designs from search = Rs. 200'
                }
            },
            'page_options': [50, 100, 200, 300, 500],
            'max_pages': 500,
            'supported_formats': ['PDF'],
            'browsing': {
                'designs_per_page': BusinessConfig.get_minimum_required_designs_onboard(),
                'description': f'Browse designs with {BusinessConfig.get_minimum_required_designs_onboard()} designs per page for easy selection'
            }
        }
    })


@swagger_auto_schema(
    method='post',
    operation_summary='Search Products For Pdf',
    operation_description='Search Products For Pdf endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def search_products_for_pdf(request):
    """
    Search products with filters for PDF generation.
    Returns exactly the configured number of designs per page for easy browsing.
    """
    query = request.data.get('query', '')
    category_id = request.data.get('category_id')
    tags = request.data.get('tags', [])
    color = request.data.get('color')
    min_price = request.data.get('min_price')
    max_price = request.data.get('max_price')
    page = int(request.data.get('page', 1))
    page_size = BusinessConfig.get_minimum_required_designs_onboard()  # Fixed at configured designs per page as per requirements
    
    # Base queryset
    products = Product.objects.filter(
        status='active',
        visibility_status='show'
    )
    
    # Apply filters
    if query:
        products = products.filter(
            Q(title__icontains=query) | 
            Q(description__icontains=query) |
            Q(product_number__icontains=query) |
            Q(studio_design_number__icontains=query)
        )
    
    if category_id:
        products = products.filter(category_id=category_id)
    
    if tags:
        products = products.filter(tags__id__in=tags).distinct()
    
    if color:
        products = products.filter(color__icontains=color)
    
    if min_price is not None:
        products = products.filter(price__gte=min_price)
    
    if max_price is not None:
        products = products.filter(price__lte=max_price)
    
    # Order by creation date
    products = products.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(products, page_size)
    page_obj = paginator.get_page(page)
    
    return Response({
        'products': ProductSerializer(page_obj.object_list, many=True, context={'request': request}).data,
        'total_pages': paginator.num_pages,
        'current_page': page,
        'total_count': paginator.count,
        'page_size': page_size,
        'filters_applied': {
            'query': query,
            'category_id': category_id,
            'tags': tags,
            'color': color,
            'min_price': min_price,
            'max_price': max_price
        }
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Browse Designs Catalog',
    operation_description='Browse Designs Catalog endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([AllowAny])
def browse_designs_catalog(request):
    """
    Browse design catalog with advanced filters for PDF generation.
    Returns configured number of designs per page with pagination.
    """
    query = request.GET.get('q', '')
    category_id = request.GET.get('category')
    tags = request.GET.getlist('tags')
    color = request.GET.get('color')
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')
    page = int(request.GET.get('page', 1))
    page_size = BusinessConfig.get_minimum_required_designs_onboard()  # Fixed at configured designs per page
    
    # Base queryset
    products = Product.objects.filter(
        status='active',
        visibility_status='show'
    )
    
    # Apply search query
    if query:
        products = products.filter(
            Q(title__icontains=query) | Q(description__icontains=query)
        )
    
    # Apply filters
    if category_id:
        products = products.filter(category_id=category_id)
    
    if tags:
        products = products.filter(tags__id__in=tags).distinct()
    
    if color:
        products = products.filter(color__icontains=color)
    
    if min_price is not None:
        try:
            products = products.filter(price__gte=float(min_price))
        except ValueError:
            pass
    
    if max_price is not None:
        try:
            products = products.filter(price__lte=float(max_price))
        except ValueError:
            pass
    
    # Order by creation date
    products = products.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(products, page_size)
    page_obj = paginator.get_page(page)
    
    return Response({
        'products': ProductSerializer(page_obj.object_list, many=True, context={'request': request}).data,
        'total_pages': paginator.num_pages,
        'current_page': page,
        'total_count': paginator.count,
        'page_size': page_size,
        'has_next': page_obj.has_next(),
        'has_previous': page_obj.has_previous(),
        'filters_applied': {
            'query': query,
            'category_id': category_id,
            'tags': tags,
            'color': color,
            'min_price': min_price,
            'max_price': max_price
        }
    })
