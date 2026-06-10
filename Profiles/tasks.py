from celery import shared_task
from django.db import transaction, connection
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.conf import settings
import zipfile
import io
import os
import logging
import json
from openpyxl import load_workbook

from .models import DesignProcessingTask, DesignerProfile, Studio, StudioMember
from Catalog.models import Product, Category, Tags
from Plans.models import Plan
from MediaFiles.models import Media
from common.avif_converter import create_avif_from_media_file

logger = logging.getLogger(__name__)

@shared_task(bind=True, name='Profiles.tasks.process_design_upload_task')
def process_design_upload_task(self, task_id, zip_file_path):
    """
    Celery task to process design upload asynchronously.
    
    Args:
        task_id: ID of the DesignProcessingTask record
        zip_file_path: Path to the stored zip file
    """
    logger.info(f"process_design_upload_task: starting for task_id={task_id}")
    # Ensure clean DB connection (previous task e.g. cleanup_design_pdf_files may have left it aborted)
    try:
        connection.close()
    except Exception:
        pass
    try:
        # Get the task record
        try:
            task = DesignProcessingTask.objects.get(id=task_id)
        except DesignProcessingTask.DoesNotExist:

            # Task doesn't exist, can't proceed - raise error to mark Celery task as failed
            raise ValueError(f"DesignProcessingTask with id={task_id} does not exist. Cannot process design upload.")
        task.status = 'processing'
        task.save(update_fields=['status', 'updated_at'])

        # Determine product owner and track if uploaded by studio member
        # For studio members, owner should be studio owner (not the member)
        # For studio owners or individual designers, owner is themselves
        product_owner = task.user
        uploaded_by_member_id = None
        
        studio_membership = StudioMember.objects.filter(
            member=task.user,
            status='active'
        ).select_related('studio').first()
        
        if studio_membership:
            # User is a studio member - owner should be studio owner
            studio = studio_membership.studio
            product_owner = studio.created_by
            uploaded_by_member_id = task.user.id

        else:
            pass

        # Use the path from the database record as the source of truth
        # The parameter might be outdated, but the database record is always current
        actual_zip_file_path = task.zip_file_path

        # If database path differs from parameter, log a warning
        if actual_zip_file_path != zip_file_path:
            pass

        # Read zip file from storage

        # Check if file exists in storage
        _zip_exists = default_storage.exists(actual_zip_file_path)
        if not _zip_exists:
            # Try to construct absolute path as fallback for debugging
            absolute_path = None
            try:
                if hasattr(settings, 'MEDIA_ROOT') and settings.MEDIA_ROOT:
                    absolute_path = os.path.join(settings.MEDIA_ROOT, actual_zip_file_path)

                    if os.path.exists(absolute_path):
                        pass
            except Exception as path_error:
                pass

            # Log available storage information for debugging

            if absolute_path:
                pass

            raise FileNotFoundError(
                f"Zip file not found at path: {actual_zip_file_path}. "
                f"MEDIA_ROOT: {getattr(settings, 'MEDIA_ROOT', 'Not set')}. "
                f"Storage backend: {type(default_storage).__name__}"
            )

        zip_file = default_storage.open(actual_zip_file_path, 'rb')
        zip_content = zip_file.read()
        zip_file.close()

        zip_buffer = io.BytesIO(zip_content)
        
        with zipfile.ZipFile(zip_buffer, 'r') as zip_ref:
            all_files = zip_ref.namelist()

            # Find metadata.xlsx

            metadata_file = None
            for file_name in all_files:
                lower_name = file_name.lower()
                if lower_name == 'metadata.xlsx' or lower_name.endswith('/metadata.xlsx'):
                    metadata_file = file_name
                    break
            
            if not metadata_file:
                raise ValueError("metadata.xlsx file not found in zip")

            # Parse metadata.xlsx
            # Fixed column positions (metadata.xlsx format):
            # Column 1: folder_name
            # Column 2: title
            # Column 3: description
            # Column 4: category
            # Column 5: subcategory
            # Column 6: tags
            # Note: plan, color, and visible are no longer in metadata - defaults applied:
            # - plan: defaults to '4' (premium)
            # - visible: defaults to '1' (show)
            # - color: defaults to None
            # Note: Price is now managed globally via SystemConfig, not from metadata

            metadata_data = zip_ref.read(metadata_file)
            metadata_buffer = io.BytesIO(metadata_data)
            workbook = load_workbook(metadata_buffer, data_only=True)
            sheet = workbook.active

            # Create metadata mapping - start from row 2 (row 1 is headers)

            metadata_dict = {}
            row_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=False):  # Start from row 2
                # Column indices (0-based): 0=folder_name, 1=title, 2=description, 3=category, 4=subcategory, etc.
                folder_name_cell = row[0] if len(row) > 0 else None
                
                if folder_name_cell and folder_name_cell.value:
                    folder_name = str(folder_name_cell.value).strip()
                    if folder_name:
                        row_data = {
                            'folder_name': folder_name,
                            'title': str(row[1].value).strip() if len(row) > 1 and row[1].value else '',
                            'description': str(row[2].value).strip() if len(row) > 2 and row[2].value else '',
                            'category': str(row[3].value).strip() if len(row) > 3 and row[3].value else '',
                            'subcategory': str(row[4].value).strip() if len(row) > 4 and row[4].value else '',
                            'tags': str(row[5].value).strip() if len(row) > 5 and row[5].value else '',  # Tags now at column 5
                        }
                        metadata_dict[folder_name] = row_data
                        row_count += 1
                        # Log first entry for debugging
                        if row_count == 1:
                            pass

            # Find root folder - handle both 2-part and 3-part paths
            # 2-part: design_folder/file.ext (zip created from inside root folder)
            # 3-part: root_folder/design_folder/file.ext (normal case)

            root_folder = ''
            if '/' in metadata_file:
                # metadata.xlsx is in a subfolder, use that folder as root
                root_folder = metadata_file.rsplit('/', 1)[0]  # Get parent folder

            else:
                # metadata.xlsx is in root, find root folder from valid design folders
                SYSTEM_FOLDERS = ['__macosx', '.ds_store', 'rar', '.rar', 'thumbs.db']
                for file_name in all_files:
                    if '/' in file_name and not file_name.endswith('/'):
                        parts = file_name.split('/')
                        if len(parts) >= 3:
                            root_folder_name = parts[0].lower()
                            if root_folder_name not in SYSTEM_FOLDERS:
                                root_folder = parts[0]
                                break
                        elif len(parts) == 2:
                            # 2-part path means no root folder (zip created from inside root)
                            root_folder = ''
                            break

            # Get valid design folders - handle both 2-part and 3-part paths (matching validation logic)

            SYSTEM_FOLDERS = ['__macosx', '.ds_store', 'rar', '.rar', 'thumbs.db']
            required_files = {'.eps', '.cdr', '.jpg', '.png'}
            valid_design_folders = {}
            
            for file_name in all_files:
                if file_name == metadata_file or file_name.endswith('/'):
                    continue
                
                if '/' in file_name:
                    parts = file_name.split('/')
                    
                    # Handle both cases:
                    # 1. root_folder/design_folder/file.ext (3 parts) - e.g., "123/WD01/WD01.eps"
                    # 2. design_folder/file.ext (2 parts) - e.g., "WD01/WD01.eps" (if zip created from inside root folder)
                    
                    if len(parts) < 2:
                        # Skip files at root level (not in any folder)
                        continue
                    
                    # Determine folder name and file info based on path length
                    if len(parts) == 2:
                        # Case 2: design_folder/file.ext (zip created from inside root folder)
                        folder_name = parts[0]
                        file_name_only = parts[1]
                    elif len(parts) == 3:
                        # Case 1: root_folder/design_folder/file.ext (normal case)
                        root_folder_name = parts[0].lower()
                        folder_name = parts[1]
                        file_name_only = parts[2]
                        
                        # Skip system folders
                        if root_folder_name in SYSTEM_FOLDERS or folder_name.lower() in SYSTEM_FOLDERS:
                            continue
                    else:
                        # More than 3 parts - skip for now
                        continue
                    
                    file_ext = os.path.splitext(file_name_only)[1].lower()
                    file_name_lower = file_name_only.lower()
                    
                    # Check if it's an optional mockup file
                    is_mockup = file_name_lower == 'mockup.jpg' or file_name_lower == 'mockup.png'
                    
                    # Skip system folders (check folder name)
                    if folder_name.lower() in SYSTEM_FOLDERS:
                        continue
                    
                    # Process files in design folders
                    if len(parts) == 2 or len(parts) == 3:
                        if file_ext in required_files or is_mockup:
                            if folder_name not in valid_design_folders:
                                valid_design_folders[folder_name] = set()
                            # Only add required files to the set (mockup is optional)
                            if file_ext in required_files:
                                valid_design_folders[folder_name].add(file_ext)
            
            # Filter folders with all required files
            valid_folders = {}
            for folder_name, files in valid_design_folders.items():
                if all(ext in files for ext in required_files):
                    valid_folders[folder_name] = files

            # Process each design folder
            processed_count = 0
            failed_count = 0
            failed_details = []
            
            # Set total designs count
            task.total_designs = len(valid_folders)
            task.save(update_fields=['total_designs', 'updated_at'])

            for idx, folder_name in enumerate(valid_folders.keys(), 1):

                try:
                    # Force a fresh DB connection so we never run on an aborted connection (e.g. from another task)
                    try:
                        connection.close()
                    except Exception:
                        pass
                    try:
                        transaction.rollback()
                    except Exception:
                        pass

                    # Process each design in its own transaction
                    with transaction.atomic():
                        # Get metadata for this folder
                        metadata = metadata_dict.get(folder_name, {})

                        # Extract files from zip - handle both 2-part and 3-part paths
                        if root_folder:
                            folder_path = f"{root_folder}/{folder_name}/"
                        else:
                            folder_path = f"{folder_name}/"  # Handle 2-part paths

                        design_files = {}
                        mockup_file = None
                        for file_name in all_files:
                            # Handle both 2-part and 3-part paths
                            if root_folder:
                                expected_path = f"{root_folder}/{folder_name}/"
                            else:
                                expected_path = f"{folder_name}/"
                            
                            if file_name.startswith(expected_path) and not file_name.endswith('/'):
                                file_name_only = os.path.basename(file_name)
                                file_name_lower = file_name_only.lower()
                                file_ext = os.path.splitext(file_name)[1].lower()
                                
                                # Check for required design files
                                if file_ext in ['.eps', '.cdr', '.jpg', '.png']:
                                    # Check if it's a mockup file (case insensitive)
                                    is_mockup = file_name_lower == 'mockup.jpg' or file_name_lower == 'mockup.png'
                                    if is_mockup:
                                        mockup_file = file_name
                                    else:
                                        design_files[file_ext] = file_name

                        if mockup_file:
                            pass

                        # Get category and subcategory from metadata (using fixed column positions)
                        category_name = metadata.get('category', '').strip() if metadata.get('category') else ''
                        if not category_name:
                            category_name = 'other'
                        else:
                            pass

                        subcategory_name = metadata.get('subcategory', '').strip() if metadata.get('subcategory') else ''

                        # Create or get parent category (parent=None means it's a top-level category)
                        # Use nested atomic so a failed create rolls back and doesn't abort the outer transaction
                        parent_category = None
                        try:
                            with transaction.atomic():
                                parent_category = Category.objects.filter(name=category_name, parent__isnull=True).first()
                                if not parent_category:
                                    parent_category = Category.objects.create(
                                        name=category_name,
                                        parent=None,
                                        created_by=task.user
                                    )
                        except Exception:
                            try:
                                with transaction.atomic():
                                    parent_category = Category.objects.filter(name='other', parent__isnull=True).first()
                                    if not parent_category:
                                        parent_category = Category.objects.create(
                                            name='other',
                                            parent=None,
                                            created_by=task.user
                                        )
                            except Exception:
                                raise ValueError(f"Failed to get or create parent category: '{category_name}' or fallback 'other'")
                        
                        if not parent_category:
                            raise ValueError(f"Failed to get or create parent category: '{category_name}'")
                        
                        # Create or get subcategory if specified (nested atomic so failure doesn't abort outer transaction)
                        if subcategory_name:
                            try:
                                with transaction.atomic():
                                    category = Category.objects.filter(
                                        name=subcategory_name,
                                        parent=parent_category
                                    ).first()
                                    if not category:
                                        category = Category.objects.create(
                                            name=subcategory_name,
                                            parent=parent_category,
                                            created_by=task.user
                                        )
                            except Exception:
                                category = parent_category
                        else:
                            category = parent_category

                        # Plan default: 4 (premium) - removed from metadata template
                        plan_value = '4'  # Default to premium plan
                        plan_mapping = {
                            '0': 'free',
                            '1': 'basic',
                            '2': 'prime',
                            '3': 'premium',
                            '4': 'premium',
                            '9': 'basic'
                        }
                        product_plan_type = plan_mapping.get(str(plan_value).strip(), 'premium')  # Default to premium
                        
                        # Price is now managed globally via SystemConfig
                        # Use global design price for paid designs, None for free designs
                        from common.business_config import BusinessConfig
                        if product_plan_type == 'free':
                            price = None
                        else:
                            price = BusinessConfig.get_design_price()
                        
                        # Visibility default: 1 (show) - removed from metadata template
                        visible_value = '1'  # Always default to visible
                        visibility_status = 'show'
                        
                        # Color default: None - removed from metadata template
                        color = None  # Color column removed, always set to None
                        
                        # Get title and description from metadata (using fixed column positions)
                        title = metadata.get('title', '').strip()[:200] if metadata.get('title') else f"Design {folder_name}"
                        if not title:
                            title = f"Design {folder_name}"
                        
                        description = metadata.get('description', '').strip() if metadata.get('description') else f"Design from folder {folder_name}"
                        if not description:
                            description = f"Design from folder {folder_name}"

                        # Create Product
                        try:

                            # Build product_metadata with bulk upload info and member tracking
                            product_metadata = {
                                'source': 'bulk_upload',
                                'folder_name': folder_name
                            }
                            if uploaded_by_member_id:
                                product_metadata['uploaded_by_member_id'] = uploaded_by_member_id
                            
                            product = Product.objects.create(
                                title=title,
                                description=description,
                                category=category,
                                product_plan_type=product_plan_type,
                                color=color,
                                price=price,
                                visibility_status=visibility_status,
                                status='draft',
                                created_by=product_owner,  # Use studio owner if member uploaded
                                product_metadata=product_metadata  # Include member tracking
                            )

                        except Exception as e:

                            raise
                        
                        # Create Media instances

                        # Ensure product_number is available
                        if not product.product_number:

                            product.refresh_from_db()
                            if not product.product_number:

                                raise ValueError(f"Product {product.id} has no product_number")
                        
                        product_number = product.product_number
                        
                        # Set product context for file path generation
                        # #region agent log
                        log_path = os.getenv('DEBUG_LOG_PATH', os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs', 'debug.log'))
                        try:
                            with open(log_path, 'a') as f:
                                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"B","location":"Profiles/tasks.py:process_design_upload_task","message":"Setting product context before Media.create (bulk upload)","data":{"product_id":product.id,"task_id":task_id},"timestamp":int(__import__('time').time()*1000)})+'\n')
                        except: pass
                        # #endregion
                        Media.set_product_context(product.id)
                        try:
                            for file_ext, file_path in design_files.items():
                                try:
                                    with transaction.atomic():
                                        file_data = zip_ref.read(file_path)
                                        file_name = os.path.basename(file_path)
                                        media_type = 'image'
                                        new_filename = f'{product_number}{file_ext}'
                                        media_file = ContentFile(file_data, name=new_filename)
                                        media = Media(
                                            file=media_file,
                                            media_type=media_type,
                                            created_by=product_owner
                                        )
                                        media.set_temp_product_id(product.id)
                                        media.save()
                                        # #region agent log
                                        try:
                                            with open(log_path, 'a') as f:
                                                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"B","location":"Profiles/tasks.py:process_design_upload_task","message":"Media object created (bulk upload)","data":{"media_id":media.id,"saved_path":media.file.name,"product_id":product.id,"filename":new_filename},"timestamp":int(__import__('time').time()*1000)})+'\n')
                                        except: pass
                                        # #endregion
                                        expected_path_prefix = f'{product_owner.id}/designs/{product.id}/'
                                        if not media.file.name.startswith(expected_path_prefix):
                                            try:
                                                with open(log_path, 'a') as f:
                                                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"B","location":"Profiles/tasks.py:process_design_upload_task","message":"VALIDATION ERROR: File in wrong location","data":{"media_id":media.id,"saved_path":media.file.name,"expected_prefix":expected_path_prefix,"product_id":product.id,"task_id":task_id},"timestamp":int(__import__('time').time()*1000)})+'\n')
                                            except: pass
                                        meta_info = {
                                            'type': file_ext[1:].upper(),
                                            'folder_name': folder_name,
                                            'original_filename': file_name
                                        }
                                        product.attach_media(media, meta=meta_info, created_by=task.user)
                                        if file_ext in ['.jpg', '.jpeg', '.png']:
                                            try:
                                                media_file_path = media.file.name
                                                avif_path, avif_media_obj = create_avif_from_media_file(
                                                    media_file_path, product_number, is_mockup=False,
                                                    product=product, created_by=product_owner
                                                )
                                            except Exception as avif_error:
                                                logger.warning(
                                                    "AVIF conversion failed for product_id=%s file=%s: %s",
                                                    product.id,
                                                    media.file.name,
                                                    avif_error,
                                                )
                                except Exception:
                                    pass

                        finally:
                            # Clear product context
                            Media.clear_product_context()
                        
                        # Process optional mockup file if present
                        if mockup_file:
                            Media.set_product_context(product.id)
                            try:
                                with transaction.atomic():
                                    file_data = zip_ref.read(mockup_file)
                                    file_name = os.path.basename(mockup_file)
                                    file_ext = os.path.splitext(file_name)[1].lower()
                                    new_filename = f'{product_number}_MOCKUP{file_ext}'
                                    media_file = ContentFile(file_data, name=new_filename)
                                    media = Media(
                                        file=media_file,
                                        media_type='image',
                                        created_by=product_owner
                                    )
                                    media.set_temp_product_id(product.id)
                                    media.save()
                                    meta_info = {
                                        'type': 'MOCKUP',
                                        'folder_name': folder_name,
                                        'original_filename': file_name,
                                        'is_mockup': True
                                    }
                                    product.attach_media(media, meta=meta_info, created_by=task.user)
                                    try:
                                        avif_path, _ = create_avif_from_media_file(
                                            media.file.name, product_number, is_mockup=True,
                                            product=product, created_by=product_owner
                                        )
                                    except Exception as avif_error:
                                        logger.warning(
                                            "AVIF mockup conversion failed for product_id=%s file=%s: %s",
                                            product.id,
                                            media.file.name,
                                            avif_error,
                                        )
                            except Exception:
                                pass
                            finally:
                                Media.clear_product_context()

                        # Process Tags (column 9, previously column 10)
                        tags_str = metadata.get('tags', '').strip() if metadata.get('tags') else ''
                        if tags_str:
                            tag_names = [tag.strip() for tag in tags_str.split(',') if tag.strip()]
                            for tag_name in tag_names:
                                if tag_name:
                                    tag, _ = Tags.objects.get_or_create(
                                        name=tag_name[:100],
                                        defaults={
                                            'tags_type': 'metadata',
                                            'created_by': task.user
                                        }
                                    )
                                    product.attach_tag(tag, meta={'source': 'bulk_upload'}, created_by=task.user)
                        
                        # Attach Plan if specified (nested atomic so failure doesn't abort design transaction)
                        if plan_value and str(plan_value).strip() in ['1', '2', '3']:
                            try:
                                with transaction.atomic():
                                    plan_name_mapping = {'1': 'basic', '2': 'prime', '3': 'premium'}
                                    plan_name = plan_name_mapping.get(str(plan_value).strip())
                                    if plan_name:
                                        plan, _ = Plan.objects.get_or_create(
                                            plan_name=plan_name,
                                            plan_duration='monthly',
                                            defaults={'created_by': task.user}
                                        )
                                        product.attach_plan(plan, meta={'source': 'bulk_upload'}, created_by=task.user)
                            except Exception:
                                pass

                        # SubProduct removed - using Product only

                    # Transaction committed - design is now saved

                    processed_count += 1

                    # Update progress every 5 designs (outside transaction for immediate commit)
                    if idx % 5 == 0 or idx == len(valid_folders):
                        task.processed_designs = processed_count
                        task.failed_designs = failed_count
                        task.save(update_fields=['processed_designs', 'failed_designs', 'updated_at'])

                except Exception as e:
                    failed_count += 1
                    error_msg = str(e)

                    failed_details.append({
                        'folder_name': folder_name,
                        'error': error_msg
                    })

                    # Force rollback so connection is not left in "aborted" state for next design (PostgreSQL)
                    try:
                        transaction.rollback()
                    except Exception:
                        pass

                    # Update progress even on failure (outside transaction)
                    if idx % 5 == 0 or idx == len(valid_folders):
                        task.processed_designs = processed_count
                        task.failed_designs = failed_count
                        task.save(update_fields=['processed_designs', 'failed_designs', 'updated_at'])

            # Note: Onboarding status is set manually in designer_onboarding_step4 view
            # We don't automatically update it here to prevent overriding manual settings
            
            # Update task status
            task.processed_designs = processed_count
            task.failed_designs = failed_count
            
            if failed_count > 0:
                task.error_message = json.dumps(failed_details[:50])  # Store first 50 failures
                if processed_count == 0:
                    task.status = 'failed'
                else:
                    task.status = 'completed'  # Partial success
            else:
                task.status = 'completed'
            
            task.save(update_fields=['status', 'processed_designs', 'failed_designs', 'error_message', 'updated_at'])
            logger.info(f"process_design_upload_task: completed for task_id={task_id}, processed={processed_count}, failed={failed_count}")

    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()

        # Update task with error
        try:
            task = DesignProcessingTask.objects.get(id=task_id)
            task.status = 'failed'
            task.error_message = str(e)
            task.save(update_fields=['status', 'error_message', 'updated_at'])
        except DesignProcessingTask.DoesNotExist:
            pass
        except Exception as update_error:
            pass

        logger.info(f"process_design_upload_task: failed for task_id={task_id}, error={e}")
        raise

