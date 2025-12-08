from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import StreamingHttpResponse
from django.db.models import Q, Avg, Count
from django.contrib.auth.models import User
from django.utils import timezone
from django.conf import settings
from datetime import timedelta
import random
import string
from Authentication.models import Email, MobileNumber, OTP
from MediaFiles.models import Media
from Catalog.models import Product, Category, Tags
from Plans.models import Plan
from django.db import transaction
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from .models import Addresses, DesignerProfile, Studio, StudioBusinessDetails, StudioMember, Ratings, DesignProcessingTask
from .tasks import process_design_upload_task
from .serializers import (
    AddressesSerializer, DesignerProfileSerializer, StudioSerializer,
    StudioBusinessDetailsSerializer, StudioMemberSerializer, StudioMemberListSerializer,
    CreateStudioMemberWithUserSerializer, RatingsSerializer,
    DesignProcessingTaskSerializer
)


@swagger_auto_schema(
    method='get',
    operation_summary='Addresses List',
    operation_description='Addresses List endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def addresses_list(request):
    """
    Get user's addresses or create a new address.
    """
    if request.method == 'GET':
        addresses = Addresses.objects.filter(created_by=request.user).order_by('-created_at')
        return Response({
            'addresses': AddressesSerializer(addresses, many=True).data,
            'total_addresses': addresses.count()
        })
    
    elif request.method == 'POST':
        serializer = AddressesSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=request.user)
            return Response({
                'message': 'Address created successfully',
                'address': serializer.data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='get',
    operation_summary='Address Detail',
    operation_description='Address Detail endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def address_detail(request, address_id):
    """
    Get, update, or delete a specific address.
    """
    try:
        address = Addresses.objects.get(id=address_id, created_by=request.user)
    except Addresses.DoesNotExist:
        return Response({
            'error': 'Address not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        return Response({
            'address': AddressesSerializer(address).data
        })
    
    elif request.method == 'PUT':
        serializer = AddressesSerializer(address, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=request.user)
            return Response({
                'message': 'Address updated successfully',
                'address': serializer.data
            })
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        address.delete()
        return Response({
            'message': 'Address deleted successfully'
        })


@swagger_auto_schema(
    method='get',
    operation_summary='Designer Profile',
    operation_description='Designer Profile endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def designer_profile(request):
    """
    Get or create/update user's designer profile.
    """
    if request.method == 'GET':
        try:
            profile = DesignerProfile.objects.get(created_by=request.user)
            return Response({
                'designer_profile': DesignerProfileSerializer(profile).data
            })
        except DesignerProfile.DoesNotExist:
            return Response({
                'designer_profile': None,
                'message': 'No designer profile found'
            })
    
    elif request.method == 'POST':
        try:
            profile = DesignerProfile.objects.get(created_by=request.user)
            serializer = DesignerProfileSerializer(profile, data=request.data, partial=True)
        except DesignerProfile.DoesNotExist:
            serializer = DesignerProfileSerializer(data=request.data)
        
        if serializer.is_valid():
            serializer.save(created_by=request.user, updated_by=request.user)
            return Response({
                'message': 'Designer profile updated successfully',
                'designer_profile': serializer.data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='get',
    operation_summary='Studios List',
    operation_description='Studios List endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def studios_list(request):
    """
    Get all studios with filtering options.
    """
    status_filter = request.GET.get('status')
    industry_type = request.GET.get('industry_type')
    search = request.GET.get('search')
    
    studios = Studio.objects.all()
    
    if status_filter:
        studios = studios.filter(status=status_filter)
    
    if industry_type:
        studios = studios.filter(studio_industry_type=industry_type)
    
    if search:
        studios = studios.filter(
            Q(name__icontains=search) | 
            Q(wedesignz_auto_name__icontains=search)
        )
    
    studios = studios.order_by('-created_at')
    
    return Response({
        'studios': StudioSerializer(studios, many=True).data,
        'total_studios': studios.count()
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Studio Detail',
    operation_description='Studio Detail endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def studio_detail(request, studio_id):
    """
    Get or update detailed information about a specific studio.
    """
    try:
        studio = Studio.objects.get(id=studio_id, created_by=request.user)
    except Studio.DoesNotExist:
        return Response({
            'error': 'Studio not found or you do not have permission'
        }, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        return Response({
            'studio': StudioSerializer(studio).data
        })
    
    elif request.method == 'PUT':
        serializer = StudioSerializer(studio, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=request.user)
            return Response({
                'message': 'Studio updated successfully',
                'studio': serializer.data
            })
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='post',
    operation_summary='Create Studio',
    operation_description='Create Studio endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_studio(request):
    """
    Create a new studio.
    """
    serializer = StudioSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save(created_by=request.user)
        return Response({
            'message': 'Studio created successfully',
            'studio': serializer.data
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='get',
    operation_summary='Studio Business Details',
    operation_description='Studio Business Details endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def studio_business_details(request, studio_id):
    """
    Get or create/update studio business details.
    """
    try:
        studio = Studio.objects.get(id=studio_id, created_by=request.user)
    except Studio.DoesNotExist:
        return Response({
            'error': 'Studio not found or you do not have permission'
        }, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        try:
            business_details = StudioBusinessDetails.objects.get(studio=studio)
            return Response({
                'business_details': StudioBusinessDetailsSerializer(business_details).data
            })
        except StudioBusinessDetails.DoesNotExist:
            return Response({
                'business_details': None,
                'message': 'No business details found'
            })
    
    elif request.method == 'POST':
        from MediaFiles.models import Media
        import logging
        logger = logging.getLogger(__name__)
        
        try:
            # Prepare data, handling FormData and JSON parsing
            # For FormData, request.data is a QueryDict, so we need to handle it differently
            if hasattr(request.data, 'dict'):
                # FormData - convert QueryDict to regular dict
                data = request.data.dict()
            elif hasattr(request.data, 'copy'):
                data = request.data.copy()
            else:
                data = dict(request.data)
            
            # Ensure studio_id is set (required by serializer) and convert to int
            data['studio_id'] = int(studio_id)
            
            # Parse registered_addresses_json if it's a string (from FormData)
            if 'registered_addresses_json' in data:
                import json
                if isinstance(data['registered_addresses_json'], str):
                    try:
                        parsed_json = json.loads(data['registered_addresses_json'])
                        data['registered_addresses_json'] = parsed_json if parsed_json else {}
                    except (json.JSONDecodeError, TypeError):
                        # If parsing fails, set to empty dict
                        data['registered_addresses_json'] = {}
            
            # Convert empty strings to None for optional fields
            optional_fields = ['gst_number', 'msme_udyam_number', 'business_model', 'pan_number', 'legal_business_name']
            for field in optional_fields:
                if field in data and data[field] == '':
                    data[field] = None
            
            # Handle file uploads - PAN Card
            if 'pan_card' in request.FILES:
                pan_card_file = request.FILES['pan_card']
                # Determine media type based on file extension
                file_name = pan_card_file.name.lower()
                if file_name.endswith('.pdf'):
                    media_type = 'pdf'
                elif any(file_name.endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp']):
                    media_type = 'image'
                else:
                    media_type = 'other'
                pan_card_media = Media.objects.create(
                    file=pan_card_file,
                    media_type=media_type,
                    created_by=request.user
                )
                pan_card_url = pan_card_media.file.url if hasattr(pan_card_media.file, 'url') else None
                if pan_card_url:
                    # Build absolute URL if relative
                    if pan_card_url.startswith('/'):
                        pan_card_url = request.build_absolute_uri(pan_card_url)
                    elif not pan_card_url.startswith('http'):
                        pan_card_url = request.build_absolute_uri('/' + pan_card_url)
                    data['pan_card'] = pan_card_url
            
            # Handle file uploads - MSME Certificate
            if 'msme_certificate_annexure' in request.FILES:
                msme_cert_file = request.FILES['msme_certificate_annexure']
                # Determine media type based on file extension
                file_name = msme_cert_file.name.lower()
                if file_name.endswith('.pdf'):
                    media_type = 'pdf'
                elif any(file_name.endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp']):
                    media_type = 'image'
                else:
                    media_type = 'other'
                msme_cert_media = Media.objects.create(
                    file=msme_cert_file,
                    media_type=media_type,
                    created_by=request.user
                )
                msme_cert_url = msme_cert_media.file.url if hasattr(msme_cert_media.file, 'url') else None
                if msme_cert_url:
                    # Build absolute URL if relative
                    if msme_cert_url.startswith('/'):
                        msme_cert_url = request.build_absolute_uri(msme_cert_url)
                    elif not msme_cert_url.startswith('http'):
                        msme_cert_url = request.build_absolute_uri('/' + msme_cert_url)
                    data['msme_certificate_annexure'] = msme_cert_url
            
            try:
                business_details = StudioBusinessDetails.objects.get(studio=studio)
                serializer = StudioBusinessDetailsSerializer(business_details, data=data, partial=True)
            except StudioBusinessDetails.DoesNotExist:
                serializer = StudioBusinessDetailsSerializer(data=data)
            
            if serializer.is_valid():
                serializer.save(studio=studio, created_by=request.user, updated_by=request.user)
                return Response({
                    'message': 'Business details updated successfully',
                    'business_details': serializer.data
                }, status=status.HTTP_201_CREATED)
            
            # Log validation errors for debugging
            logger.error(f'Studio business details validation failed: {serializer.errors}')
            logger.error(f'Data sent: {data}')
            
            # Return detailed validation errors
            return Response({
                'error': 'Validation failed',
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            logger.error(f'Error updating studio business details: {str(e)}', exc_info=True)
            return Response({
                'error': 'An error occurred while updating business details',
                'detail': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Studio Members',
    operation_description='Studio Members endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def studio_members(request, studio_id):
    """
    Get studio members or add a new member.
    """
    try:
        studio = Studio.objects.get(id=studio_id, created_by=request.user)
    except Studio.DoesNotExist:
        return Response({
            'error': 'Studio not found or you do not have permission'
        }, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        members = StudioMember.objects.filter(studio=studio).order_by('-created_at')
        return Response({
            'members': StudioMemberSerializer(members, many=True).data,
            'total_members': members.count()
        })
    
    elif request.method == 'POST':
        from django.db import transaction
        
        serializer = StudioMemberSerializer(data=request.data)
        if serializer.is_valid():
            try:
                with transaction.atomic():
                    # Extract member_id from validated data
                    member_id = serializer.validated_data.get('member_id')
                    if not member_id:
                        return Response({
                            'error': 'member_id is required'
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    # Get the member user
                    try:
                        member_user = User.objects.get(id=member_id)
                    except User.DoesNotExist:
                        return Response({
                            'error': 'Member user does not exist'
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    # Save the StudioMember with member field
                    studio_member = serializer.save(
                        studio=studio,
                        member=member_user,
                        created_by=request.user
                    )
                    
                    # Auto-create DesignerProfile for the member user if it doesn't exist
                    DesignerProfile.objects.get_or_create(
                        created_by=member_user,
                        defaults={
                            'is_individual': False,
                            'status': 'pending'
                        }
                    )
                    
                    return Response({
                        'message': 'Studio member added successfully',
                        'member': StudioMemberSerializer(studio_member).data
                    }, status=status.HTTP_201_CREATED)
            except Exception as e:
                from django.conf import settings
                import traceback
                error_message = str(e)
                if settings.DEBUG:
                    error_message += f"\n{traceback.format_exc()}"
                return Response({
                    'error': f'Failed to add studio member: {error_message}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='post',
    operation_summary='Create New Studio Member with User Account',
    operation_description='Create a new user account and add them as a studio member. Only studio owners can create members.',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'email': openapi.Schema(type=openapi.TYPE_STRING, format=openapi.FORMAT_EMAIL, description='Email address for the new user'),
            'password': openapi.Schema(type=openapi.TYPE_STRING, description='Password (minimum 8 characters)'),
            'confirm_password': openapi.Schema(type=openapi.TYPE_STRING, description='Password confirmation'),
            'first_name': openapi.Schema(type=openapi.TYPE_STRING, description='First name (optional)'),
            'last_name': openapi.Schema(type=openapi.TYPE_STRING, description='Last name (optional)'),
            'role': openapi.Schema(type=openapi.TYPE_STRING, enum=['designer', 'design_lead'], description='Member role', default='designer'),
        },
        required=['email', 'password', 'confirm_password']
    ),
    responses={
        201: openapi.Response(description='Studio member created successfully'),
        400: openapi.Response(description='Bad request - validation errors'),
        404: openapi.Response(description='Studio not found or permission denied')
    },
    tags=['API']
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_studio_member_with_user(request, studio_id):
    """
    Create a new user account and add them as a studio member.
    Only studio owners can create members.
    """
    try:
        studio = Studio.objects.get(id=studio_id, created_by=request.user)
    except Studio.DoesNotExist:
        return Response({
            'error': 'Studio not found or you do not have permission'
        }, status=status.HTTP_404_NOT_FOUND)
    
    serializer = CreateStudioMemberWithUserSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        with transaction.atomic():
            # Generate unique username from email
            email = serializer.validated_data['email']
            base_username = email.split('@')[0]
            username = base_username
            counter = 1
            
            # Ensure username is unique
            while User.objects.filter(username=username).exists():
                username = f"{base_username}{counter}"
                counter += 1
            
            # Create User
            user = User.objects.create_user(
                username=username,
                email=email,
                first_name=serializer.validated_data.get('first_name', ''),
                last_name=serializer.validated_data.get('last_name', ''),
                password=serializer.validated_data['password'],
                is_active=True
            )
            
            # Create Email record
            Email.objects.create(
                email=email,
                is_primary=True,
                is_verified=False,
                created_by=user
            )
            
            # Create StudioMember (NO DesignerProfile)
            studio_member = StudioMember.objects.create(
                studio=studio,
                member=user,
                role=serializer.validated_data.get('role', 'designer'),
                status='active',
                created_by=request.user
            )
            
            return Response({
                'message': 'Studio member created successfully',
                'member': StudioMemberListSerializer(studio_member).data,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                }
            }, status=status.HTTP_201_CREATED)
            
    except Exception as e:
        import traceback
        error_message = str(e)
        if settings.DEBUG:
            error_message += f"\n{traceback.format_exc()}"
        return Response({
            'error': f'Failed to create studio member: {error_message}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='post',
    operation_summary='Send Login Credentials to Studio Member',
    operation_description='Send login credentials (email and password) to a studio member via email.',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={},
    ),
    responses={
        200: openapi.Response(description='Credentials sent successfully'),
        404: openapi.Response(description='Studio member not found'),
        403: openapi.Response(description='Permission denied')
    },
    tags=['API']
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_studio_member_credentials(request, studio_id, member_id):
    """
    Send login credentials to a studio member.
    Only studio owners can send credentials.
    """
    try:
        studio = Studio.objects.get(id=studio_id, created_by=request.user)
    except Studio.DoesNotExist:
        return Response({
            'error': 'Studio not found or you do not have permission'
        }, status=status.HTTP_404_NOT_FOUND)
    
    try:
        studio_member = StudioMember.objects.get(id=member_id, studio=studio)
        member_user = studio_member.member
        
        if not member_user:
            return Response({
                'error': 'Member user not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Get the password from request (it should be stored temporarily or retrieved)
        # Since we can't retrieve password, we'll need to pass it in the request
        # For security, this should only work immediately after creation
        password = request.data.get('password')
        
        if not password:
            return Response({
                'error': 'Password is required to send credentials'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Send credentials email
        from common.email_service import EmailService
        success = EmailService.send_studio_member_credentials_email(
            member_user, 
            studio, 
            studio_member.role,
            password
        )
        
        if success:
            return Response({
                'message': 'Login credentials sent successfully'
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                'error': 'Failed to send credentials email'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    except StudioMember.DoesNotExist:
        return Response({
            'error': 'Studio member not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        import traceback
        error_message = str(e)
        if settings.DEBUG:
            error_message += f"\n{traceback.format_exc()}"
        return Response({
            'error': f'Failed to send credentials: {error_message}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Studio Member Detail',
    operation_description='Studio Member Detail endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def studio_member_detail(request, studio_id, member_id):
    """
    Get, update, or delete a specific studio member.
    """
    try:
        studio = Studio.objects.get(id=studio_id, created_by=request.user)
        member = StudioMember.objects.get(id=member_id, studio=studio)
    except (Studio.DoesNotExist, StudioMember.DoesNotExist):
        return Response({
            'error': 'Studio or member not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        return Response({
            'member': StudioMemberSerializer(member).data
        })
    
    elif request.method == 'PUT':
        serializer = StudioMemberSerializer(member, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=request.user)
            return Response({
                'message': 'Studio member updated successfully',
                'member': serializer.data
            })
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        from django.db import transaction
        
        try:
            with transaction.atomic():
                # Get the member user before deleting the StudioMember entry
                member_user = member.member
                member_id = member.id
                member_username = member_user.username if member_user else None
                
                # Delete the StudioMember entry first
                member.delete()
                
                # Delete the User account
                # Django will handle cascading deletes based on model on_delete settings
                if member_user:
                    member_user.delete()
                
                return Response({
                    'message': 'Studio member and user account removed successfully',
                    'deleted_member_id': member_id,
                    'deleted_username': member_username
                })
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error deleting studio member: {e}')
            import traceback
            logger.error(traceback.format_exc())
            return Response({
                'error': f'Failed to delete studio member: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Ratings List',
    operation_description='Ratings List endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def ratings_list(request):
    """
    Get ratings or create a new rating.
    """
    if request.method == 'GET':
        rating_type = request.GET.get('rating_type')
        studio_id = request.GET.get('studio_id')
        product_id = request.GET.get('product_id')
        
        ratings = Ratings.objects.filter(status='show')
        
        if rating_type:
            ratings = ratings.filter(rating_type=rating_type)
        
        if studio_id:
            ratings = ratings.filter(studio_id=studio_id)
        
        if product_id:
            ratings = ratings.filter(product_id=product_id)
        
        ratings = ratings.order_by('-created_at')
        
        return Response({
            'ratings': RatingsSerializer(ratings, many=True).data,
            'total_ratings': ratings.count()
        })
    
    elif request.method == 'POST':
        serializer = RatingsSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=request.user)
            return Response({
                'message': 'Rating created successfully',
                'rating': serializer.data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='get',
    operation_summary='Studio Ratings',
    operation_description='Studio Ratings endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def studio_ratings(request, studio_id):
    """
    Get ratings for a specific studio.
    """
    try:
        studio = Studio.objects.get(id=studio_id)
    except Studio.DoesNotExist:
        return Response({
            'error': 'Studio not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    ratings = Ratings.objects.filter(studio=studio, status='show').order_by('-created_at')
    average_rating = ratings.aggregate(avg_rating=Avg('rating_value'))['avg_rating'] or 0
    
    return Response({
        'studio': StudioSerializer(studio).data,
        'ratings': RatingsSerializer(ratings, many=True).data,
        'average_rating': round(average_rating, 2),
        'total_ratings': ratings.count()
    })


@swagger_auto_schema(
    method='get',
    operation_summary='My Studios',
    operation_description='My Studios endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_studios(request):
    """
    Get studios created by the current user.
    """
    studios = Studio.objects.filter(created_by=request.user).order_by('-created_at')
    
    return Response({
        'studios': StudioSerializer(studios, many=True).data,
        'total_studios': studios.count()
    })


@swagger_auto_schema(
    method='get',
    operation_summary='My Ratings',
    operation_description='My Ratings endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_ratings(request):
    """
    Get ratings created by the current user.
    """
    ratings = Ratings.objects.filter(created_by=request.user).order_by('-created_at')
    
    return Response({
        'ratings': RatingsSerializer(ratings, many=True).data,
        'total_ratings': ratings.count()
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Top Studios',
    operation_description='Top Studios endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def top_studios(request):
    """
    Get top-rated studios.
    """
    studios = Studio.objects.filter(
        status='active',
        ratings__status='show'
    ).annotate(
        avg_rating=Avg('ratings__rating_value'),
        total_ratings=Count('ratings')
    ).filter(
        total_ratings__gt=0
    ).order_by('-avg_rating', '-total_ratings')
    
    return Response({
        'top_studios': StudioSerializer(studios, many=True).data,
        'total_studios': studios.count()
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Studio Member Ratings',
    operation_description='Studio Member Ratings endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def studio_member_ratings(request, studio_id, member_id):
    """
    Get ratings for a specific studio member.
    """
    try:
        studio = Studio.objects.get(id=studio_id)
        member = StudioMember.objects.get(id=member_id, studio=studio)
    except (Studio.DoesNotExist, StudioMember.DoesNotExist):
        return Response({
            'error': 'Studio or member not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    ratings = Ratings.objects.filter(
        studio_member=member,
        status='show'
    ).order_by('-created_at')
    
    average_rating = ratings.aggregate(avg_rating=Avg('rating_value'))['avg_rating'] or 0
    
    return Response({
        'member': StudioMemberSerializer(member).data,
        'ratings': RatingsSerializer(ratings, many=True).data,
        'average_rating': round(average_rating, 2),
        'total_ratings': ratings.count()
    })


# ==================== DESIGNER CONSOLE VIEWS ====================

@swagger_auto_schema(
    method='post',
    operation_summary='Designer Onboarding Step1',
    operation_description='Designer Onboarding Step1 endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def designer_onboarding_step1(request):
    """
    Designer Onboarding Step 1: Basic information and profile creation.
    Links DesignerProfile to existing authenticated user.
    Accepts: first_name, last_name, email, phone, is_individual, profile_photo
    """
    from django.db import transaction
    from Authentication.models import Email, MobileNumber
    from MediaFiles.models import Media
    from common.relations import attach_relation, detach_relation
    
    # Validate required fields
    required_fields = ['first_name', 'last_name', 'email', 'phone', 'is_individual']
    for field in required_fields:
        if field not in request.data:
            return Response({
                'error': f'{field} is required'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    user = request.user
    
    # Convert is_individual from string to boolean (FormData sends strings)
    is_individual_str = request.data.get('is_individual', 'False')
    is_individual = is_individual_str.lower() in ('true', '1', 'yes', 'on')
    
    try:
        with transaction.atomic():
            # Update user's basic info
            user.first_name = request.data['first_name']
            user.last_name = request.data['last_name']
            user.save()
            
            # Update or create email record
            email_obj, _ = Email.objects.get_or_create(
                email=request.data['email'],
                created_by=user,
                defaults={'is_primary': True}
            )
            if not email_obj.is_primary:
                # Make this email primary
                Email.objects.filter(created_by=user, is_primary=True).update(is_primary=False)
                email_obj.is_primary = True
                email_obj.save()
            
            # Update or create mobile number record
            mobile_obj, _ = MobileNumber.objects.get_or_create(
                mobile_number=request.data['phone'],
                created_by=user,
                defaults={'is_primary': True}
            )
            if not mobile_obj.is_primary:
                # Make this mobile primary
                MobileNumber.objects.filter(created_by=user, is_primary=True).update(is_primary=False)
                mobile_obj.is_primary = True
                mobile_obj.save()
            
            # Get or create designer profile
            designer_profile, created = DesignerProfile.objects.get_or_create(
                created_by=user,
                defaults={'is_individual': is_individual}
            )
            
            # Update is_individual if profile already exists
            if not created:
                designer_profile.is_individual = is_individual
                designer_profile.updated_by = user
                designer_profile.save()
            
            # Handle profile photo if provided
            if 'profile_photo' in request.FILES:
                # Remove old profile photo if exists
                from MediaFiles.models import Relation
                old_relations = Relation.objects.filter(
                    relation_type='DesignerProfile:Media',
                    id_1=designer_profile.pk
                )
                for relation in old_relations:
                    if relation.meta and relation.meta.get('type') == 'profile_photo':
                        relation.delete()  # Delete the relation
                
                # Create new profile photo
                profile_photo = Media.objects.create(
                    file=request.FILES['profile_photo'],
                    media_type='image',
                    created_by=user
                )
                attach_relation('DesignerProfile:Media', designer_profile, profile_photo, 
                              meta={'type': 'profile_photo'}, created_by=user)
            
            # Get profile photo URL if exists
            profile_photo_url = None
            from MediaFiles.models import Relation
            profile_photo_relations = Relation.objects.filter(
                relation_type='DesignerProfile:Media',
                id_1=designer_profile.pk
            )
            for relation in profile_photo_relations:
                if relation.meta and relation.meta.get('type') == 'profile_photo':
                    try:
                        profile_photo = Media.objects.get(pk=relation.id_2)
                        if hasattr(profile_photo.file, 'url'):
                            # Build absolute URL for the media file
                            relative_url = profile_photo.file.url
                            if relative_url.startswith('/'):
                                profile_photo_url = request.build_absolute_uri(relative_url)
                            elif relative_url.startswith('http'):
                                profile_photo_url = relative_url
                            else:
                                # If relative URL doesn't start with /, prepend it
                                profile_photo_url = request.build_absolute_uri('/' + relative_url)
                            break
                    except Media.DoesNotExist:
                        continue
            
            return Response({
                'message': 'Step 1 data saved successfully',
                'data': {
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'email': email_obj.email,
                    'phone': mobile_obj.mobile_number,
                    'is_individual': designer_profile.is_individual,
                    'profile_photo_url': profile_photo_url,
                    'designer_profile_id': designer_profile.id
                }
            }, status=status.HTTP_200_OK)
        
    except Exception as e:
        import traceback
        from django.conf import settings
        error_trace = traceback.format_exc()
        print(f"Error in designer_onboarding_step1: {str(e)}")
        print(error_trace)
        return Response({
            'error': f'Failed to save Step 1 data: {str(e)}',
            'detail': str(e) if settings.DEBUG else 'An error occurred while saving your data'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='post',
    operation_summary='Verify Otp',
    operation_description='Verify Otp endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def verify_otp(request):
    """
    Verify OTP for email or mobile verification.
    """
    from Accounts.models import OTP, Email, MobileNumber
    from django.utils import timezone
    
    otp = request.data.get('otp')
    otp_type = request.data.get('otp_type')  # 'E' for email, 'M' for mobile
    otp_for = request.data.get('otp_for')  # 'email_verification' or 'mobile_verification'
    
    if not all([otp, otp_type, otp_for]):
        return Response({
            'error': 'OTP, otp_type, and otp_for are required'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        otp_obj = OTP.objects.get(
            otp=otp,
            otp_type=otp_type,
            otp_for=otp_for,
            created_by=request.user,
            is_verified=False
        )
        
        if otp_obj.is_expired():
            return Response({
                'error': 'OTP has expired'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Delete OTP record after successful verification
        otp_obj.delete()
        
        # Update email or mobile verification status
        if otp_type == 'E':
            email_obj = Email.objects.get(email=request.user.email, created_by=request.user)
            email_obj.is_verified = True
            email_obj.save()
        elif otp_type == 'M':
            mobile_obj = MobileNumber.objects.get(created_by=request.user, is_primary=True)
            mobile_obj.is_verified = True
            mobile_obj.save()
        
        # Check if both email and mobile are verified
        email_verified = Email.objects.filter(email=request.user.email, created_by=request.user, is_verified=True).exists()
        mobile_verified = MobileNumber.objects.filter(created_by=request.user, is_primary=True, is_verified=True).exists()
        
        if email_verified and mobile_verified:
            # Activate user account
            request.user.is_active = True
            request.user.save()
            
            return Response({
                'message': 'Account verified and activated successfully',
                'account_activated': True
            })
        
        return Response({
            'message': 'OTP verified successfully',
            'account_activated': False
        })
        
    except OTP.DoesNotExist:
        return Response({
            'error': 'Invalid OTP'
        }, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='post',
    operation_summary='Designer Onboarding Step2',
    operation_description='Designer Onboarding Step2 endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def designer_onboarding_step2(request):
    """
    Designer Onboarding Step 2: Business details for Razorpay Linked Account creation.
    Accepts: legal_business_name, business_type, business_model, business_category, 
    business_sub_category, street, city, state, postal_code, country, 
    studio_email, studio_mobile_number, msme_udyam_number, msme_certificate_annexure, gst_number
    """
    from django.db import transaction
    from MediaFiles.models import Media
    from common.relations import attach_relation, detach_relation
    
    # Validate required fields
    required_fields = [
        'legal_business_name', 'business_type', 'business_model', 
        'business_category', 'business_sub_category',
        'street', 'city', 'state', 'postal_code',
        'studio_email', 'studio_mobile_number'
    ]
    
    for field in required_fields:
        if field not in request.data:
            return Response({
                'error': f'{field} is required'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        with transaction.atomic():
            # Prepare registered_addresses_json
            registered_address = {
                'street': request.data['street'],
                'city': request.data['city'],
                'state': request.data['state'],
                'postal_code': request.data['postal_code'],
                'country': request.data.get('country', 'India')
            }
            
            # Get or create studio
            studio, studio_created = Studio.objects.get_or_create(
                created_by=request.user,
                defaults={
                    'name': request.data['legal_business_name'],
                    'wedesignz_auto_name': f"WD{request.user.id:09d}",
                    'studio_industry_type': 'design_studio',
                }
            )
            
            # Update studio name if it changed
            if not studio_created and studio.name != request.data['legal_business_name']:
                studio.name = request.data['legal_business_name']
                studio.updated_by = request.user
                studio.save()
            
            # Get or create studio business details
            business_details, details_created = StudioBusinessDetails.objects.get_or_create(
                studio=studio,
                defaults={
                    'studio_email': request.data['studio_email'],
                    'studio_mobile_number': request.data['studio_mobile_number'],
                    'legal_business_name': request.data['legal_business_name'],
                    'business_type': request.data['business_type'],
                    'business_category': request.data['business_category'],
                    'business_sub_category': request.data['business_sub_category'],
                    'business_model': request.data['business_model'],
                    'registered_addresses_json': registered_address,
                    'gst_number': request.data.get('gst_number', ''),
                    'msme_udyam_number': request.data.get('msme_udyam_number', ''),
                    'created_by': request.user
                }
            )
            
            # Update if already exists
            if not details_created:
                business_details.studio_email = request.data['studio_email']
                business_details.studio_mobile_number = request.data['studio_mobile_number']
                business_details.legal_business_name = request.data['legal_business_name']
                business_details.business_type = request.data['business_type']
                business_details.business_category = request.data['business_category']
                business_details.business_sub_category = request.data['business_sub_category']
                business_details.business_model = request.data['business_model']
                business_details.registered_addresses_json = registered_address
                business_details.gst_number = request.data.get('gst_number', '')
                business_details.msme_udyam_number = request.data.get('msme_udyam_number', '')
                business_details.updated_by = request.user
                business_details.save()
            
            # Handle MSME certificate annexure file upload
            msme_cert_url = None
            if 'msme_certificate_annexure' in request.FILES:
                # Remove old MSME certificate if exists
                if business_details.msme_certificate_annexure:
                    # Delete old file logic can be added here if needed
                    pass
                
                # Create new MSME certificate media
                msme_cert = Media.objects.create(
                    file=request.FILES['msme_certificate_annexure'],
                    media_type='document',
                    created_by=request.user
                )
                msme_cert_url = msme_cert.file.url if hasattr(msme_cert.file, 'url') else None
                business_details.msme_certificate_annexure = msme_cert_url
                business_details.save()
            
            return Response({
                'message': 'Step 2 data saved successfully',
                'data': {
                    'studio_id': studio.id,
                    'business_details_id': business_details.id,
                    'legal_business_name': business_details.legal_business_name,
                    'business_type': business_details.business_type,
                    'business_model': business_details.business_model,
                    'business_category': business_details.business_category,
                    'business_sub_category': business_details.business_sub_category,
                    'registered_address': business_details.registered_addresses_json,
                    'studio_email': business_details.studio_email,
                    'studio_mobile_number': business_details.studio_mobile_number,
                    'gst_number': business_details.gst_number,
                    'msme_udyam_number': business_details.msme_udyam_number,
                    'msme_certificate_annexure_url': business_details.msme_certificate_annexure
                }
            }, status=status.HTTP_200_OK)
            
    except Exception as e:
        return Response({
            'error': f'Failed to save Step 2 data: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='post',
    operation_summary='Designer Onboarding Step3',
    operation_description='Designer Onboarding Step3 endpoint for PAN details',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'pan_number': openapi.Schema(type=openapi.TYPE_STRING, description='PAN number'),
            'pan_card': openapi.Schema(type=openapi.TYPE_STRING, format=openapi.FORMAT_BINARY, description='PAN card document file')
        },
        required=['pan_number']
    ),
    responses={
        200: openapi.Response(description='Success'),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def designer_onboarding_step3(request):
    """
    Designer Onboarding Step 3: Legal Info - PAN details.
    Accepts: pan_number, pan_card (file)
    For individuals: PAN is saved to DesignerProfile
    For companies: PAN is saved to StudioBusinessDetails
    """
    from django.db import transaction
    from MediaFiles.models import Media
    
    # Validate required fields
    if 'pan_number' not in request.data:
        return Response({
            'error': 'pan_number is required'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        with transaction.atomic():
            # Get designer profile to check if individual
            designer_profile = DesignerProfile.objects.get(created_by=request.user)
            is_individual = designer_profile.is_individual
            
            pan_number = request.data['pan_number']
            pan_card_url = None
            
            # Handle PAN card document file upload
            if 'pan_card' in request.FILES:
                # Create new PAN card media
                pan_card = Media.objects.create(
                    file=request.FILES['pan_card'],
                    media_type='document',
                    created_by=request.user
                )
                pan_card_url = pan_card.file.url if hasattr(pan_card.file, 'url') else None
            
            if is_individual:
                # For individuals, store PAN in DesignerProfile (we'll add a field or use a relation)
                # For now, we'll create a minimal studio for consistency
                studio, studio_created = Studio.objects.get_or_create(
                    created_by=request.user,
                    defaults={
                        'name': f"{request.user.first_name} {request.user.last_name}",
                        'wedesignz_auto_name': f"WD{request.user.id:09d}",
                        'studio_industry_type': 'design_studio',
                    }
                )
                
                # Get or create business details for individual
                business_details, created = StudioBusinessDetails.objects.get_or_create(
                    studio=studio,
                    defaults={'created_by': request.user}
                )
                business_details.pan_number = pan_number
                if pan_card_url:
                    business_details.pan_card = pan_card_url
                business_details.updated_by = request.user
                business_details.save()
            else:
                # For companies, use existing studio
                studio = Studio.objects.filter(created_by=request.user).first()
                if not studio:
                    return Response({
                        'error': 'Studio not found. Please complete Step 2 first.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Get or create business details
                business_details, created = StudioBusinessDetails.objects.get_or_create(
                    studio=studio,
                    defaults={'created_by': request.user}
                )
                business_details.pan_number = pan_number
                if pan_card_url:
                    business_details.pan_card = pan_card_url
            business_details.updated_by = request.user
            business_details.save()
            
            return Response({
                'message': 'Step 3 data saved successfully',
                'data': {
                    'pan_number': pan_number,
                    'pan_card_url': pan_card_url,
                    'business_details_id': business_details.id
                }
            }, status=status.HTTP_200_OK)
            
    except DesignerProfile.DoesNotExist:
        return Response({
            'error': 'Designer profile not found. Please complete Step 1 first.'
        }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({
            'error': f'Failed to save Step 3 data: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='post',
    operation_summary='Designer Onboarding Step4',
    operation_description='Designer Onboarding Step4 endpoint for bank details or bulk design upload with validation',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'bank_account_number': openapi.Schema(type=openapi.TYPE_STRING, description='Bank account number'),
            'bank_ifsc_code': openapi.Schema(type=openapi.TYPE_STRING, description='IFSC code'),
            'bank_account_holder_name': openapi.Schema(type=openapi.TYPE_STRING, description='Account holder name'),
            'account_type': openapi.Schema(type=openapi.TYPE_STRING, description='Account type (savings/current)'),
            'zip_file': openapi.Schema(type=openapi.TYPE_STRING, format=openapi.FORMAT_BINARY, description='Zip file containing designs')
        }
    ),
    responses={
        200: openapi.Response(description='Success'),
        400: openapi.Response(description='Bad request - validation failed')
    },
    tags=['API']
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def designer_onboarding_step4(request):
    """
    Designer Onboarding Step 4: Bank Details or Upload Designs.
    If bank details are provided (JSON), saves to StudioBusinessDetails.
    If zip_file is provided, validates zip file structure, metadata.xlsx, and design folders.
    """
    from django.db import transaction
    
    # Check if this is a bank details request (JSON data with bank fields)
    if 'bank_account_number' in request.data or 'bank_ifsc_code' in request.data:
        # Handle bank details
        return save_bank_details(request)
    
    # Otherwise, handle design upload (zip file)
    return handle_design_upload(request)


def save_bank_details(request):
    """
    Save bank details to StudioBusinessDetails.
    """
    from django.db import transaction
    
    # Validate required fields
    required_fields = ['bank_account_number', 'bank_ifsc_code', 'bank_account_holder_name', 'account_type']
    for field in required_fields:
        if field not in request.data:
            return Response({
                'error': f'{field} is required'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        with transaction.atomic():
            # Get designer profile to check if individual
            designer_profile = DesignerProfile.objects.get(created_by=request.user)
            
            # Get or find studio (should exist from Step 2 or Step 3)
            studio = Studio.objects.filter(created_by=request.user).first()
            
            if not studio:
                return Response({
                    'error': 'Studio not found. Please complete previous onboarding steps first.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get or create business details
            business_details, created = StudioBusinessDetails.objects.get_or_create(
                studio=studio,
                defaults={'created_by': request.user}
            )
            
            # Update bank details
            business_details.bank_account_number = request.data['bank_account_number']
            business_details.bank_ifsc_code = request.data['bank_ifsc_code']
            business_details.bank_account_holder_name = request.data['bank_account_holder_name']
            business_details.account_type = request.data['account_type']
            business_details.updated_by = request.user
            business_details.save()
            
            return Response({
                'message': 'Bank details saved successfully',
                'data': {
                    'bank_account_number': business_details.bank_account_number,  # Note: In production, you might want to mask this
                    'bank_ifsc_code': business_details.bank_ifsc_code,
                    'bank_account_holder_name': business_details.bank_account_holder_name,
                    'account_type': business_details.account_type,
                    'business_details_id': business_details.id
                }
            }, status=status.HTTP_200_OK)
            
    except DesignerProfile.DoesNotExist:
        return Response({
            'error': 'Designer profile not found. Please complete Step 1 first.'
        }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        return Response({
            'error': f'Failed to save bank details: {str(e)}',
            'traceback': error_traceback if settings.DEBUG else None
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def handle_design_upload(request):
    """
    Handle design upload (zip file).
    Validates zip file structure, metadata.xlsx, and design folders.
    """
    import zipfile
    import io
    import os
    from openpyxl import load_workbook
    
    # Check if zip file is provided
    if 'zip_file' not in request.FILES:
        return Response({
            'error': '❌ No file uploaded: Please select a zip file to upload.',
            'validation_errors': ['❌ No file uploaded: Please select a zip file to upload.']
        }, status=status.HTTP_400_BAD_REQUEST)
    
    zip_file = request.FILES['zip_file']
    
    # Validate file size (1GB = 1073741824 bytes)
    MAX_FILE_SIZE = 1073741824  # 1GB
    if zip_file.size > MAX_FILE_SIZE:
        file_size_mb = zip_file.size / (1024*1024)
        return Response({
            'error': f'❌ File too large: Your file is {file_size_mb:.2f} MB, but the maximum allowed size is 1GB (1024 MB). Please compress your files or split them into smaller zip files.',
            'validation_errors': [f'❌ File too large: Your file is {file_size_mb:.2f} MB, but the maximum allowed size is 1GB (1024 MB). Please compress your files or split them into smaller zip files.']
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Validate file extension
    if not zip_file.name.lower().endswith('.zip'):
        return Response({
            'error': '❌ Invalid file type: Please upload a .zip file. The file you selected is not a zip archive.',
            'validation_errors': ['❌ Invalid file type: Please upload a .zip file. The file you selected is not a zip archive.']
        }, status=status.HTTP_400_BAD_REQUEST)
    
    validation_errors = []
    zip_folders = {}  # Initialize to avoid scope issues
    valid_design_folders = {}  # Initialize to avoid scope issues
    all_files = []  # Initialize to avoid scope issues
    root_folder = ''  # Initialize to avoid scope issues
    metadata_file = None  # Initialize to avoid scope issues
    zip_content = None  # Store zip content for reuse
    
    try:
        # Read zip file into memory (only once)
        zip_content = zip_file.read()
        zip_buffer = io.BytesIO(zip_content)
        
        # Open zip file
        with zipfile.ZipFile(zip_buffer, 'r') as zip_ref:
            # Get all file names in zip
            all_files = zip_ref.namelist()
            
            # Check for metadata.xlsx (can be at root or in a subfolder)
            metadata_file = None
            for file_name in all_files:
                # Check if it's metadata.xlsx (case-insensitive)
                lower_name = file_name.lower()
                # Match: metadata.xlsx, folder/metadata.xlsx, or any path ending with /metadata.xlsx
                if lower_name == 'metadata.xlsx' or lower_name.endswith('/metadata.xlsx'):
                    metadata_file = file_name
                    break
            
            if not metadata_file:
                validation_errors.append('❌ metadata.xlsx file not found: Your zip file must contain a file named "metadata.xlsx" at the root level (inside the main folder).')
            else:
                # Read and parse metadata.xlsx
                try:
                    metadata_data = zip_ref.read(metadata_file)
                    metadata_buffer = io.BytesIO(metadata_data)
                    workbook = load_workbook(metadata_buffer, data_only=True)
                    sheet = workbook.active
                    
                    # Find folder_name column
                    header_row = None
                    folder_name_col = None
                    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=False), 1):
                        values = [str(cell.value).lower() if cell.value else '' for cell in row]
                        if 'folder_name' in values:
                            header_row = idx
                            folder_name_col = values.index('folder_name') + 1
                            break
                    
                    if not folder_name_col:
                        validation_errors.append('❌ Missing required column: The metadata.xlsx file must have a column named "folder_name" (case-insensitive). This column should list all your design folder names.')
                    else:
                        # Extract folder names from Excel
                        excel_folders = set()
                        for row in sheet.iter_rows(min_row=header_row + 1, values_only=False):
                            folder_name_cell = row[folder_name_col - 1]
                            if folder_name_cell.value:
                                folder_name = str(folder_name_cell.value).strip()
                                if folder_name:
                                    excel_folders.add(folder_name)
                        
                        # System folders to ignore
                        SYSTEM_FOLDERS = ['__macosx', '.ds_store', 'rar', '.rar', 'thumbs.db']
                        
                        # Find root folder (first folder in path, e.g., "dummy" from "dummy/WD1/file.eps")
                        root_folder = ''
                        for file_name in all_files:
                            if '/' in file_name and not file_name.endswith('/'):
                                parts = file_name.split('/')
                                if len(parts) >= 2:
                                    root_folder = parts[0]
                                    break
                        
                        # Extract actual design folders from zip (skip root folder and system folders)
                        zip_folders = {}
                        required_files = {'.eps', '.cdr', '.jpg', '.png'}
                        
                        for file_name in all_files:
                            # Skip metadata.xlsx and directories
                            if file_name == metadata_file or file_name.endswith('/'):
                                continue
                            
                            # Extract folder structure
                            if '/' in file_name:
                                parts = file_name.split('/')
                                
                                # Skip if not enough parts (should have at least root/folder/file)
                                if len(parts) < 3:
                                    continue
                                
                                root_folder_name = parts[0].lower()
                                folder_name = parts[1]
                                file_name_only = parts[-1]
                                file_ext = os.path.splitext(file_name_only)[1].lower()
                                file_name_lower = file_name_only.lower()
                                
                                # Check if it's an optional mockup file (case insensitive)
                                is_mockup_file = file_name_lower == 'mockup.jpg' or file_name_lower == 'mockup.png'
                                
                                # Skip system folders
                                if root_folder_name in SYSTEM_FOLDERS or folder_name.lower() in SYSTEM_FOLDERS:
                                    continue
                                
                                # Only process files in design folders (second level, e.g., dummy/WD1/file.eps)
                                # Accept required files or optional mockup files
                                if len(parts) == 3 and (file_ext in required_files or is_mockup_file):
                                    if folder_name not in zip_folders:
                                        zip_folders[folder_name] = set()
                                    # Only add required files to the set (mockup is optional)
                                    if file_ext in required_files:
                                        zip_folders[folder_name].add(file_ext)
                        
                        # Filter folders that have all required files
                        valid_design_folders.clear()  # Clear and reuse the initialized dict
                        for folder_name, files in zip_folders.items():
                            if all(ext in files for ext in required_files):
                                valid_design_folders[folder_name] = files
                        
                        # Validate folder count (use configurable minimum) - only for onboarding
                        # Check if user has already completed onboarding
                        from .models import DesignerProfile
                        is_onboarding = False
                        try:
                            designer_profile = DesignerProfile.objects.get(created_by=request.user)
                            is_onboarding = not designer_profile.onboarding_completed
                        except DesignerProfile.DoesNotExist:
                            # If no profile exists, this is onboarding
                            is_onboarding = True
                        
                        # Only enforce minimum design requirement during onboarding
                        if is_onboarding:
                            from common.business_config import BusinessConfig
                            minimum_required = BusinessConfig.get_minimum_required_designs_onboard()
                            if len(valid_design_folders) < minimum_required:
                                validation_errors.append(f'❌ Insufficient design folders: You have {len(valid_design_folders)} design folders, but a minimum of {minimum_required} is required. Please add {minimum_required - len(valid_design_folders)} more design folders.')
                        
                        # Validate folder_name mapping
                        zip_folder_names = set(valid_design_folders.keys())
                        missing_in_zip = excel_folders - zip_folder_names
                        missing_in_excel = zip_folder_names - excel_folders
                        
                        if missing_in_zip:
                            missing_list = list(missing_in_zip)[:10]
                            more_count = len(missing_in_zip) - 10 if len(missing_in_zip) > 10 else 0
                            validation_errors.append(f'❌ Folders listed in metadata.xlsx but not found in zip file: {", ".join(missing_list)}{f" (and {more_count} more)" if more_count > 0 else ""}. Please ensure these folders exist in your zip file.')
                        
                        if missing_in_excel:
                            missing_list = list(missing_in_excel)[:10]
                            more_count = len(missing_in_excel) - 10 if len(missing_in_excel) > 10 else 0
                            validation_errors.append(f'❌ Folders found in zip file but not listed in metadata.xlsx: {", ".join(missing_list)}{f" (and {more_count} more)" if more_count > 0 else ""}. Please add these folders to the "folder_name" column in your metadata.xlsx file.')
                        
                        # Validate each design folder has required files (should already be filtered, but double-check)
                        invalid_folders = []
                        for folder_name, files in valid_design_folders.items():
                            missing_files = required_files - files
                            if missing_files:
                                invalid_folders.append(f'{folder_name} (missing: {", ".join(missing_files)})')
                        
                        if invalid_folders:
                            more_count = len(invalid_folders) - 10 if len(invalid_folders) >= 10 else 0
                            validation_errors.append(f'❌ Some design folders are missing required files. Each folder must contain all 4 file types (.eps, .cdr, .jpg, .png). Affected folders: {"; ".join(invalid_folders[:10])}{f" (and {more_count} more)" if more_count > 0 else ""}')
                        
                        # Update zip_folders to only include valid design folders for count
                        zip_folders = valid_design_folders
                
                except Exception as e:
                    validation_errors.append(f'❌ Error reading metadata.xlsx: {str(e)}. The Excel file appears to be corrupted or in an invalid format. Please ensure it\'s a valid .xlsx file.')
        
        # If validation errors exist, return them
        if validation_errors:
            return Response({
                'error': 'Zip file validation failed',
                'validation_errors': validation_errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get total designs count (valid_design_folders should contain valid design folders at this point)
        total_designs = len(valid_design_folders)
        
        # Save zip file to storage
        try:
            # Generate unique file path
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            # Ensure user directory exists
            import os
            from django.conf import settings
            user_upload_dir = os.path.join(settings.MEDIA_ROOT, 'design_uploads', str(request.user.id))
            os.makedirs(user_upload_dir, exist_ok=True)
            
            zip_file_path = f'design_uploads/{request.user.id}/{timestamp}_{zip_file.name}'
            
            # Reset file pointer before saving
            zip_file.seek(0)
            # Save zip file to storage
            saved_path = default_storage.save(zip_file_path, zip_file)
            
            # Create DesignProcessingTask record
            with transaction.atomic():
                processing_task = DesignProcessingTask.objects.create(
                    user=request.user,
                    zip_file_path=saved_path,
                    total_designs=total_designs,
                    status='pending'
                )
                
                # Mark onboarding as completed immediately
                # Note: We set this to True here because Step 4 upload is complete
                # The actual product creation happens in background, but onboarding is considered complete
                try:
                    designer_profile = DesignerProfile.objects.get(created_by=request.user)
                    designer_profile.onboarding_completed = True
                    designer_profile.save(update_fields=['onboarding_completed', 'updated_at'])
                    # Don't call check_and_update_onboarding_status() here as it might override
                    # since products aren't created yet (they're being processed in background)
                except DesignerProfile.DoesNotExist:
                    # If DesignerProfile doesn't exist, create it
                    designer_profile = DesignerProfile.objects.create(
                        created_by=request.user,
                        onboarding_completed=True
                    )
                
                # Queue Celery task to process designs asynchronously
                process_design_upload_task.delay(processing_task.id, saved_path)
            
            # Verify onboarding_completed was set
            try:
                designer_profile = DesignerProfile.objects.get(created_by=request.user)
                onboarding_completed = designer_profile.onboarding_completed
            except DesignerProfile.DoesNotExist:
                onboarding_completed = False
            
            # Return success response immediately with task_id
            return Response({
                'message': 'Zip file uploaded successfully. Designs are being processed in the background.',
                'data': {
                    'task_id': processing_task.id,
                    'total_designs': total_designs,
                    'status': 'pending',
                    'onboarding_completed': onboarding_completed
                }
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            import traceback
            error_traceback = traceback.format_exc()
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error saving zip file: {str(e)}")
            logger.error(error_traceback)
            return Response({
                'error': f'Failed to save zip file: {str(e)}',
                'traceback': error_traceback if settings.DEBUG else None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    except zipfile.BadZipFile:
        return Response({
            'error': 'Invalid zip file format'
        }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Error in designer_onboarding_step4: {str(e)}")
        print(f"Traceback: {error_traceback}")
        return Response({
            'error': f'Failed to process zip file: {str(e)}',
            'traceback': error_traceback if settings.DEBUG else None
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Get Design Processing Progress',
    operation_description='Get the current progress of a design processing task',
    manual_parameters=[
        openapi.Parameter('task_id', openapi.IN_QUERY, description='Task ID', type=openapi.TYPE_INTEGER, required=True)
    ],
    responses={
        200: openapi.Response(description='Success'),
        404: openapi.Response(description='Task not found')
    },
    tags=['API']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_design_processing_progress(request):
    """
    Get the current progress of a design processing task.
    """
    task_id = request.GET.get('task_id')
    if not task_id:
        return Response({
            'error': 'task_id parameter is required'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        task = DesignProcessingTask.objects.get(id=task_id, user=request.user)
        serializer = DesignProcessingTaskSerializer(task)
        return Response({
            'data': serializer.data
        }, status=status.HTTP_200_OK)
    except DesignProcessingTask.DoesNotExist:
        return Response({
            'error': 'Task not found'
        }, status=status.HTTP_404_NOT_FOUND)


@swagger_auto_schema(
    method='get',
    operation_summary='Get Design Processing Status',
    operation_description='Get the latest design processing task status for the current user',
    responses={
        200: openapi.Response(description='Success'),
        404: openapi.Response(description='No task found')
    },
    tags=['API']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_design_processing_status(request):
    """
    Get the latest design processing task status for the current user.
    """
    try:
        task = DesignProcessingTask.objects.filter(user=request.user).order_by('-created_at').first()
        if not task:
            return Response({
                'data': None,
                'message': 'No processing task found'
            }, status=status.HTTP_200_OK)
        
        serializer = DesignProcessingTaskSerializer(task)
        return Response({
            'data': serializer.data
        }, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def stream_design_processing_progress(request):
    """
    Server-Sent Events endpoint for real-time design processing progress updates.
    Bypasses DRF's content negotiation to return raw SSE stream.
    Note: EventSource doesn't support custom headers, so token is passed as query parameter.
    """
    from django.http import HttpResponse
    import json
    import time
    
    # Check authentication manually (bypass DRF for SSE)
    # EventSource doesn't support custom headers, so we accept token as query parameter
    from rest_framework_simplejwt.authentication import JWTAuthentication
    from rest_framework.exceptions import AuthenticationFailed
    from rest_framework_simplejwt.tokens import UntypedToken
    from rest_framework_simplejwt.exceptions import InvalidToken, TokenError
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    user = None
    
    # Try to get token from query parameter first (for EventSource compatibility)
    # EventSource doesn't support custom headers, so we accept token as query parameter
    token = request.GET.get('token')
    
    if token:
        try:
            # Decode token directly to get user_id
            from rest_framework_simplejwt.tokens import AccessToken
            from rest_framework_simplejwt.exceptions import TokenError as JWTTokenError
            
            # Validate and decode the token
            access_token = AccessToken(token)
            user_id = access_token.get('user_id')
            
            if user_id:
                # Get user from database
                try:
                    user = User.objects.get(id=user_id)
                except User.DoesNotExist:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f'SSE: User {user_id} not found')
                    user = None
        except JWTTokenError as e:
            # Token is invalid (expired, malformed, etc.)
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f'SSE: Invalid token - {str(e)}')
            user = None
        except Exception as e:
            # Log unexpected errors (including database connection errors)
            import logging
            from django.db import OperationalError
            logger = logging.getLogger(__name__)
            if isinstance(e, OperationalError) and 'too many clients' in str(e):
                logger.error('SSE authentication failed: Database connection pool exhausted')
            else:
                logger.error(f'SSE authentication unexpected error: {str(e)}', exc_info=True)
            user = None
    
    # Fallback: Try JWT authentication from Authorization header if no token in query
    if not user:
        try:
            jwt_auth = JWTAuthentication()
            auth_result = jwt_auth.authenticate(request)
            if auth_result:
                user, _ = auth_result
        except (AuthenticationFailed, InvalidToken, TokenError) as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f'SSE JWT header authentication failed: {str(e)}')
            user = None
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'SSE JWT header authentication error: {str(e)}', exc_info=True)
            user = None
    
    # Don't try request.user fallback - it requires database access for session lookup
    # If JWT auth failed, return 401 immediately
    if not user:
        return HttpResponse(
            json.dumps({'error': 'Authentication required'}),
            status=401,
            content_type='application/json'
        )
    
    # Set the authenticated user
    request.user = user
    
    task_id = request.GET.get('task_id')
    if not task_id:
        return HttpResponse(
            json.dumps({'error': 'task_id parameter is required'}),
            status=400,
            content_type='application/json'
        )
    
    try:
        task = DesignProcessingTask.objects.get(id=task_id, user=user)
    except DesignProcessingTask.DoesNotExist:
        return HttpResponse(
            json.dumps({'error': 'Task not found'}),
            status=404,
            content_type='application/json'
        )
    
    def event_stream():
        """Generator function for SSE events."""
        last_processed = -1
        last_status = None
        
        while True:
            try:
                # Re-fetch task from database to get latest data
                try:
                    current_task = DesignProcessingTask.objects.get(id=task_id, user=user)
                except DesignProcessingTask.DoesNotExist:
                    error_data = {
                        'type': 'error',
                        'data': {
                            'error': 'Task not found'
                        }
                    }
                    yield f"data: {json.dumps(error_data)}\n\n"
                    break
                
                # Check if status or progress changed
                if current_task.processed_designs != last_processed or current_task.status != last_status:
                    # Send progress update
                    event_data = {
                        'type': 'progress',
                        'data': {
                            'status': current_task.status,
                            'processed': current_task.processed_designs,
                            'total': current_task.total_designs,
                            'failed': current_task.failed_designs,
                            'progress_percentage': current_task.progress_percentage
                        }
                    }
                    
                    yield f"data: {json.dumps(event_data)}\n\n"
                    
                    last_processed = current_task.processed_designs
                    last_status = current_task.status
                
                # If task is completed or failed, send final event and close
                if current_task.status in ['completed', 'failed']:
                    final_data = {
                        'type': 'complete',
                        'data': {
                            'status': current_task.status,
                            'processed': current_task.processed_designs,
                            'total': current_task.total_designs,
                            'failed': current_task.failed_designs,
                            'progress_percentage': current_task.progress_percentage,
                            'error_message': current_task.error_message if current_task.status == 'failed' else None
                        }
                    }
                    yield f"data: {json.dumps(final_data)}\n\n"
                    break
                
                # Wait before next check (poll every 2 seconds)
                time.sleep(2)
                
            except Exception as e:
                error_data = {
                    'type': 'error',
                    'data': {
                        'error': str(e)
                    }
                }
                yield f"data: {json.dumps(error_data)}\n\n"
                break
    
    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'  # Disable buffering in nginx
    return response


@swagger_auto_schema(
    method='get',
    operation_summary='Get Designer Onboarding Step 1 Data',
    operation_description='Get saved Step 1 data for prefill',
    responses={
        200: openapi.Response(description='Success'),
        404: openapi.Response(description='No data found')
    },
    tags=['API']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_designer_onboarding_step1(request):
    """
    Get saved Step 1 data for prefill.
    Returns email and phone verification status.
    """
    try:
        designer_profile = DesignerProfile.objects.get(created_by=request.user)
        user = request.user
        
        # Get primary email and check verification status
        email_obj = Email.objects.filter(created_by=user, is_primary=True).first()
        email = email_obj.email if email_obj else user.email
        email_verified = email_obj.is_verified if email_obj else False
        
        # Get primary mobile and check verification status
        mobile_obj = MobileNumber.objects.filter(created_by=user, is_primary=True).first()
        phone = mobile_obj.mobile_number if mobile_obj else ''
        phone_verified = mobile_obj.is_verified if mobile_obj else False
        
        # Get profile photo
        profile_photo_url = None
        from MediaFiles.models import Relation
        profile_photo_relations = Relation.objects.filter(
            relation_type='DesignerProfile:Media',
            id_1=designer_profile.pk
        )
        for relation in profile_photo_relations:
            if relation.meta and relation.meta.get('type') == 'profile_photo':
                try:
                    profile_photo = Media.objects.get(pk=relation.id_2)
                    if hasattr(profile_photo.file, 'url'):
                        # Build absolute URL for the media file
                        relative_url = profile_photo.file.url
                        if relative_url.startswith('/'):
                            profile_photo_url = request.build_absolute_uri(relative_url)
                        elif relative_url.startswith('http'):
                            profile_photo_url = relative_url
                        else:
                            # If relative URL doesn't start with /, prepend it
                            profile_photo_url = request.build_absolute_uri('/' + relative_url)
                        break
                except Media.DoesNotExist:
                    continue
        
        return Response({
            'data': {
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': email,
                'phone': phone,
                'is_individual': designer_profile.is_individual,
                'profile_photo_url': profile_photo_url,
                'email_verified': email_verified,
                'phone_verified': phone_verified
            }
        }, status=status.HTTP_200_OK)
        
    except DesignerProfile.DoesNotExist:
        return Response({
            'data': None,
            'message': 'No Step 1 data found'
        }, status=status.HTTP_404_NOT_FOUND)


@swagger_auto_schema(
    method='get',
    operation_summary='Get Designer Onboarding Step 2 Data',
    operation_description='Get saved Step 2 data for prefill',
    responses={
        200: openapi.Response(description='Success'),
        404: openapi.Response(description='No data found')
    },
    tags=['API']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_designer_onboarding_step2(request):
    """
    Get saved Step 2 data for prefill.
    """
    try:
        studio = Studio.objects.filter(created_by=request.user).first()
        if not studio:
            return Response({
                'data': None,
                'message': 'No Step 2 data found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        business_details = StudioBusinessDetails.objects.filter(studio=studio).first()
        if not business_details:
            return Response({
                'data': None,
                'message': 'No Step 2 data found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Extract address from registered_addresses_json
        address = business_details.registered_addresses_json or {}
        
        # Check if studio_email matches user's personal email and get verification status
        studio_email_verified = False
        if business_details.studio_email:
            email_obj = Email.objects.filter(
                email=business_details.studio_email,
                created_by=request.user,
                is_primary=True
            ).first()
            if email_obj:
                studio_email_verified = email_obj.is_verified
        
        # Check if studio_mobile_number matches user's personal phone and get verification status
        studio_mobile_verified = False
        if business_details.studio_mobile_number:
            mobile_obj = MobileNumber.objects.filter(
                mobile_number=business_details.studio_mobile_number,
                created_by=request.user,
                is_primary=True
            ).first()
            if mobile_obj:
                studio_mobile_verified = mobile_obj.is_verified
        
        # Build absolute URL for MSME certificate if exists
        msme_certificate_url = None
        if business_details.msme_certificate_annexure:
            if business_details.msme_certificate_annexure.startswith('/'):
                msme_certificate_url = request.build_absolute_uri(business_details.msme_certificate_annexure)
            elif business_details.msme_certificate_annexure.startswith('http'):
                msme_certificate_url = business_details.msme_certificate_annexure
            else:
                msme_certificate_url = request.build_absolute_uri('/' + business_details.msme_certificate_annexure)
        
        return Response({
            'data': {
                'legal_business_name': business_details.legal_business_name,
                'business_type': business_details.business_type,
                'business_model': business_details.business_model,
                'business_category': business_details.business_category,
                'business_sub_category': business_details.business_sub_category,
                'street': address.get('street', ''),
                'city': address.get('city', ''),
                'state': address.get('state', ''),
                'postal_code': address.get('postal_code', ''),
                'country': address.get('country', 'India'),
                'studio_email': business_details.studio_email,
                'studio_mobile_number': business_details.studio_mobile_number,
                'studio_email_verified': studio_email_verified,
                'studio_mobile_verified': studio_mobile_verified,
                'gst_number': business_details.gst_number or '',
                'msme_udyam_number': business_details.msme_udyam_number or '',
                'msme_certificate_annexure_url': msme_certificate_url
            }
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({
            'error': f'Failed to retrieve Step 2 data: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Get Designer Onboarding Step 3 Data',
    operation_description='Get saved Step 3 data for prefill',
    responses={
        200: openapi.Response(description='Success'),
        404: openapi.Response(description='No data found')
    },
    tags=['API']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_designer_onboarding_step3(request):
    """
    Get saved Step 3 data for prefill.
    Works for both individuals and companies.
    """
    try:
        # Check if designer profile exists
        designer_profile = DesignerProfile.objects.get(created_by=request.user)
        
        # Try to get studio (may not exist for individuals who haven't completed Step 3)
        studio = Studio.objects.filter(created_by=request.user).first()
        if not studio:
            return Response({
                'data': None,
                'message': 'No Step 3 data found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        business_details = StudioBusinessDetails.objects.filter(studio=studio).first()
        if not business_details or not business_details.pan_number:
            return Response({
                'data': None,
                'message': 'No Step 3 data found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        return Response({
            'data': {
                'pan_number': business_details.pan_number,
                'pan_card_url': business_details.pan_card
            }
        }, status=status.HTTP_200_OK)
        
    except DesignerProfile.DoesNotExist:
        return Response({
            'data': None,
            'message': 'No Step 3 data found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            'error': f'Failed to retrieve Step 3 data: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Get Designer Onboarding Step4',
    operation_description='Get saved Step 4 (bank details) data for prefill',
    responses={
        200: openapi.Response(description='Success'),
        404: openapi.Response(description='No data found')
    },
    tags=['API']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_designer_onboarding_step4(request):
    """
    Get saved Step 4 (bank details) data for prefill.
    Works for both individuals and companies.
    """
    try:
        # Find the studio for this designer
        studio = Studio.objects.filter(created_by=request.user).first()
        
        if not studio:
            return Response({
                'data': None,
                'message': 'No Step 4 data found. Please complete previous steps first.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        business_details = StudioBusinessDetails.objects.filter(studio=studio).first()
        if not business_details or not business_details.bank_account_number:
            return Response({
                'data': None,
                'message': 'No Step 4 data found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        return Response({
            'data': {
                'bank_account_number': business_details.bank_account_number,
                'bank_ifsc_code': business_details.bank_ifsc_code,
                'bank_account_holder_name': business_details.bank_account_holder_name,
                'account_type': business_details.account_type
            }
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({
            'error': f'Failed to retrieve Step 4 data: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Designer Onboarding Status',
    operation_description='Designer Onboarding Status endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def designer_onboarding_status(request):
    """
    Get designer onboarding status and linked account verification status.
    
    Handles three scenarios:
    1. Studio Member (no DesignerProfile) → Access studio owner's console with limited access
    2. Studio Owner (has DesignerProfile + owns studio) → Own console with full access
    3. Individual Designer (has DesignerProfile, no studio) → Own console with full access
    """
    # Check if user is a StudioMember
    is_studio_member = StudioMember.objects.filter(member=request.user, status='active').exists()
    
    # Check if user has DesignerProfile
    has_designer_profile = DesignerProfile.objects.filter(created_by=request.user).exists()
    
    # ========== SCENARIO 1: Studio Member (no DesignerProfile) ==========
    if is_studio_member and not has_designer_profile:
        # Get membership details
        membership = StudioMember.objects.filter(
            member=request.user, 
            status='active'
        ).select_related('studio', 'studio__created_by').first()
        
        if not membership:
            return Response({
                'error': 'Studio membership not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        studio = membership.studio
        studio_owner = studio.created_by
        
        # Build minimal onboarding status for studio member
        onboarding_status = {
            'step1_completed': True,  # User is authenticated
            'step2_completed': False,  # Members don't create studios
            'email_verified': Email.objects.filter(
                email=request.user.email, 
                created_by=request.user, 
                is_verified=True
            ).exists(),
            'mobile_verified': MobileNumber.objects.filter(
                created_by=request.user, 
                is_primary=True, 
                is_verified=True
            ).exists(),
            'designer_profile_status': None,  # No DesignerProfile
            'studio_created': False,  # Members don't own studios
            'is_studio_member': True,
            'business_details_completed': False,  # Not applicable
            'onboarding_completed': False,  # Members don't need onboarding
            'profile_type': 'member',
            'is_studio_owner': False,
            'has_full_console_access': False,  # Limited access
            'can_upload_designs': True  # Members can upload designs
        }
        
        # Get latest design processing task status (members can upload designs)
        latest_task = DesignProcessingTask.objects.filter(user=request.user).order_by('-created_at').first()
        if latest_task:
            onboarding_status['design_processing_status'] = {
                'task_id': latest_task.id,
                'status': latest_task.status,
                'total_designs': latest_task.total_designs,
                'processed_designs': latest_task.processed_designs,
                'failed_designs': latest_task.failed_designs,
                'progress_percentage': latest_task.progress_percentage
            }
        else:
            onboarding_status['design_processing_status'] = None
        
        # Studio context - which studio's console they're accessing
        # Include comprehensive studio information for dashboard display
        active_members_count = StudioMember.objects.filter(studio=studio, status='active').count()
        design_lead = StudioMember.objects.filter(studio=studio, role='design_lead', status='active').select_related('member').first()
        business_details = None
        try:
            business_details = StudioBusinessDetails.objects.get(studio=studio)
        except StudioBusinessDetails.DoesNotExist:
            pass
        
        # Format address from business details - include all address fields
        address = None
        if business_details and business_details.registered_addresses_json:
            addr = business_details.registered_addresses_json
            if isinstance(addr, dict):
                address_parts = []
                # Include all possible address fields (handle both formats)
                # Street/Address Line 1
                street = addr.get('address_line_1') or addr.get('street') or ''
                if street:
                    address_parts.append(street)
                
                # Address Line 2
                if addr.get('address_line_2'):
                    address_parts.append(addr['address_line_2'])
                
                # City
                if addr.get('city'):
                    address_parts.append(addr['city'])
                
                # State
                if addr.get('state'):
                    address_parts.append(addr['state'])
                
                # Pincode/Postal Code
                pincode = addr.get('pincode') or addr.get('postal_code') or ''
                if pincode:
                    address_parts.append(pincode)
                
                # Country
                if addr.get('country'):
                    address_parts.append(addr['country'])
                
                if address_parts:
                    # Filter out empty strings and join with commas
                    address = ', '.join([part for part in address_parts if part and part.strip()])
        
        accessing_studio_data = {
            'id': studio.id,
            'name': studio.name,
            'wedesignz_auto_name': studio.wedesignz_auto_name,
            'studio_industry_type': studio.studio_industry_type,
            'industry_display': dict(Studio.INDUSTRY_TYPE_CHOICES).get(studio.studio_industry_type, studio.studio_industry_type.replace('_', ' ').title()),
            'status': studio.status,
            'remarks': studio.remarks,
            'owner_id': studio_owner.id,
            'owner_name': f"{studio_owner.first_name} {studio_owner.last_name}".strip() or studio_owner.username,
            'owner_first_name': studio_owner.first_name or '',
            'owner_last_name': studio_owner.last_name or '',
            'owner_email': studio_owner.email or '',
            'total_active_members': active_members_count,
            'design_lead': None,
            'address': address
        }
        
        # Include design lead information if exists
        if design_lead and design_lead.member:
            accessing_studio_data['design_lead'] = {
                'id': design_lead.member.id,
                'first_name': design_lead.member.first_name or '',
                'last_name': design_lead.member.last_name or '',
                'email': design_lead.member.email or '',
                'full_name': f"{design_lead.member.first_name or ''} {design_lead.member.last_name or ''}".strip() or design_lead.member.email or design_lead.member.username
            }
        
        # Return response for studio member accessing studio owner's console
        return Response({
            'onboarding_status': onboarding_status,
            'can_access_console': True,  # ✅ Allow access to studio owner's console
            'profile_info': {
                'profile_type': 'member',
                'is_studio_owner': False,
                'is_studio_member': True,
                'owned_studio': None,  # Member doesn't own a studio
                'studio_membership': StudioMemberSerializer(membership).data,
                'has_full_console_access': False,  # Limited access
                'can_upload_designs': True,
                'accessing_studio': accessing_studio_data
            }
        })
    
    # ========== SCENARIO 2 & 3: Has DesignerProfile ==========
    # (Studio Owner or Individual Designer)
    try:
        designer_profile = DesignerProfile.objects.get(created_by=request.user)
        studio = Studio.objects.filter(created_by=request.user).first()
        
        # Note: Onboarding status is set manually, not automatically updated here
        # This prevents signals or automatic checks from overriding manual settings
        
        # Check if user is also a StudioMember (in case they're both owner and member of another studio)
        is_studio_member = StudioMember.objects.filter(member=request.user, status='active').exists()
        
        # Get profile type information
        is_owner = designer_profile.is_studio_owner()
        owned_studio = designer_profile.get_owned_studio()
        membership = designer_profile.get_studio_membership()
        
        onboarding_status = {
            'step1_completed': True,  # User is authenticated, so step 1 is complete
            'step2_completed': studio is not None,
            'email_verified': Email.objects.filter(email=request.user.email, created_by=request.user, is_verified=True).exists(),
            'mobile_verified': MobileNumber.objects.filter(created_by=request.user, is_primary=True, is_verified=True).exists(),
            'designer_profile_status': designer_profile.status,
            'studio_created': studio is not None,
            'is_studio_member': is_studio_member,
            'business_details_completed': False,
            'onboarding_completed': designer_profile.onboarding_completed,  # Use the boolean field
            # Profile type information
            'profile_type': designer_profile.profile_type,
            'is_studio_owner': is_owner,
            'has_full_console_access': designer_profile.has_full_console_access,
            'can_upload_designs': designer_profile.can_upload_designs
        }
        
        # Get latest design processing task status
        latest_task = DesignProcessingTask.objects.filter(user=request.user).order_by('-created_at').first()
        if latest_task:
            onboarding_status['design_processing_status'] = {
                'task_id': latest_task.id,
                'status': latest_task.status,
                'total_designs': latest_task.total_designs,
                'processed_designs': latest_task.processed_designs,
                'failed_designs': latest_task.failed_designs,
                'progress_percentage': latest_task.progress_percentage
            }
        else:
            onboarding_status['design_processing_status'] = None
        
        if studio:
            try:
                business_details = StudioBusinessDetails.objects.get(studio=studio)
                onboarding_status['business_details_completed'] = True
            except StudioBusinessDetails.DoesNotExist:
                pass
        
        # Allow console access if:
        # 1. User owns a studio and completed onboarding steps, OR
        # 2. User is an active StudioMember (they can access console to upload designs)
        can_access_console = (
            is_studio_member or  # Studio members can access console
            designer_profile.onboarding_completed  # Use the boolean field for performance
        )
        
        return Response({
            'onboarding_status': onboarding_status,
            'can_access_console': can_access_console,
            'profile_info': {
                'profile_type': designer_profile.profile_type,
                'is_studio_owner': is_owner,
                'is_studio_member': is_studio_member,
                'owned_studio': StudioSerializer(owned_studio).data if owned_studio else None,
                'studio_membership': StudioMemberSerializer(membership).data if membership else None,
                'has_full_console_access': designer_profile.has_full_console_access,
                'can_upload_designs': designer_profile.can_upload_designs
            }
        })
        
    except DesignerProfile.DoesNotExist:
        # User is not a studio member and has no DesignerProfile
        # They need to complete onboarding
        return Response({
            'error': 'Designer profile not found. Please complete onboarding.',
            'can_access_console': False
        }, status=status.HTTP_404_NOT_FOUND)


@swagger_auto_schema(
    method='get',
    operation_summary='Designer Dashboard',
    operation_description='Designer Dashboard endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def designer_dashboard(request):
    """
    Get designer dashboard with KPIs and analytics.
    Returns default/empty data if designer profile doesn't exist yet.
    For studio members: returns limited stats (no financial data)
    For studio owners: returns full stats including financial data
    """
    from django.db.models import Count, Sum, Avg, Q
    from Catalog.models import Product
    from Wallet.models import Wallet, WalletTransaction
    from Orders.models import Order
    from Profiles.models import Studio, StudioMember
    
    # Try to get designer profile, but don't fail if it doesn't exist
    try:
        designer_profile = DesignerProfile.objects.get(created_by=request.user)
        profile_data = DesignerProfileSerializer(designer_profile).data
        has_full_access = designer_profile.has_full_console_access
        is_studio_owner = designer_profile.is_studio_owner()
        is_studio_member = designer_profile.is_studio_member()
    except DesignerProfile.DoesNotExist:
        # Return default profile data if profile doesn't exist
        # Check if user is a studio member without DesignerProfile
        profile_data = None
        has_full_access = False
        is_studio_owner = Studio.objects.filter(created_by=request.user).exists()
        is_studio_member = StudioMember.objects.filter(member=request.user, status='active').exists()
    
    # Get products based on user type
    if is_studio_owner:
        # Studio owners: Get all designs from their studio
        studio = Studio.objects.filter(created_by=request.user).first()
        if studio:
            studio_members = StudioMember.objects.filter(studio=studio, status='active').values_list('member', flat=True)
            # Show designs:
            # 1. Created by studio owner directly (created_by = owner)
            # 2. Uploaded by any studio member (product_metadata['uploaded_by_member_id'] in studio_members)
            # Use Python filtering for JSONB queries to avoid type casting issues
            owner_products = Product.objects.filter(created_by=request.user)
            # For member products, use Python filtering
            all_products = Product.objects.exclude(product_metadata__isnull=True)
            member_product_ids = []
            for product in all_products.iterator(chunk_size=100):
                if product.product_metadata and isinstance(product.product_metadata, dict):
                    member_id = product.product_metadata.get('uploaded_by_member_id')
                    if member_id and member_id in list(studio_members):
                        member_product_ids.append(product.id)
            member_products = Product.objects.filter(id__in=member_product_ids) if member_product_ids else Product.objects.none()
            products = (owner_products | member_products).distinct()
        else:
            products = Product.objects.filter(created_by=request.user)
    elif is_studio_member:
        # Studio members: Only their own uploaded designs
        # Query by product_metadata['uploaded_by_member_id'] == request.user.id
        # Use Python filtering to avoid JSONB type casting issues with PostgreSQL
        all_products = Product.objects.exclude(product_metadata__isnull=True)
        product_ids = []
        for product in all_products.iterator(chunk_size=100):
            if product.product_metadata and isinstance(product.product_metadata, dict):
                if product.product_metadata.get('uploaded_by_member_id') == request.user.id:
                    product_ids.append(product.id)
        products = Product.objects.filter(id__in=product_ids) if product_ids else Product.objects.none()
    else:
        # Individual designer: Only their own designs
        products = Product.objects.filter(created_by=request.user)
    
    # Calculate KPIs
    total_uploaded = products.count()
    total_approved = products.filter(status='active').count()
    total_rejected = products.filter(status='inactive').count()  # Rejected designs have status='inactive'
    
    # TODO: Get downloads, views, purchases from analytics
    # For now, using placeholder values
    total_downloads = 0  # TODO: Calculate from analytics
    total_views = 0      # TODO: Calculate from analytics
    total_purchases = 0  # TODO: Calculate from analytics
    
    # Financial data - only for studio owners
    wallet_balance = 0
    monthly_earnings = 0
    lifetime_earnings = 0
    pending_withdrawals = 0
    performance_score = 0
    
    if has_full_access:
        # Get wallet information (only for studio owners)
        try:
            wallet = Wallet.objects.get(created_by=request.user)
            wallet_balance = wallet.balance
        except Wallet.DoesNotExist:
            wallet_balance = 0
        
        # Calculate monthly earnings (placeholder)
        monthly_earnings = 0  # TODO: Calculate from wallet transactions
        lifetime_earnings = 0  # TODO: Calculate from wallet transactions
        
        # Get pending withdrawals
        pending_withdrawals = 0  # TODO: Calculate from withdrawal requests
        
        # Calculate performance score
        # score = 0.2*norm(views) + 0.6*norm(purchases) + 0.2*norm(downloads)
        performance_score = 0  # TODO: Implement performance score calculation
    
    dashboard_data = {
        'designer_profile': profile_data,
        'kpis': {
            'total_uploaded': total_uploaded,
            'total_approved': total_approved,
            'total_rejected': total_rejected,
            'total_downloads': total_downloads if has_full_access else 0,
            'total_views': total_views if has_full_access else 0,
            'total_purchases': total_purchases if has_full_access else 0,
            'monthly_earnings': monthly_earnings,
            'lifetime_earnings': lifetime_earnings,
            'pending_withdrawals': pending_withdrawals,
            'performance_score': performance_score
        },
        'wallet_balance': wallet_balance
    }
    
    return Response(dashboard_data)


# ==================== STUDIO MANAGEMENT VIEWS ====================

@swagger_auto_schema(
    method='post',
    operation_summary='Create Studio',
    operation_description='Create Studio endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_studio(request):
    """
    Create a new studio with auto-generated unique name.
    """
    from common.studio_name_generator import generate_studio_name, generate_design_numbers
    
    # Validate required fields
    required_fields = ['name', 'studio_industry_type']
    for field in required_fields:
        if field not in request.data:
            return Response({
                'error': f'{field} is required'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Check if user already has a studio
        existing_studio = Studio.objects.filter(created_by=request.user).first()
        if existing_studio:
            return Response({
                'error': 'User already has a studio',
                'existing_studio': StudioSerializer(existing_studio).data
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Generate unique studio name
        auto_name = generate_studio_name(strategy="hybrid")
        if not auto_name:
            return Response({
                'error': 'Failed to generate unique studio name'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Create studio
        studio = Studio.objects.create(
            name=request.data['name'],
            wedesignz_auto_name=auto_name,
            studio_industry_type=request.data['studio_industry_type'],
            daily_design_generation_capacity=request.data.get('daily_design_generation_capacity', 0),
            remarks=request.data.get('remarks', ''),
            created_by=request.user
        )
        
        # Generate sample design numbers for preview
        sample_design_numbers = generate_design_numbers(auto_name)
        
        return Response({
            'message': 'Studio created successfully',
            'studio': StudioSerializer(studio).data,
            'auto_generated_name': auto_name,
            'sample_design_numbers': sample_design_numbers,
            'design_number_format': {
                'general_format': 'WDG00000001 (8 digits)',
                'studio_format': f'{auto_name[:2].upper()}0000001 (7 digits)'
            }
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return Response({
            'error': f'Failed to create studio: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Studio Design Number Info',
    operation_description='Studio Design Number Info endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def studio_design_number_info(request):
    """
    Get information about design number generation for a studio.
    """
    try:
        studio = Studio.objects.filter(created_by=request.user).first()
        if not studio:
            return Response({
                'error': 'No studio found for user'
            }, status=status.HTTP_404_NOT_FOUND)
        
        from common.studio_name_generator import generate_design_numbers
        
        # Generate sample numbers
        sample_numbers = generate_design_numbers(studio.wedesignz_auto_name)
        
        return Response({
            'studio': StudioSerializer(studio).data,
            'design_number_info': {
                'studio_auto_name': studio.wedesignz_auto_name,
                'studio_prefix': studio.wedesignz_auto_name[:2].upper(),
                'sample_general_number': sample_numbers['general_number'],
                'sample_studio_number': sample_numbers['studio_number'],
                'formats': {
                    'general': 'WDG00000001 (8 digits, global sequence)',
                    'studio': f'{studio.wedesignz_auto_name[:2].upper()}0000001 (7 digits, studio sequence)'
                }
            }
        })
        
    except Exception as e:
        return Response({
            'error': f'Failed to get design number info: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='post',
    operation_summary='Regenerate Studio Name',
    operation_description='Regenerate Studio Name endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def regenerate_studio_name(request):
    """
    Regenerate studio auto name (admin only or for testing).
    """
    try:
        studio = Studio.objects.filter(created_by=request.user).first()
        if not studio:
            return Response({
                'error': 'No studio found for user'
            }, status=status.HTTP_404_NOT_FOUND)
        
        from common.studio_name_generator import generate_studio_name
        
        # Generate new unique name
        new_auto_name = generate_studio_name(strategy="hybrid")
        if not new_auto_name:
            return Response({
                'error': 'Failed to generate new unique studio name'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Update studio
        old_name = studio.wedesignz_auto_name
        studio.wedesignz_auto_name = new_auto_name
        studio.save()
        
        return Response({
            'message': 'Studio name regenerated successfully',
            'old_name': old_name,
            'new_name': new_auto_name,
            'studio': StudioSerializer(studio).data
        })
        
    except Exception as e:
        return Response({
            'error': f'Failed to regenerate studio name: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
