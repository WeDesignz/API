from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Sum, Q
from django.utils import timezone
from django.http import HttpResponse
from django.contrib.auth.models import User
from .models import Wallet, WalletTransaction, WalletWithdrawalRequest, SettlementRequest
from .serializers import (
    WalletSerializer, WalletTransactionSerializer, WalletWithdrawalRequestSerializer
)
from CoreAdmin.auth import admin_required
from common.relations import get_related
import csv
import io
from datetime import datetime, date
import pytz
from decimal import Decimal


@swagger_auto_schema(
    method='get',
    operation_summary='Get Wallet Balance',
    operation_description='Get user wallet balance and details.',
    responses={
        200: openapi.Response(
            description='Wallet balance retrieved successfully',
            examples={
                'application/json': {
                    'wallet': {
                        'id': 1,
                        'balance': 150.75,
                        'created_at': '2024-01-01T00:00:00Z',
                        'updated_at': '2024-01-01T00:00:00Z'
                    },
                    'balance': 150.75
                }
            }
        ),
        401: openapi.Response(description='Unauthorized - authentication required')
    },
    tags=['Wallet']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def wallet_balance(request):
    """
    Get user's wallet balance and details.
    """
    try:
        wallet = Wallet.objects.get(created_by=request.user)
    except Wallet.DoesNotExist:
        # Create wallet if it doesn't exist
        wallet = Wallet.objects.create(created_by=request.user)
    
    try:
        wallet_data = WalletSerializer(wallet).data
    except Exception as e:
        # If serialization fails, return minimal wallet data
        wallet_data = {
            'id': wallet.id,
            'balance': float(wallet.balance),
            'created_at': wallet.created_at.isoformat() if wallet.created_at else None,
            'updated_at': wallet.updated_at.isoformat() if wallet.updated_at else None,
            'transactions': [],
            'withdrawal_requests': []
        }
    
    return Response({
        'wallet': wallet_data,
        'balance': float(wallet.balance)
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Wallet Transactions',
    operation_description='Wallet Transactions endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def wallet_transactions(request):
    """
    Get user's wallet transaction history.
    """
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        wallet = Wallet.objects.get(created_by=request.user)
    except Wallet.DoesNotExist:
        wallet = Wallet.objects.create(created_by=request.user)
    
    # Get transactions both by created_by and through wallet relations
    from common.relations import get_related
    
    # Get transactions created by user
    transactions_by_user = WalletTransaction.objects.filter(created_by=request.user)
    
    # Get transactions linked to wallet through relations
    transactions_by_wallet = get_related(wallet, 'Wallet:WalletTransaction', WalletTransaction)
    
    # Combine both querysets and remove duplicates
    all_transaction_ids = set(transactions_by_user.values_list('id', flat=True))
    all_transaction_ids.update(transactions_by_wallet.values_list('id', flat=True))
    
    # Get all unique transactions ordered by created_at
    transactions = WalletTransaction.objects.filter(id__in=all_transaction_ids).order_by('-created_at')
    
    # Log transaction count for debugging
    transaction_count = transactions.count()
    logger.info(f"Wallet transactions query for user {request.user.id}: Found {transaction_count} transactions (by user: {transactions_by_user.count()}, by wallet: {transactions_by_wallet.count()})")
    
    # Get transaction summary
    total_credit = transactions.filter(wallet_transaction_type='credit').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    total_debit = transactions.filter(wallet_transaction_type='debit').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    try:
        wallet_data = WalletSerializer(wallet).data
    except Exception as e:
        logger.error(f"Failed to serialize wallet data: {str(e)}", exc_info=True)
        # If serialization fails, return minimal wallet data
        wallet_data = {
            'id': wallet.id,
            'balance': float(wallet.balance),
            'created_at': wallet.created_at.isoformat() if wallet.created_at else None,
            'updated_at': wallet.updated_at.isoformat() if wallet.updated_at else None,
            'transactions': [],
            'withdrawal_requests': []
        }
    
    try:
        transactions_data = WalletTransactionSerializer(transactions, many=True).data
        logger.info(f"Successfully serialized {len(transactions_data)} transactions")
    except Exception as e:
        logger.error(f"Failed to serialize wallet transactions: {str(e)}", exc_info=True)
        # If serialization fails, return empty list
        transactions_data = []
    
    response_data = {
        'wallet': wallet_data,
        'transactions': transactions_data,
        'summary': {
            'total_credit': float(total_credit),
            'total_debit': float(total_debit),
            'current_balance': float(wallet.balance),
            'total_transactions': transaction_count
        }
    }
    
    logger.info(f"Returning wallet transactions response with {len(transactions_data)} transactions")
    
    return Response(response_data)


@swagger_auto_schema(
    method='post',
    operation_summary='Add Wallet Balance',
    operation_description='Add money to user wallet balance.',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'amount': openapi.Schema(
                type=openapi.TYPE_NUMBER,
                description='Amount to add to wallet',
                example=50.00
            ),
            'payment_method': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Payment method',
                example='razorpay',
                enum=['razorpay', 'bank_transfer']
            ),
            'description': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Transaction description',
                example='Wallet top-up'
            )
        },
        required=['amount', 'payment_method']
    ),
    responses={
        201: openapi.Response(
            description='Balance added successfully',
            examples={
                'application/json': {
                    'message': 'Balance added successfully',
                    'transaction': {
                        'id': 1,
                        'amount': 50.00,
                        'type': 'credit',
                        'description': 'Wallet top-up',
                        'created_at': '2024-01-01T00:00:00Z'
                    },
                    'new_balance': 200.75
                }
            }
        ),
        400: openapi.Response(description='Bad request - invalid amount or payment method'),
        401: openapi.Response(description='Unauthorized - authentication required')
    },
    tags=['Wallet']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_wallet_balance(request):
    """
    Add balance to user's wallet.
    """
    amount = request.data.get('amount')
    description = request.data.get('description', 'Wallet top-up')
    reference_id = request.data.get('reference_id')
    
    if not amount or float(amount) <= 0:
        return Response({
            'error': 'Valid amount is required'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        wallet = Wallet.objects.get(created_by=request.user)
    except Wallet.DoesNotExist:
        wallet = Wallet.objects.create(created_by=request.user)
    
    # Create credit transaction
    transaction = WalletTransaction.objects.create(
        wallet_transaction_type='credit',
        amount=amount,
        description=description,
        reference_id=reference_id,
        created_by=request.user
    )
    
    # Update wallet balance
    wallet.balance += float(amount)
    wallet.save()
    
    # Attach transaction to wallet
    wallet.attach_wallet_transaction(transaction)
    
    return Response({
        'message': 'Balance added successfully',
        'transaction': WalletTransactionSerializer(transaction).data,
        'new_balance': float(wallet.balance)
    }, status=status.HTTP_201_CREATED)


@swagger_auto_schema(
    method='post',
    operation_summary='Deduct Wallet Balance',
    operation_description='Deduct Wallet Balance endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def deduct_wallet_balance(request):
    """
    Deduct balance from user's wallet.
    """
    amount = request.data.get('amount')
    description = request.data.get('description', 'Wallet deduction')
    reference_id = request.data.get('reference_id')
    
    if not amount or float(amount) <= 0:
        return Response({
            'error': 'Valid amount is required'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        wallet = Wallet.objects.get(created_by=request.user)
    except Wallet.DoesNotExist:
        wallet = Wallet.objects.create(created_by=request.user)
    
    # Check if sufficient balance
    if wallet.balance < float(amount):
        return Response({
            'error': 'Insufficient wallet balance'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Create debit transaction
    transaction = WalletTransaction.objects.create(
        wallet_transaction_type='debit',
        amount=amount,
        description=description,
        reference_id=reference_id,
        created_by=request.user
    )
    
    # Update wallet balance
    wallet.balance -= float(amount)
    wallet.save()
    
    # Attach transaction to wallet
    wallet.attach_wallet_transaction(transaction)
    
    return Response({
        'message': 'Balance deducted successfully',
        'transaction': WalletTransactionSerializer(transaction).data,
        'new_balance': float(wallet.balance)
    }, status=status.HTTP_201_CREATED)


@swagger_auto_schema(
    method='post',
    operation_summary='Create Withdrawal Request',
    operation_description='Create a withdrawal request from wallet balance.',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'amount': openapi.Schema(
                type=openapi.TYPE_NUMBER,
                description='Amount to withdraw',
                example=100.00
            ),
            'reason': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Reason for withdrawal',
                example='Monthly payout request'
            )
        },
        required=['amount']
    ),
    responses={
        201: openapi.Response(
            description='Withdrawal request created successfully',
            examples={
                'application/json': {
                    'message': 'Withdrawal request created successfully',
                    'withdrawal_request': {
                        'id': 1,
                        'amount': 100.00,
                        'status': 'pending',
                        'reason': 'Monthly payout request',
                        'created_at': '2024-01-01T00:00:00Z'
                    }
                }
            }
        ),
        400: openapi.Response(description='Bad request - insufficient balance or pending request exists'),
        401: openapi.Response(description='Unauthorized - authentication required')
    },
    tags=['Wallet']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_withdrawal_request(request):
    """
    Create a withdrawal request from wallet.
    """
    amount = request.data.get('amount')
    reason = request.data.get('reason', '')
    
    if not amount or float(amount) <= 0:
        return Response({
            'error': 'Valid amount is required'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        wallet = Wallet.objects.get(created_by=request.user)
    except Wallet.DoesNotExist:
        wallet = Wallet.objects.create(created_by=request.user)
    
    # Check if sufficient balance
    if wallet.balance < float(amount):
        return Response({
            'error': 'Insufficient wallet balance'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Check for pending withdrawal requests
    pending_requests = WalletWithdrawalRequest.objects.filter(
        wallet=wallet,
        status='pending'
    ).count()
    
    if pending_requests > 0:
        return Response({
            'error': 'You already have a pending withdrawal request'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Create withdrawal request
    withdrawal_request = WalletWithdrawalRequest.objects.create(
        wallet=wallet,
        amount=amount,
        reason=reason,
        created_by=request.user
    )
    
    return Response({
        'message': 'Withdrawal request created successfully',
        'withdrawal_request': WalletWithdrawalRequestSerializer(withdrawal_request).data
    }, status=status.HTTP_201_CREATED)


@swagger_auto_schema(
    method='get',
    operation_summary='Withdrawal Requests',
    operation_description='Withdrawal Requests endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def withdrawal_requests(request):
    """
    Get user's withdrawal requests.
    """
    try:
        wallet = Wallet.objects.get(created_by=request.user)
    except Wallet.DoesNotExist:
        wallet = Wallet.objects.create(created_by=request.user)
    
    requests = WalletWithdrawalRequest.objects.filter(
        wallet=wallet
    ).order_by('-created_at')
    
    return Response({
        'withdrawal_requests': WalletWithdrawalRequestSerializer(requests, many=True).data,
        'total_requests': requests.count()
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Withdrawal Request Detail',
    operation_description='Withdrawal Request Detail endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def withdrawal_request_detail(request, request_id):
    """
    Get details of a specific withdrawal request.
    """
    try:
        withdrawal_request = WalletWithdrawalRequest.objects.get(
            id=request_id,
            wallet__created_by=request.user
        )
        return Response({
            'withdrawal_request': WalletWithdrawalRequestSerializer(withdrawal_request).data
        })
    except WalletWithdrawalRequest.DoesNotExist:
        return Response({
            'error': 'Withdrawal request not found'
        }, status=status.HTTP_404_NOT_FOUND)


@swagger_auto_schema(
    method='post',
    operation_summary='Cancel Withdrawal Request',
    operation_description='Cancel Withdrawal Request endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def cancel_withdrawal_request(request, request_id):
    """
    Cancel a pending withdrawal request.
    """
    try:
        withdrawal_request = WalletWithdrawalRequest.objects.get(
            id=request_id,
            wallet__created_by=request.user,
            status='pending'
        )
        
        withdrawal_request.status = 'rejected'
        withdrawal_request.admin_remarks = 'Cancelled by user'
        withdrawal_request.save()
        
        return Response({
            'message': 'Withdrawal request cancelled successfully',
            'withdrawal_request': WalletWithdrawalRequestSerializer(withdrawal_request).data
        })
    
    except WalletWithdrawalRequest.DoesNotExist:
        return Response({
            'error': 'Withdrawal request not found or cannot be cancelled'
        }, status=status.HTTP_404_NOT_FOUND)


@swagger_auto_schema(
    method='get',
    operation_summary='Wallet Summary',
    operation_description='Wallet Summary endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def wallet_summary(request):
    """
    Get comprehensive wallet summary.
    """
    try:
        wallet = Wallet.objects.get(created_by=request.user)
    except Wallet.DoesNotExist:
        wallet = Wallet.objects.create(created_by=request.user)
    
    # Get transaction statistics
    transactions = WalletTransaction.objects.filter(created_by=request.user)
    
    credit_transactions = transactions.filter(wallet_transaction_type='credit')
    debit_transactions = transactions.filter(wallet_transaction_type='debit')
    
    total_credit = credit_transactions.aggregate(total=Sum('amount'))['total'] or 0
    total_debit = debit_transactions.aggregate(total=Sum('amount'))['total'] or 0
    
    # Get withdrawal requests
    withdrawal_requests = WalletWithdrawalRequest.objects.filter(wallet=wallet)
    pending_withdrawals = withdrawal_requests.filter(status='pending').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    try:
        wallet_data = WalletSerializer(wallet).data
    except Exception as e:
        # If serialization fails, return minimal wallet data
        wallet_data = {
            'id': wallet.id,
            'balance': float(wallet.balance),
            'created_at': wallet.created_at.isoformat() if wallet.created_at else None,
            'updated_at': wallet.updated_at.isoformat() if wallet.updated_at else None,
            'transactions': [],
            'withdrawal_requests': []
        }
    
    return Response({
        'wallet': wallet_data,
        'summary': {
            'current_balance': float(wallet.balance),
            'total_credit': float(total_credit),
            'total_debit': float(total_debit),
            'pending_withdrawals': float(pending_withdrawals),
            'total_transactions': transactions.count(),
            'credit_transactions': credit_transactions.count(),
            'debit_transactions': debit_transactions.count(),
            'withdrawal_requests': withdrawal_requests.count()
        }
    })


@swagger_auto_schema(
    method='get',
    operation_summary='Recent Transactions',
    operation_description='Recent Transactions endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def recent_transactions(request):
    """
    Get recent wallet transactions.
    """
    try:
        wallet = Wallet.objects.get(created_by=request.user)
    except Wallet.DoesNotExist:
        wallet = Wallet.objects.create(created_by=request.user)
    
    limit = int(request.GET.get('limit', 10))
    transactions = WalletTransaction.objects.filter(
        created_by=request.user
    ).order_by('-created_at')[:limit]
    
    try:
        transactions_data = WalletTransactionSerializer(transactions, many=True).data
    except Exception as e:
        # If serialization fails, return empty list
        transactions_data = []
    
    return Response({
        'recent_transactions': transactions_data,
        'total_count': transactions.count()
    })


@swagger_auto_schema(
    method='post',
    operation_summary='Transfer To Wallet',
    operation_description='Transfer To Wallet endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def transfer_to_wallet(request):
    """
    Transfer amount from one user's wallet to another.
    """
    recipient_username = request.data.get('recipient_username')
    amount = request.data.get('amount')
    description = request.data.get('description', 'Wallet transfer')
    
    if not all([recipient_username, amount]):
        return Response({
            'error': 'Recipient username and amount are required'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if float(amount) <= 0:
        return Response({
            'error': 'Valid amount is required'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        from django.contrib.auth.models import User
        recipient = User.objects.get(username=recipient_username)
    except User.DoesNotExist:
        return Response({
            'error': 'Recipient user not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    # Get sender's wallet
    try:
        sender_wallet = Wallet.objects.get(created_by=request.user)
    except Wallet.DoesNotExist:
        sender_wallet = Wallet.objects.create(created_by=request.user)
    
    # Check sender's balance
    if sender_wallet.balance < float(amount):
        return Response({
            'error': 'Insufficient wallet balance'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Get or create recipient's wallet
    try:
        recipient_wallet = Wallet.objects.get(created_by=recipient)
    except Wallet.DoesNotExist:
        recipient_wallet = Wallet.objects.create(created_by=recipient)
    
    # Create debit transaction for sender
    sender_transaction = WalletTransaction.objects.create(
        wallet_transaction_type='debit',
        amount=amount,
        description=f"Transfer to {recipient_username}: {description}",
        reference_id=f"transfer_{request.user.id}_{recipient.id}",
        created_by=request.user
    )
    
    # Create credit transaction for recipient
    recipient_transaction = WalletTransaction.objects.create(
        wallet_transaction_type='credit',
        amount=amount,
        description=f"Transfer from {request.user.username}: {description}",
        reference_id=f"transfer_{request.user.id}_{recipient.id}",
        created_by=recipient
    )
    
    # Update wallet balances
    sender_wallet.balance -= float(amount)
    sender_wallet.save()
    
    recipient_wallet.balance += float(amount)
    recipient_wallet.save()
    
    # Attach transactions to wallets
    sender_wallet.attach_wallet_transaction(sender_transaction)
    recipient_wallet.attach_wallet_transaction(recipient_transaction)
    
    return Response({
        'message': 'Transfer completed successfully',
        'sender_transaction': WalletTransactionSerializer(sender_transaction).data,
        'recipient_transaction': WalletTransactionSerializer(recipient_transaction).data,
        'sender_balance': float(sender_wallet.balance)
    }, status=status.HTTP_201_CREATED)


# ==================== DESIGNER CONSOLE - SETTLEMENT & PAYOUTS ====================

@swagger_auto_schema(
    method='get',
    operation_summary='Settlement Status',
    operation_description='Settlement Status endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def settlement_status(request):
    """
    Get current settlement status and check if settlement window is active.
    Settlement window: Days 1-5 of each month
    """
    from datetime import datetime, date
    import pytz
    from Wallet.models import SettlementRequest
    from Authentication.user_relations import get_user_wallets
    
    # Get current date in Asia/Kolkata timezone
    kolkata_tz = pytz.timezone('Asia/Kolkata')
    current_date = datetime.now(kolkata_tz).date()
    current_day = current_date.day
    
    # Check if we're in settlement window (days 1-5)
    settlement_window_active = 1 <= current_day <= 5
    
    # Calculate current period (previous month)
    if current_date.month == 1:
        period_start = date(current_date.year - 1, 12, 1)
    else:
        period_start = date(current_date.year, current_date.month - 1, 1)
    
    # Get settlement request for current period
    settlement_request = SettlementRequest.objects.filter(
        designer_id=request.user.id,
        settlement_period_start=period_start
    ).first()
    
    # Get current wallet balance
    wallets = get_user_wallets(request.user)
    current_wallet_balance = wallets.first().balance if wallets.exists() else 0
    
    # Check if bank details are provided (required for settlement)
    # Get bank details from StudioBusinessDetails
    from Profiles.models import Studio, StudioBusinessDetails
    studio = Studio.objects.filter(created_by=request.user).first()
    has_bank_details = False
    if studio:
        business_details = StudioBusinessDetails.objects.filter(studio=studio).first()
        has_bank_details = business_details and business_details.bank_account_number and business_details.bank_ifsc_code and business_details.bank_account_holder_name if business_details else False
    
    settlement_data = {
        'settlement_window_active': settlement_window_active,
        'current_day': current_day,
        'settlement_window_days': [1, 2, 3, 4, 5],
        'settlement_request': None,
        'can_accept_settlement': False,
        'has_bank_details': has_bank_details,
        'current_wallet_balance': float(current_wallet_balance)
    }
    
    if settlement_request:
        settlement_data['settlement_request'] = {
            'id': settlement_request.id,
            'settlement_period_start': settlement_request.settlement_period_start.isoformat(),
            'settlement_period_end': settlement_request.settlement_period_end.isoformat(),
            'wallet_balance_at_period_end': float(settlement_request.wallet_balance_at_period_end),
            'settlement_amount': float(settlement_request.settlement_amount),
            'status': settlement_request.status,
            'opted_in': settlement_request.opted_in,
            'opted_in_at': settlement_request.opted_in_at.isoformat() if settlement_request.opted_in_at else None,
            'settlement_date': settlement_request.settlement_date.isoformat() if settlement_request.settlement_date else None
        }
        
        if settlement_window_active and settlement_request.status == 'pending':
            settlement_data['can_accept_settlement'] = has_bank_details
    
    return Response(settlement_data)


@swagger_auto_schema(
    method='post',
    operation_summary='Accept Settlement',
    operation_description='Accept Settlement endpoint',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'data': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Request data'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def accept_settlement(request):
    """
    Designer opts in to settlement during the settlement window (days 1-5).
    """
    from datetime import datetime, date
    import pytz
    from django.utils import timezone
    from Wallet.models import SettlementRequest
    
    # Check if we're in settlement window (days 1-5)
    kolkata_tz = pytz.timezone('Asia/Kolkata')
    current_date = datetime.now(kolkata_tz)
    current_day = current_date.day
    
    if not (1 <= current_day <= 5):
        return Response({
            'error': 'Settlement window is not active. Settlement can only be accepted between days 1-5 of each month.'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Calculate current period (previous month)
    today = current_date.date()
    if today.month == 1:
        period_start = date(today.year - 1, 12, 1)
    else:
        period_start = date(today.year, today.month - 1, 1)
    
    # Get pending settlement request for current period
    try:
        settlement_request = SettlementRequest.objects.get(
            designer_id=request.user.id,
            settlement_period_start=period_start,
            status='pending'
        )
    except SettlementRequest.DoesNotExist:
        return Response({
            'error': 'No pending settlement request found for this period'
        }, status=status.HTTP_404_NOT_FOUND)
    
    # Check if bank details are provided
    # Get bank details from StudioBusinessDetails
    from Profiles.models import Studio, StudioBusinessDetails
    studio = Studio.objects.filter(created_by=request.user).first()
    
    if not studio:
        return Response({
            'error': 'Studio not found. Please complete onboarding first.'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    business_details = StudioBusinessDetails.objects.filter(studio=studio).first()
    if not business_details or not business_details.bank_account_number or not business_details.bank_ifsc_code or not business_details.bank_account_holder_name:
        return Response({
            'error': 'Bank account details are incomplete. Please provide bank account number, IFSC code, and account holder name.'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Opt in to settlement
        settlement_request.opted_in = True
        settlement_request.opted_in_at = timezone.now()
        settlement_request.status = 'opted_in'
        settlement_request.save()
        
        # Schedule async task to generate and send monthly bill
        try:
            from common.tasks import generate_and_send_designer_bill_async
            import logging
            logger = logging.getLogger(__name__)
            
            generate_and_send_designer_bill_async.delay(
                designer_id=request.user.id,
                settlement_period_start=settlement_request.settlement_period_start.isoformat(),
                settlement_period_end=settlement_request.settlement_period_end.isoformat(),
                settlement_request_id=settlement_request.id
            )
            logger.info(f"Scheduled bill generation task for designer {request.user.id} for settlement {settlement_request.id}")
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Failed to queue bill generation task for designer {request.user.id}: {str(e)}', exc_info=True)
            # Don't fail the settlement opt-in if bill generation fails
        
        return Response({
            'message': 'Settlement accepted successfully',
            'settlement_request': {
                'id': settlement_request.id,
                'amount': float(settlement_request.settlement_amount),
                'period_start': settlement_request.settlement_period_start.isoformat(),
                'period_end': settlement_request.settlement_period_end.isoformat(),
                'status': settlement_request.status,
                'settlement_date': 'Day 6 of current month'  # Will be processed on Day 6
            }
        })
        
    except Exception as e:
        return Response({
            'error': f'Failed to accept settlement: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method='get',
    operation_summary='Settlement History',
    operation_description='Settlement History endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def settlement_history(request):
    """
    Get settlement history for the designer.
    """
    # TODO: Get settlement history from SettlementRequest model
    # settlements = SettlementRequest.objects.filter(
    #     designer=request.user
    # ).order_by('-created_at')
    
    settlement_history = {
        'settlements': [],  # TODO: Serialize settlement requests
        'total_settlements': 0,
        'total_paid': 0,
        'pending_settlements': 0
    }
    
    return Response(settlement_history)


@swagger_auto_schema(
    method='get',
    operation_summary='Payout History',
    operation_description='Payout History endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def payout_history(request):
    """
    Get payout history for the designer.
    """
    # TODO: Get payout history from Payout model
    # payouts = Payout.objects.filter(
    #     designer=request.user
    # ).order_by('-created_at')
    
    payout_history = {
        'payouts': [],  # TODO: Serialize payouts
        'total_payouts': 0,
        'total_amount_paid': 0,
        'pending_payouts': 0
    }
    
    return Response(payout_history)


@swagger_auto_schema(
    method='get',
    operation_summary='Earnings Summary',
    operation_description='Earnings Summary endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def earnings_summary(request):
    """
    Get earnings summary for the designer.
    """
    from django.db.models import Sum
    from datetime import datetime, timedelta
    
    # Get current month earnings
    current_month = datetime.now().replace(day=1)
    monthly_earnings = WalletTransaction.objects.filter(
        created_by=request.user,
        wallet_transaction_type='credit',
        created_at__gte=current_month
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Get lifetime earnings
    lifetime_earnings = WalletTransaction.objects.filter(
        created_by=request.user,
        wallet_transaction_type='credit'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Get pending withdrawals
    pending_withdrawals = WalletWithdrawalRequest.objects.filter(
        wallet__created_by=request.user,
        status='pending'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # TODO: Get earnings from last withdrawal to now for settlement
    # last_withdrawal_date = get_last_withdrawal_date(request.user)
    # earnings_since_last_withdrawal = calculate_earnings_since(request.user, last_withdrawal_date)
    
    earnings_summary = {
        'monthly_earnings': float(monthly_earnings),
        'lifetime_earnings': float(lifetime_earnings),
        'pending_withdrawals': float(pending_withdrawals),
        'current_wallet_balance': float(Wallet.objects.get(created_by=request.user).balance),
        'earnings_since_last_settlement': 0,  # TODO: Calculate
        'platform_fee_percentage': 10,  # TODO: Get from settings
        'next_settlement_eligible': True  # TODO: Check if eligible for next settlement
    }
    
    return Response(earnings_summary)


@swagger_auto_schema(
    method='get',
    operation_summary='Linked Account Status',
    operation_description='Linked Account Status endpoint',
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'message': 'Success',
                    'data': {}
                }
            }
        ),
        400: openapi.Response(description='Bad request')
    },
    tags=['API']
)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def linked_account_status(request):
    """
    Get Razorpay linked account status and verification details.
    """
    # TODO: Get linked account status from Razorpay
    # account_status = razorpay_service.get_linked_account_status(request.user)
    
    account_status = {
        'account_created': False,  # TODO: Check if account exists
        'verification_status': 'pending',  # TODO: Get from Razorpay
        'verification_documents_required': [],  # TODO: Get required documents
        'verification_documents_submitted': [],  # TODO: Get submitted documents
        'account_id': None,  # TODO: Get Razorpay account ID
        'payouts_enabled': False,  # TODO: Check if payouts are enabled
        'verification_errors': []  # TODO: Get any verification errors
    }
    
    return Response(account_status)


# ==================== ADMIN - SETTLEMENT SHEET DOWNLOAD ====================

@swagger_auto_schema(
    method='get',
    operation_summary='Download Settlement Sheet',
    operation_description='Download settlement sheet as CSV or Excel file (Admin only). Filter by status, period, or settlement date.',
    manual_parameters=[
        openapi.Parameter(
            'format',
            openapi.IN_QUERY,
            description='File format (csv or xlsx)',
            type=openapi.TYPE_STRING,
            enum=['csv', 'xlsx'],
            default='xlsx'
        ),
        openapi.Parameter(
            'status',
            openapi.IN_QUERY,
            description='Filter by settlement status (pending, opted_in, processing, completed, failed, expired)',
            type=openapi.TYPE_STRING
        ),
        openapi.Parameter(
            'period_start',
            openapi.IN_QUERY,
            description='Filter by settlement period start date (YYYY-MM-DD)',
            type=openapi.TYPE_STRING
        ),
        openapi.Parameter(
            'settlement_date',
            openapi.IN_QUERY,
            description='Filter by settlement date (YYYY-MM-DD)',
            type=openapi.TYPE_STRING
        ),
    ],
    responses={
        200: openapi.Response(
            description='Settlement sheet file',
            content={
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': {},
                'text/csv': {}
            }
        ),
        403: openapi.Response(description='Access denied - Admin privileges required'),
        400: openapi.Response(description='Invalid parameters')
    },
    tags=['Wallet Admin']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@admin_required()
def download_settlement_sheet(request):
    """
    Download settlement sheet as CSV or Excel file.
    Admin-only endpoint to generate settlement reports for manual payouts.
    """
    try:
        # Get query parameters
        file_format = request.GET.get('format', 'xlsx').lower()
        status_filter = request.GET.get('status')
        period_start_str = request.GET.get('period_start')
        settlement_date_str = request.GET.get('settlement_date')
        
        # Validate format
        if file_format not in ['csv', 'xlsx']:
            return Response({
                'error': 'Invalid format. Use "csv" or "xlsx"'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Build query
        queryset = SettlementRequest.objects.all()
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        if period_start_str:
            try:
                period_start = datetime.strptime(period_start_str, '%Y-%m-%d').date()
                queryset = queryset.filter(settlement_period_start=period_start)
            except ValueError:
                return Response({
                    'error': 'Invalid period_start format. Use YYYY-MM-DD'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        if settlement_date_str:
            try:
                settlement_date = datetime.strptime(settlement_date_str, '%Y-%m-%d').date()
                queryset = queryset.filter(settlement_date=settlement_date)
            except ValueError:
                return Response({
                    'error': 'Invalid settlement_date format. Use YYYY-MM-DD'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Order by settlement period and designer ID
        queryset = queryset.order_by('-settlement_period_start', 'designer_id')
        
        # Prepare data
        settlement_data = []
        total_amount = Decimal('0.00')
        
        for settlement in queryset:
            try:
                # Get designer
                designer = settlement.designer
                if not designer:
                    continue
                
                # Get bank details from StudioBusinessDetails
                from Profiles.models import Studio, StudioBusinessDetails
                studio = Studio.objects.filter(created_by=designer).first()
                bank_account_number = ''
                bank_ifsc_code = ''
                bank_account_holder_name = ''
                
                if studio:
                    business_details = StudioBusinessDetails.objects.filter(studio=studio).first()
                    if business_details:
                        bank_account_number = business_details.bank_account_number or ''
                        bank_ifsc_code = business_details.bank_ifsc_code or ''
                        bank_account_holder_name = business_details.bank_account_holder_name or ''
                
                # Mask account number for security (show last 4 digits)
                if bank_account_number:
                    if len(bank_account_number) > 4:
                        masked_account = '****' + bank_account_number[-4:]
                    else:
                        masked_account = '****'
                else:
                    masked_account = 'N/A'
                
                # Get designer contact info
                designer_email = designer.email if hasattr(designer, 'email') else ''
                designer_phone = ''
                try:
                    from Authentication.models import MobileNumber
                    mobile_numbers = get_related(designer, 'User:MobileNumber', MobileNumber)
                    if mobile_numbers.exists():
                        mobile = mobile_numbers.first()
                        designer_phone = str(mobile.mobile_number) if hasattr(mobile, 'mobile_number') else ''
                except:
                    pass
                
                # Format period
                period_str = f"{settlement.settlement_period_start.strftime('%b %d, %Y')} - {settlement.settlement_period_end.strftime('%b %d, %Y')}"
                
                settlement_data.append({
                    'Settlement ID': settlement.id,
                    'Designer ID': designer.id,
                    'Designer Name': f"{designer.first_name} {designer.last_name}".strip() or designer.username,
                    'Email': designer_email,
                    'Phone': designer_phone,
                    'Account Holder Name': bank_account_holder_name or 'N/A',
                    'Account Number': masked_account,
                    'IFSC Code': bank_ifsc_code or 'N/A',
                    'Settlement Period': period_str,
                    'Settlement Amount (₹)': float(settlement.settlement_amount),
                    'Status': settlement.get_status_display(),
                    'Opted In': 'Yes' if settlement.opted_in else 'No',
                    'Opted In At': settlement.opted_in_at.strftime('%Y-%m-%d %H:%M:%S') if settlement.opted_in_at else 'N/A',
                    'Settlement Date': settlement.settlement_date.strftime('%Y-%m-%d') if settlement.settlement_date else 'N/A',
                    'Failure Reason': settlement.failure_reason or 'N/A',
                    'Created At': settlement.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                })
                
                total_amount += Decimal(str(settlement.settlement_amount))
                
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error processing settlement {settlement.id}: {str(e)}", exc_info=True)
                continue
        
        # Generate filename
        current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'settlement_sheet_{current_date}.{file_format}'
        
        # Generate file
        if file_format == 'csv':
            return _generate_csv_response(settlement_data, filename, total_amount)
        else:
            return _generate_excel_response(settlement_data, filename, total_amount)
    
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error generating settlement sheet: {str(e)}", exc_info=True)
        return Response({
            'error': f'Failed to generate settlement sheet: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _generate_csv_response(settlement_data, filename, total_amount):
    """Generate CSV response"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    if not settlement_data:
        writer = csv.writer(response)
        writer.writerow(['No settlements found'])
        return response
    
    # Write header
    writer = csv.DictWriter(response, fieldnames=settlement_data[0].keys())
    writer.writeheader()
    
    # Write data
    for row in settlement_data:
        writer.writerow(row)
    
    # Write summary
    writer.writerow({})
    writer.writerow({'Settlement ID': 'SUMMARY', 'Designer Name': f'Total Settlements: {len(settlement_data)}', 'Settlement Amount (₹)': f'Total Amount: ₹{float(total_amount):.2f}'})
    
    return response


def _generate_excel_response(settlement_data, filename, total_amount):
    """Generate Excel response using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Settlement Sheet"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if not settlement_data:
            ws['A1'] = 'No settlements found'
            ws['A1'].font = Font(bold=True)
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
        
        # Write header
        headers = list(settlement_data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_num, row_data in enumerate(settlement_data, 2):
            for col_num, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Write summary row
        summary_row = len(settlement_data) + 3
        ws.cell(row=summary_row, column=1, value='SUMMARY').font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=f'Total Settlements: {len(settlement_data)}').font = Font(bold=True)
        
        # Find the amount column
        amount_col = headers.index('Settlement Amount (₹)') + 1
        ws.cell(row=summary_row, column=amount_col, value=f'Total Amount: ₹{float(total_amount):.2f}').font = Font(bold=True)
        
        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(header))
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        return _generate_csv_response(settlement_data, filename.replace('.xlsx', '.csv'), total_amount)


# ==================== ADMIN - SETTLEMENT STATUS UPDATE ====================

@swagger_auto_schema(
    method='put',
    operation_summary='Update Settlement Status',
    operation_description='Update settlement request status after manual payout processing (Admin only).',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'status': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='New settlement status',
                enum=['processing', 'completed', 'failed'],
                example='completed'
            ),
            'failure_reason': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Failure reason (required if status is failed)',
                example='Bank account details incorrect'
            ),
            'manual_reference_id': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Manual transaction reference ID (e.g., UTR number, transaction ID)',
                example='UTR123456789'
            ),
            'admin_notes': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Admin notes about the settlement',
                example='Payout processed via NEFT on 2024-02-06'
            )
        },
        required=['status']
    ),
    responses={
        200: openapi.Response(description='Settlement status updated successfully'),
        400: openapi.Response(description='Invalid data'),
        404: openapi.Response(description='Settlement not found'),
        403: openapi.Response(description='Access denied - Admin privileges required')
    },
    tags=['Wallet Admin']
)
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@admin_required()
def update_settlement_status(request, settlement_id):
    """
    Update settlement request status after manual payout processing.
    Admin-only endpoint to mark settlements as completed or failed.
    """
    try:
        settlement = SettlementRequest.objects.get(id=settlement_id)
    except SettlementRequest.DoesNotExist:
        return Response({
            'error': 'Settlement request not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    new_status = request.data.get('status')
    failure_reason = request.data.get('failure_reason', '')
    manual_reference_id = request.data.get('manual_reference_id', '')
    admin_notes = request.data.get('admin_notes', '')
    
    # Validate status
    valid_statuses = ['processing', 'completed', 'failed']
    if new_status not in valid_statuses:
        return Response({
            'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Validate failure reason if status is failed
    if new_status == 'failed' and not failure_reason:
        return Response({
            'error': 'failure_reason is required when status is failed'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    old_status = settlement.status
    
    # Update settlement
    settlement.status = new_status
    
    if new_status == 'failed':
        settlement.failure_reason = failure_reason
    elif new_status == 'completed':
        # Clear failure reason if completing
        if settlement.failure_reason:
            settlement.failure_reason = None
    
    # Store manual reference ID in razorpay_transfer_id field (repurposed for manual transactions)
    if manual_reference_id:
        settlement.razorpay_transfer_id = manual_reference_id
    
    settlement.save()
    
    # Log activity
    try:
        from CoreAdmin.models import AdminActivityLog
        AdminActivityLog.log_activity(
            user=request.user,
            activity_type='settlement_management',
            description=f'Updated settlement {settlement_id} status from {old_status} to {new_status}',
            request=request,
            metadata={
                'settlement_id': settlement_id,
                'old_status': old_status,
                'new_status': new_status,
                'manual_reference_id': manual_reference_id,
                'admin_notes': admin_notes,
                'designer_id': settlement.designer_id
            }
        )
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Failed to log activity: {str(e)}', exc_info=True)
    
    return Response({
        'message': 'Settlement status updated successfully',
        'settlement': {
            'id': settlement.id,
            'designer_id': settlement.designer_id,
            'settlement_amount': float(settlement.settlement_amount),
            'old_status': old_status,
            'new_status': settlement.status,
            'settlement_period_start': settlement.settlement_period_start.isoformat(),
            'settlement_period_end': settlement.settlement_period_end.isoformat(),
            'manual_reference_id': settlement.razorpay_transfer_id,
            'failure_reason': settlement.failure_reason
        }
    })


@swagger_auto_schema(
    method='post',
    operation_summary='Bulk Update Settlement Status',
    operation_description='Bulk update multiple settlement request statuses after manual payout processing (Admin only).',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'settlement_ids': openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(type=openapi.TYPE_INTEGER),
                description='List of settlement IDs to update',
                example=[1, 2, 3]
            ),
            'status': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='New settlement status for all settlements',
                enum=['processing', 'completed', 'failed'],
                example='completed'
            ),
            'failure_reason': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Failure reason (applied to all if status is failed)',
                example='Bank account details incorrect'
            ),
            'manual_reference_ids': openapi.Schema(
                type=openapi.TYPE_OBJECT,
                description='Mapping of settlement_id to manual_reference_id (optional)',
                additional_properties=openapi.Schema(type=openapi.TYPE_STRING),
                example={'1': 'UTR123456789', '2': 'UTR987654321'}
            ),
            'admin_notes': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Admin notes about the bulk update',
                example='All payouts processed via NEFT on 2024-02-06'
            )
        },
        required=['settlement_ids', 'status']
    ),
    responses={
        200: openapi.Response(description='Settlements updated successfully'),
        400: openapi.Response(description='Invalid data'),
        403: openapi.Response(description='Access denied - Admin privileges required')
    },
    tags=['Wallet Admin']
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@admin_required()
def bulk_update_settlement_status(request):
    """
    Bulk update multiple settlement request statuses after manual payout processing.
    Admin-only endpoint to mark multiple settlements as completed or failed at once.
    """
    settlement_ids = request.data.get('settlement_ids', [])
    new_status = request.data.get('status')
    failure_reason = request.data.get('failure_reason', '')
    manual_reference_ids = request.data.get('manual_reference_ids', {})
    admin_notes = request.data.get('admin_notes', '')
    
    if not settlement_ids:
        return Response({
            'error': 'settlement_ids is required and cannot be empty'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if not isinstance(settlement_ids, list):
        return Response({
            'error': 'settlement_ids must be a list'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Validate status
    valid_statuses = ['processing', 'completed', 'failed']
    if new_status not in valid_statuses:
        return Response({
            'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Validate failure reason if status is failed
    if new_status == 'failed' and not failure_reason:
        return Response({
            'error': 'failure_reason is required when status is failed'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    updated_count = 0
    failed_updates = []
    
    for settlement_id in settlement_ids:
        try:
            settlement = SettlementRequest.objects.get(id=settlement_id)
            
            old_status = settlement.status
            settlement.status = new_status
            
            if new_status == 'failed':
                settlement.failure_reason = failure_reason
            elif new_status == 'completed':
                # Clear failure reason if completing
                if settlement.failure_reason:
                    settlement.failure_reason = None
            
            # Store manual reference ID if provided
            if str(settlement_id) in manual_reference_ids:
                settlement.razorpay_transfer_id = manual_reference_ids[str(settlement_id)]
            
            settlement.save()
            updated_count += 1
            
        except SettlementRequest.DoesNotExist:
            failed_updates.append(f"Settlement {settlement_id} not found")
        except Exception as e:
            failed_updates.append(f"Error updating settlement {settlement_id}: {str(e)}")
    
    # Log activity
    try:
        from CoreAdmin.models import AdminActivityLog
        AdminActivityLog.log_activity(
            user=request.user,
            activity_type='settlement_management',
            description=f'Bulk updated {updated_count} settlements to {new_status}',
            request=request,
            metadata={
                'settlement_ids': settlement_ids,
                'old_status': 'various',
                'new_status': new_status,
                'updated_count': updated_count,
                'failed_updates': failed_updates,
                'admin_notes': admin_notes
            }
        )
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Failed to log activity: {str(e)}', exc_info=True)
    
    return Response({
        'message': f'Successfully updated {updated_count} settlements',
        'updated_count': updated_count,
        'failed_updates': failed_updates,
        'status': new_status
    })


@swagger_auto_schema(
    method='get',
    operation_summary='List Settlements',
    operation_description='Get list of settlement requests with filtering (Admin only).',
    manual_parameters=[
        openapi.Parameter(
            'status',
            openapi.IN_QUERY,
            description='Filter by settlement status',
            type=openapi.TYPE_STRING,
            enum=['pending', 'opted_in', 'processing', 'completed', 'failed', 'expired']
        ),
        openapi.Parameter(
            'period_start',
            openapi.IN_QUERY,
            description='Filter by settlement period start date (YYYY-MM-DD)',
            type=openapi.TYPE_STRING
        ),
        openapi.Parameter(
            'settlement_date',
            openapi.IN_QUERY,
            description='Filter by settlement date (YYYY-MM-DD)',
            type=openapi.TYPE_STRING
        ),
        openapi.Parameter(
            'designer_id',
            openapi.IN_QUERY,
            description='Filter by designer ID',
            type=openapi.TYPE_INTEGER
        ),
        openapi.Parameter(
            'page',
            openapi.IN_QUERY,
            description='Page number',
            type=openapi.TYPE_INTEGER
        ),
        openapi.Parameter(
            'page_size',
            openapi.IN_QUERY,
            description='Number of items per page',
            type=openapi.TYPE_INTEGER
        )
    ],
    responses={
        200: openapi.Response(description='Settlements retrieved successfully'),
        403: openapi.Response(description='Access denied - Admin privileges required')
    },
    tags=['Wallet Admin']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
@admin_required()
def list_settlements(request):
    """
    Get list of settlement requests with filtering.
    Admin-only endpoint to view and manage settlements.
    """
    status_filter = request.GET.get('status')
    period_start_str = request.GET.get('period_start')
    settlement_date_str = request.GET.get('settlement_date')
    designer_id = request.GET.get('designer_id')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    
    queryset = SettlementRequest.objects.all()
    
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    
    if period_start_str:
        try:
            period_start = datetime.strptime(period_start_str, '%Y-%m-%d').date()
            queryset = queryset.filter(settlement_period_start=period_start)
        except ValueError:
            return Response({
                'error': 'Invalid period_start format. Use YYYY-MM-DD'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    if settlement_date_str:
        try:
            settlement_date = datetime.strptime(settlement_date_str, '%Y-%m-%d').date()
            queryset = queryset.filter(settlement_date=settlement_date)
        except ValueError:
            return Response({
                'error': 'Invalid settlement_date format. Use YYYY-MM-DD'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    if designer_id:
        try:
            queryset = queryset.filter(designer_id=int(designer_id))
        except ValueError:
            return Response({
                'error': 'Invalid designer_id'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    # Order by settlement period and designer ID
    queryset = queryset.order_by('-settlement_period_start', 'designer_id')
    
    # Pagination
    total_count = queryset.count()
    start = (page - 1) * page_size
    end = start + page_size
    settlements_page = queryset[start:end]
    
    # Serialize settlements
    settlements_data = []
    for settlement in settlements_page:
        settlements_data.append({
            'id': settlement.id,
            'designer_id': settlement.designer_id,
            'settlement_period_start': settlement.settlement_period_start.isoformat(),
            'settlement_period_end': settlement.settlement_period_end.isoformat(),
            'wallet_balance_at_period_end': float(settlement.wallet_balance_at_period_end),
            'settlement_amount': float(settlement.settlement_amount),
            'status': settlement.status,
            'opted_in': settlement.opted_in,
            'opted_in_at': settlement.opted_in_at.isoformat() if settlement.opted_in_at else None,
            'settlement_date': settlement.settlement_date.isoformat() if settlement.settlement_date else None,
            'manual_reference_id': settlement.razorpay_transfer_id,
            'failure_reason': settlement.failure_reason,
            'created_at': settlement.created_at.isoformat(),
            'updated_at': settlement.updated_at.isoformat()
        })
    
    return Response({
        'settlements': settlements_data,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_count': total_count,
            'total_pages': (total_count + page_size - 1) // page_size
        }
    })
